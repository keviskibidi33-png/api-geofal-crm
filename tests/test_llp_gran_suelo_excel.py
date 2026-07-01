import io
import sys
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.gran_suelo.excel import TEMPLATE_PATH as GRAN_TEMPLATE_PATH
from app.modules.gran_suelo.excel import generate_gran_suelo_excel
from app.modules.gran_suelo.schemas import GranSueloRequest
from app.modules.llp.excel import TEMPLATE_PATH as LLP_TEMPLATE_PATH
from app.modules.llp.excel import generate_llp_excel
from app.modules.llp.schemas import LLPRequest

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _build_llp_payload() -> LLPRequest:
    return LLPRequest(
        muestra="123-SU-26",
        numero_ot="4567-26",
        fecha_ensayo="2026/05/07",
        realizado_por="OPERADOR",
        metodo_ensayo_limite_liquido="MULTIPUNTO",
        herramienta_ranurado_limite_liquido="METAL",
        dispositivo_limite_liquido="MANUAL",
        metodo_laminacion_limite_plastico="MANUAL",
        contenido_humedad_muestra_inicial_pct=35.4,
        proceso_seleccion_muestra="TRAZA",
        metodo_preparacion_muestra="SECADO AL AIRE",
        tamano_maximo_visual_in='3/4"',
        porcentaje_retenido_tamiz_40_pct=5.7,
        forma_particula="SUBREDONDEADA",
        tipo_muestra="SUELO FINO",
        condicion_muestra="ALTERADO",
        puntos=[
            {
                "recipiente_numero": "A1",
                "numero_golpes": 20,
                "masa_recipiente_suelo_humedo": 105,
                "masa_recipiente_suelo_seco_1": 75,
                "masa_recipiente": 25,
            },
            {
                "recipiente_numero": "A2",
                "numero_golpes": 18,
                "masa_recipiente_suelo_humedo": 110,
                "masa_recipiente_suelo_seco_1": 80,
                "masa_recipiente": 25,
            },
            {
                "recipiente_numero": "A3",
                "numero_golpes": 15,
                "masa_recipiente_suelo_humedo": 120,
                "masa_recipiente_suelo_seco_1": 90,
                "masa_recipiente": 30,
            },
            {
                "recipiente_numero": "P1",
                "masa_recipiente_suelo_humedo": 60,
                "masa_recipiente_suelo_seco_1": 50,
                "masa_recipiente": 20,
            },
            {
                "recipiente_numero": "P2",
                "masa_recipiente_suelo_humedo": 62,
                "masa_recipiente_suelo_seco_1": 50,
                "masa_recipiente": 20,
            },
        ],
        balanza_001g_codigo="BAL-001",
        horno_110_codigo="HOR-110",
        copa_casagrande_codigo="CAS-001",
        ranurador_codigo="RAN-0107",
        revisado_por="REVISOR R",
        revisado_fecha="2026/05/08",
        aprobado_por="APROBADOR R",
        aprobado_fecha="2026/05/09",
    )


def _build_gran_payload() -> GranSueloRequest:
    return GranSueloRequest(
        muestra="123-SU-26",
        numero_ot="4567-26",
        fecha_ensayo="2026/05/07",
        realizado_por="OPERADOR",
        descripcion_turbo_organico="ORG",
        metodo_prueba="A",
        tamizado_tipo="GLOBAL",
        metodo_muestreo="SECADO AL AIRE",
        tipo_muestra="SUELO",
        condicion_muestra="ALTERADO",
        tamano_maximo_particula_in="3/4\"",
        forma_particula="SUBREDONDEADA",
        masa_seca_porcion_gruesa_cp_md_g=100,
        masa_humeda_porcion_fina_fp_mm_g=50,
        masa_seca_porcion_fina_fp_md_g=45,
        masa_seca_muestra_s_md_g=145,
        masa_seca_global_g=1000,
        subespecie_masa_humeda_g=200,
        subespecie_masa_seca_g=180,
        contenido_agua_wfp_pct=10,
        masa_porcion_gruesa_lavada_cpwmd_g=12,
        masa_retenida_plato_cpmrpan_g=4,
        perdida_cpl_pct=8,
        masa_subespecimen_lavado_fina_g=20,
        clasificacion_visual_simbolo="SP",
        clasificacion_visual_nombre="ARENA",
        excluyo_material="SI",
        excluyo_material_descripcion="desc",
        problema_muestra="NO",
        problema_descripcion="-",
        proceso_dispersion="MANUAL",
        masa_retenida_primer_tamiz_g=1,
        masa_retenida_tamiz_g=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15],
    )


def test_llp_template_replaced_and_generation_keeps_links():
    workbook = load_workbook(LLP_TEMPLATE_PATH, data_only=False)

    assert workbook["FORMATO"]["A8"].value == "FORMATO N° F-LEM-P-SU-23.01"
    assert [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"] == [
        "FORMATO",
        "INFORME LIMITE",
        "DATOS",
        "Incertidumbre",
        "Balanza",
        "Precision",
    ]

    generated = load_workbook(io.BytesIO(generate_llp_excel(_build_llp_payload())), data_only=False)

    assert generated["INFORME NTP"]["C11"].value == "123-SU-26"
    assert generated["INFORME NTP"]["E11"].value == "4567-26"
    assert generated["INFORME NTP"]["H11"].value == "2026/05/07"
    assert generated["INFORME NTP"]["J11"].value == "OPERADOR"
    assert generated["FORMATO"]["C11"].value == "123-SU-26"
    assert generated["FORMATO"]["E11"].value == "4567-26"
    assert generated["FORMATO"]["H11"].value == "2026/05/07"
    assert generated["FORMATO"]["J11"].value == "OPERADOR"
    assert generated["FORMATO"]["J17"].value == "MULTIPUNTO"
    assert generated["FORMATO"]["J21"].value == 35.4
    assert generated["FORMATO"]["J22"].value == "TRAZA"
    assert generated["FORMATO"]["J23"].value == "SECADO AL AIRE"
    assert generated["FORMATO"]["J29"].value == "SUELO FINO"
    assert generated["FORMATO"]["J30"].value == "ALTERADO"
    assert generated["FORMATO"]["J31"].value == '3/4"'
    assert generated["FORMATO"]["J32"].value == 5.7
    assert generated["FORMATO"]["J33"].value == "SUBREDONDEADA"
    assert generated["FORMATO"]["D48"].value == "BAL-001"
    assert generated["FORMATO"]["D49"].value == "HOR-110"
    assert generated["FORMATO"]["D50"].value == "CAS-001"
    assert generated["FORMATO"]["D51"].value == "INS-0107"
    assert generated["FORMATO"]["A8"].value == "FORMATO N° F-LEM-P-SU-23.01"
    assert generated["INFORME NTP"]["A8"].value == "=+K13"
    assert generated["DATOS"]["C15"].value == "=+FORMATO!G37"
    assert generated["curva"]["C2"].value == "=DATOS!D16"
    assert generated["INFORME LIMITE"]["C5"].value == "=+FORMATO!Q2"
    assert generated["INFORME LIMITE"]["C6"].value == "=+FORMATO!Q3"
    assert generated["INFORME LIMITE"]["V23"].value == "=+DATOS!D16"
    assert generated["INFORME LIMITE"]["K52"].value == "=+FORMATO!J17"

    with zipfile.ZipFile(io.BytesIO(generate_llp_excel(_build_llp_payload())), "r") as archive:
        names = set(archive.namelist())
        assert "xl/calcChain.xml" not in names
        assert all(not name.startswith("xl/externalLinks/") for name in names)
        assert "xl/workbook.xml" in names

        workbook_root = etree.fromstring(archive.read("xl/workbook.xml"))
        ns = {"m": NS_MAIN}
        workbook_pr = workbook_root.find("m:workbookPr", ns)
        assert workbook_pr is not None
        assert workbook_pr.get("updateLinks") == "never"
        assert workbook_root.find("m:externalReferences", ns) is None
        defined_names = workbook_root.find("m:definedNames", ns)
        assert defined_names is not None
        for defined_name in defined_names.findall("m:definedName", ns):
            text = (defined_name.text or "").strip()
            assert "#REF!" not in text
            assert "[" not in text
        calc_pr = workbook_root.find("m:calcPr", ns)
        assert calc_pr is not None
        assert calc_pr.get("fullCalcOnLoad") == "1"
        assert calc_pr.get("forceFullCalc") == "1"

        rels_root = etree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = [rel.get("Target", "") for rel in rels_root]
        assert all(not target.endswith("calcChain.xml") for target in rel_targets)
        assert all("externalLinks/" not in target for target in rel_targets)

        content_types_root = etree.fromstring(archive.read("[Content_Types].xml"))
        overrides = [override.get("PartName", "") for override in content_types_root]
        assert "/xl/calcChain.xml" not in overrides
        assert all(not part_name.startswith("/xl/externalLinks/") for part_name in overrides)

        sheet1 = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        ns = {"m": NS_MAIN}
        assert sheet1.xpath(".//m:c[@r='P32']/m:f[@t='shared']", namespaces=ns)
        assert sheet1.xpath(".//m:c[@r='Q33']/m:f[@t='shared']", namespaces=ns)

        drawing2 = etree.fromstring(archive.read("xl/drawings/drawing2.xml"))
        drawing_ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        footer_texts = []
        for anchor in drawing2.findall(".//xdr:twoCellAnchor", drawing_ns):
            footer_texts.extend(
                (node.text or "").strip()
                for node in anchor.findall(".//a:t", drawing_ns)
                if (node.text or "").strip()
            )
        assert "Revisado:" in footer_texts
        assert "REVISOR R" in footer_texts
        assert "Fecha: 2026/05/08" in footer_texts
        assert "Aprobado:" in footer_texts
        assert "APROBADOR R" in footer_texts
        assert "Fecha: 2026/05/09" in footer_texts


def test_gran_suelo_template_replaced_and_generation_forces_recalc():
    workbook = load_workbook(GRAN_TEMPLATE_PATH, data_only=False)

    assert workbook["FORMATO"]["A8"].value == "FORMATO N° F-LEM-P-SU-24.01"
    assert [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"] == [
        "FORMATO",
        "A.Granul",
        "Incertidumbre",
        "Balanza",
    ]

    generated = load_workbook(io.BytesIO(generate_gran_suelo_excel(_build_gran_payload())), data_only=False)

    assert generated["FORMATO"]["D11"].value == "123-SU-26"
    assert generated["FORMATO"]["F11"].value == "4567-26"
    assert generated["FORMATO"]["H11"].value == "2026/05/07"
    assert generated["FORMATO"]["J11"].value == "OPERADOR"
    assert generated["FORMATO"]["A8"].value == "FORMATO N° F-LEM-P-SU-24.01"
    assert generated["FORMATO"]["B17"].value == "X"
    assert generated["FORMATO"]["B22"].value == "X"
    assert generated["FORMATO"]["C26"].value == "X"
    assert generated["FORMATO"]["E38"].value == "Manual X"
    assert generated["FORMATO"]["E42"].fill.fill_type is None
    assert generated["Sucs "]["L2"].value == "=+A.Granul!R2"
    assert generated["Aashto"]["C20"].value == "=+'LL-LP'!H35"

    with zipfile.ZipFile(io.BytesIO(generate_gran_suelo_excel(_build_gran_payload())), "r") as archive:
        names = set(archive.namelist())
        assert "xl/calcChain.xml" not in names
        assert all(not name.startswith("xl/externalLinks/") for name in names)
        assert "xl/workbook.xml" in names

        workbook_root = etree.fromstring(archive.read("xl/workbook.xml"))
        ns = {"m": NS_MAIN}
        workbook_pr = workbook_root.find("m:workbookPr", ns)
        assert workbook_pr is not None
        assert workbook_pr.get("updateLinks") == "never"
        assert workbook_root.find("m:externalReferences", ns) is None
        defined_names = workbook_root.find("m:definedNames", ns)
        assert defined_names is not None
        for defined_name in defined_names.findall("m:definedName", ns):
            text = (defined_name.text or "").strip()
            assert "#REF!" not in text
            assert "[" not in text
        calc_pr = workbook_root.find("m:calcPr", ns)
        assert calc_pr is not None
        assert calc_pr.get("calcMode") == "auto"
        assert calc_pr.get("fullCalcOnLoad") == "1"
        assert calc_pr.get("forceFullCalc") == "1"

        rels_root = etree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = [rel.get("Target", "") for rel in rels_root]
        assert all(not target.endswith("calcChain.xml") for target in rel_targets)
        assert all("externalLinks/" not in target for target in rel_targets)

        content_types_root = etree.fromstring(archive.read("[Content_Types].xml"))
        overrides = [override.get("PartName", "") for override in content_types_root]
        assert "/xl/calcChain.xml" not in overrides
        assert all(not part_name.startswith("/xl/externalLinks/") for part_name in overrides)
