import io
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.proctor.excel import TEMPLATE_PATH, generate_proctor_excel
from app.modules.proctor.schemas import ProctorRequest


EXPECTED_VISIBLE_SHEETS = [
    "formato",
    "Proctor",
    "DATOS",
    "Incertidumbre",
    "Balanza",
    "precision",
]

EXPECTED_HIDDEN_SHEETS = [
    "Correccion",
    "A.Granul",
    "LL-LP",
    "GR-Esp",
    "Sucs ",
    "Aashto",
    "Hoja1",
]


def _build_payload() -> ProctorRequest:
    return ProctorRequest(
        muestra="123-SU-26",
        numero_ot="4567-26",
        fecha_ensayo="2026/05/07",
        realizado_por="OPERADOR",
    )


def test_proctor_template_is_1901_and_has_expected_sheet_stack():
    workbook = load_workbook(TEMPLATE_PATH, data_only=False)

    assert workbook.sheetnames == EXPECTED_VISIBLE_SHEETS + EXPECTED_HIDDEN_SHEETS
    assert workbook["formato"]["A6"].value == "FORMATO N° F-LEM-P-SU-19.01"
    assert [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"] == EXPECTED_VISIBLE_SHEETS
    assert [ws.title for ws in workbook.worksheets if ws.sheet_state == "hidden"] == EXPECTED_HIDDEN_SHEETS
    assert len([ws for ws in workbook.worksheets if ws.sheet_state == "visible"]) == 6
    assert len([ws for ws in workbook.worksheets if ws.sheet_state == "hidden"]) == 7

    with zipfile.ZipFile(TEMPLATE_PATH, "r") as workbook_zip:
        names = set(workbook_zip.namelist())
        assert "xl/calcChain.xml" in names
        assert "xl/externalLinks/externalLink7.xml" in names
        assert "xl/workbook.xml" in names


def test_generate_proctor_excel_preserves_formulas_and_sheet_connections():
    payload = _build_payload()
    generated_bytes = generate_proctor_excel(payload)
    workbook = load_workbook(io.BytesIO(generated_bytes), data_only=False)

    assert workbook.sheetnames == EXPECTED_VISIBLE_SHEETS + EXPECTED_HIDDEN_SHEETS
    assert [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"] == EXPECTED_VISIBLE_SHEETS
    assert [ws.title for ws in workbook.worksheets if ws.sheet_state == "hidden"] == EXPECTED_HIDDEN_SHEETS

    formato = workbook["formato"]
    assert formato["B9"].value == payload.muestra
    assert formato["C9"].value == payload.numero_ot
    assert formato["F9"].value == payload.fecha_ensayo
    assert formato["H9"].value == payload.realizado_por

    # Main-sheet formulas must remain intact so linked sheets continue working.
    assert workbook["Proctor"]["C9"].value == "=+formato!L2"
    assert workbook["Correccion"]["C10"].value == "=+Proctor!E9"
    assert workbook["Sucs "]["L2"].value == "=+A.Granul!O4"
    assert workbook["Aashto"]["C20"].value == "=+'LL-LP'!H35"

    with zipfile.ZipFile(io.BytesIO(generated_bytes), "r") as workbook_zip:
        names = set(workbook_zip.namelist())
        assert "xl/calcChain.xml" in names
        assert "xl/externalLinks/externalLink7.xml" in names
