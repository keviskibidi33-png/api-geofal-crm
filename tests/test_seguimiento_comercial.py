from __future__ import annotations

import os
import sys
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Force SQLite in-memory database for testing to avoid connection errors with the external Postgres DB
os.environ["QUOTES_DATABASE_URL"] = "sqlite:///:memory:"

import app.database as app_database

from app.database import Base, get_db_session
from app.modules.seguimiento_cliente_comercial.models import SeguimientoClienteComercial
from app.modules.seguimiento_cliente_comercial.service import SeguimientoClienteComercialService

from sqlalchemy.pool import StaticPool

# Setup test SQLite database
test_engine = create_engine(
    "sqlite:///:memory:",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=test_engine)

# Patch the shared database module before importing the FastAPI app so that
# app.main binds to the in-memory SQLite engine instead of the real Postgres
# connection used in production.
app_database.engine = test_engine
app_database.SessionLocal = TestingSessionLocal

from app.main import app

class TestSeguimientoComercialEndpoints(unittest.TestCase):
    def setUp(self):
        # Create tables
        Base.metadata.create_all(bind=test_engine)
        self.db = TestingSessionLocal()
        
        # Override get_db_session dependency in FastAPI app
        def override_get_db_session():
            db = TestingSessionLocal()
            try:
                yield db
                db.commit()
            except Exception:
                db.rollback()
                raise
            finally:
                db.close()
                
        app.dependency_overrides[get_db_session] = override_get_db_session
        self.client = TestClient(app)
        
        # Seed a test record
        self.test_record = SeguimientoClienteComercial(
            no=1,
            fecha_contacto=date(2026, 5, 20),
            persona_contacto="Test Contact",
            numero_celular="987654321",
            email="test@example.com",
            razon_social="Test Company S.A.C.",
            ruc="20123456789",
            asesor="Silvia Peralta",
            contacto="WHATSAPP",
            rubro="LABORATORIO",
            estado_cliente="SE SOLICITÓ INFORMACIÓN",
            servicio_solicitado="Ensayos de concreto",
            fecha_ultimo_contacto=date(2026, 5, 21),
            comentarios_asistente="Test comments",
            numero_cotizacion="COT-1234",
            estado_seguimiento="Enviado"
        )
        self.db.add(self.test_record)
        self.db.commit()
        self.db.refresh(self.test_record)

    def tearDown(self):
        self.db.close()
        Base.metadata.drop_all(bind=test_engine)
        app.dependency_overrides.clear()

    def test_list_records(self):
        # List all records
        response = self.client.get("/api/seguimiento-comercial")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["total"], 1)
        self.assertEqual(data["items"][0]["persona_contacto"], "Test Contact")
        self.assertEqual(data["items"][0]["ruc"], "20123456789")

    def test_list_records_search_filter(self):
        # Search by RUC
        response = self.client.get("/api/seguimiento-comercial?search=20123456789")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total"], 1)

        # Search by non-existent term
        response = self.client.get("/api/seguimiento-comercial?search=invalid")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total"], 0)

    def test_get_catalogs(self):
        response = self.client.get("/api/seguimiento-comercial/catalogs")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("asesores", data)
        self.assertIn("contactos", data)
        self.assertIn("rubros", data)
        self.assertIn("estados", data)
        # Ensure our test record's values are present or merged
        self.assertIn("Silvia Peralta", data["asesores"])
        self.assertIn("WHATSAPP", data["contactos"])

    def test_get_catalogs_normalizes_legacy_advisors(self):
        # Insert a record with advisor = "SILVIA"
        legacy_record = SeguimientoClienteComercial(
            no=10,
            fecha_contacto=date(2026, 5, 20),
            persona_contacto="Legacy Contact",
            razon_social="Legacy Company",
            asesor="SILVIA"
        )
        self.db.add(legacy_record)
        self.db.commit()

        response = self.client.get("/api/seguimiento-comercial/catalogs")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("Silvia Peralta", data["asesores"])
        self.assertNotIn("SILVIA", data["asesores"])
        self.assertEqual(len(data["asesores"]), 3)

    def test_get_catalogs_excludes_removed_service_options(self):
        legacy_services = [
            "Morteros",
            "Extracción de Diamantina",
            "EMS – CIMENTACIÓN",
            "EMS – PAVIMENTACIÓN",
            "EMS – ALCANTARILLADO",
            "Estudios Geotécnicos",
        ]

        for index, servicio in enumerate(legacy_services, start=20):
            self.db.add(
                SeguimientoClienteComercial(
                    no=index,
                    fecha_contacto=date(2026, 5, 20),
                    persona_contacto=f"Legacy Service {index}",
                    razon_social="Legacy Service S.A.C.",
                    servicio_solicitado=servicio,
                )
            )

        self.db.commit()

        response = self.client.get("/api/seguimiento-comercial/catalogs")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        for servicio in legacy_services:
            self.assertNotIn(servicio, data["servicios"])
        self.assertIn("Ensayos de Laboratorio", data["servicios"])
        self.assertIn("Estudios de Suelos", data["servicios"])
        self.assertIn("Densidades", data["servicios"])
        self.assertIn("Probetas", data["servicios"])
        self.assertIn("Alquiler", data["servicios"])
        for internal_code in ["DEN", "PROB", "EMS", "ALQ", "ENS.V."]:
            self.assertNotIn(internal_code, data["servicios"])

    def test_create_record(self):
        payload = {
            "fecha_contacto": "2026-05-21",
            "persona_contacto": "New Contact",
            "numero_celular": "987000111",
            "email": "new@example.com",
            "razon_social": "New Company S.A.C.",
            "ruc": "20999888777",
            "asesor": "Juan Garcia",
            "contacto": "LLAMADA",
            "rubro": "INGENIERÍA",
            "estado_cliente": "COTIZACIÓN REALIZADA",
            "servicio_solicitado": "Ensayos de suelos",
            "comentarios_asistente": "New comments"
        }
        
        # Headers simulate an authenticated user
        headers = {"x-dev-user-id": "dev-user", "x-dev-user-name": "Test Operator"}
        response = self.client.post("/api/seguimiento-comercial", json=payload, headers=headers)
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["persona_contacto"], "New Contact")
        self.assertEqual(data["servicio_solicitado"], "Estudios de Suelos")
        self.assertEqual(data["no"], 2) # Auto-increment 'no'
        self.assertEqual(data["creado_por"], "Test Operator")

    def test_patch_record(self):
        payload = {
            "estado_cliente": "EN ESPERA DE INFORMACIÓN",
            "comentarios_asistente": "Patched comments"
        }
        headers = {"x-dev-user-id": "dev-user"}
        response = self.client.patch(f"/api/seguimiento-comercial/{self.test_record.id}", json=payload, headers=headers)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["estado_cliente"], "EN ESPERA DE INFORMACIÓN")
        self.assertEqual(data["comentarios_asistente"], "Patched comments")

    def test_import_normalizes_catalog_values(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SEG.CLIENTE"

        headers = [
            "N°", "FECHA CONTACTO", "PERSONA CONTACTO", "CELULAR", "EMAIL", "RAZÓN SOCIAL", "RUC",
            "ASESOR", "CONTACTO", "RUBRO", "ESTADO CLIENTE", "SERVICIO SOLICITADO",
            "F. ÚLTIMO CONTACTO", "N° COTIZACIÓN", "ESTADO SEGUIMIENTO",
        ]
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=4, column=index).value = header

        sheet.cell(row=5, column=1).value = 1
        sheet.cell(row=5, column=2).value = "2026-05-21"
        sheet.cell(row=5, column=3).value = "Import Contact"
        sheet.cell(row=5, column=4).value = "999999999"
        sheet.cell(row=5, column=5).value = "import@example.com"
        sheet.cell(row=5, column=6).value = "Import Company S.A.C."
        sheet.cell(row=5, column=7).value = "20111111111"
        sheet.cell(row=5, column=8).value = "silvia peralta"
        sheet.cell(row=5, column=9).value = "whatsapp"
        sheet.cell(row=5, column=10).value = "ingenieria"
        sheet.cell(row=5, column=11).value = "1. SOLICITUD INFORMACION"
        sheet.cell(row=5, column=12).value = "Ensayos de concreto"
        sheet.cell(row=5, column=13).value = "2026-05-22"
        sheet.cell(row=5, column=14).value = "COT-999"
        sheet.cell(row=5, column=15).value = "Enviado"

        payload_buffer = BytesIO()
        workbook.save(payload_buffer)

        inserted = SeguimientoClienteComercialService.importar_excel(self.db, payload_buffer.getvalue(), creado_por="Test Import")
        self.assertEqual(inserted, 1)

        imported = self.db.query(SeguimientoClienteComercial).order_by(SeguimientoClienteComercial.id.desc()).first()
        self.assertIsNotNone(imported)
        self.assertEqual(imported.asesor, "Silvia Peralta")
        self.assertEqual(imported.contacto, "WHATSAPP")
        self.assertEqual(imported.rubro, "INGENIERÍA")
        self.assertEqual(imported.estado_cliente, "SE SOLICITÓ INFORMACIÓN")

    def test_import_tsv_txt_normalizes_catalog_values(self):
        txt_content = "\n".join(
            [
                "N°\tFECHA CONTACTO\tPERSONA CONTACTO\tCELULAR\tEMAIL\tRAZÓN SOCIAL\tRUC\tASESOR\tCONTACTO\tRUBRO\tESTADO CLIENTE\tSERVICIO SOLICITADO\tF. ÚLTIMO CONTACTO\tN° COTIZACIÓN\tESTADO SEGUIMIENTO",
                "1\t23-feb.-26\tImport Contact\t999999999\timport@example.com\tImport Company S.A.C.\t20111111111\tsilvia peralta\twhatsapp\tingenieria\t4. SEG. COTIZACION\tEnsayos de concreto\t24-feb.-26\tCOT-999\tEnviado",
            ]
        )

        inserted = SeguimientoClienteComercialService.importar_excel(self.db, txt_content.encode("utf-8"), creado_por="Test Import TXT")
        self.assertEqual(inserted, 1)

        imported = self.db.query(SeguimientoClienteComercial).order_by(SeguimientoClienteComercial.id.desc()).first()
        self.assertIsNotNone(imported)
        self.assertEqual(imported.persona_contacto, "Import Contact")
        self.assertEqual(imported.asesor, "Silvia Peralta")
        self.assertEqual(imported.contacto, "WHATSAPP")
        self.assertEqual(imported.rubro, "INGENIERÍA")
        self.assertEqual(imported.estado_cliente, "COTIZACIÓN REALIZADA")
        self.assertEqual(imported.fecha_contacto.isoformat(), "2026-02-23")
        self.assertEqual(imported.fecha_ultimo_contacto.isoformat(), "2026-02-24")

    def test_import_csv_semicolon_and_misspelled_headers(self):
        csv_content = "\n".join([
            ";;;;;;SEGUIMIENTO DE CLIENTES;;;;;;;;",
            "",
            "No;FECHA CONTACTO;PERSONA CONTACTO;NMERO CELULAR;E-MAIL;RAZON SOCIAL;RUC;ASESOR;CONTACTO;RUBRO;ESTADO CLIENTE;SERVICIO SOLICITADO;FECHA LTIMO CONTACTO;N DE COTIZACIN;ESTADO DE SEGUIMIENTO",
            "1;23-feb.-26;Import Contact;999999999;import@example.com;Import Company;20111111111;SILVIA;whatsapp;ingenieria;4. SEG. COTIZACION;Ensayos de concreto;24-feb.-26;COT-999;Enviado"
        ])

        inserted = SeguimientoClienteComercialService.importar_excel(self.db, csv_content.encode("utf-8"), creado_por="Test Import CSV Semicolon")
        self.assertEqual(inserted, 1)

        imported = self.db.query(SeguimientoClienteComercial).order_by(SeguimientoClienteComercial.id.desc()).first()
        self.assertIsNotNone(imported)
        self.assertEqual(imported.persona_contacto, "Import Contact")
        self.assertEqual(imported.numero_celular, "999999999")
        self.assertEqual(imported.asesor, "Silvia Peralta")
        self.assertEqual(imported.contacto, "WHATSAPP")
        self.assertEqual(imported.rubro, "INGENIERÍA")
        self.assertEqual(imported.estado_cliente, "COTIZACIÓN REALIZADA")
        self.assertEqual(imported.fecha_contacto.isoformat(), "2026-02-23")
        self.assertEqual(imported.fecha_ultimo_contacto.isoformat(), "2026-02-24")

    def test_delete_record(self):
        headers = {"x-dev-user-id": "dev-user"}
        response = self.client.delete(f"/api/seguimiento-comercial/{self.test_record.id}", headers=headers)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)
        
        # Verify it is deleted
        db_check = self.db.query(SeguimientoClienteComercial).filter(SeguimientoClienteComercial.id == self.test_record.id).first()
        self.assertIsNone(db_check)

    def test_export_excel_with_seguimiento_template_numeric_cost(self):
        # Update test record with cost
        self.test_record.costo_cotiz_sin_igv = "S/. 17,759.00"
        self.test_record.categoria_servicio = "DEN"
        self.db.commit()

        template_path = str(PROJECT_ROOT / "app" / "templates" / "Seguimiento.xlsx")
        excel_bytes = SeguimientoClienteComercialService.exportar_excel(self.db, template_path)
        
        import openpyxl
        wb = openpyxl.load_workbook(excel_bytes, data_only=False)
        sheet = wb['COMERCIAL'] if 'COMERCIAL' in wb.sheetnames else wb.active

        # Find header row
        header_row = 8
        col_mapping = SeguimientoClienteComercialService._resolve_excel_column_mapping(sheet, header_row)
        
        cost_col = None
        for col_idx, field in col_mapping.items():
            if field == "costo_cotiz_sin_igv":
                cost_col = col_idx
                break

        self.assertIsNotNone(cost_col, "COSTO COTIZ SIN IGV column should be detected")
        self.assertEqual(cost_col, 12, "COSTO COTIZ SIN IGV should be at column 12 (L)")

        # Verify exported row 9
        row_9_cost_cell = sheet.cell(row=header_row + 1, column=cost_col)
        self.assertIn(type(row_9_cost_cell.value), (int, float), "Costo must be exported as a numeric value")
        self.assertEqual(row_9_cost_cell.value, 17759)
        self.assertEqual(row_9_cost_cell.number_format, '#,##0.00')

    def test_import_excel_14_column_seguimiento_template(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "COMERCIAL"

        headers = [
            "N°", "FECHA CONTACTO", "PERSONA CONTACTO", "CELULAR", "ASESOR COMENTARIO",
            "EMPRESA", "F.ULTIMO CONTACTO", "RUBRO", "ESTADO CLIENTE", "SERVICIO SOLICITADO",
            "N° COTIZACION", "COSTO COTIZ SIN IGV", "ESTADO SEGUIMIENTO", "CATEOGRIA CLIENTE"
        ]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=8, column=col_idx).value = header

        sheet.cell(row=9, column=1).value = 1
        sheet.cell(row=9, column=2).value = "2026-07-01"
        sheet.cell(row=9, column=3).value = "Juan Perez"
        sheet.cell(row=9, column=4).value = "987654321"
        sheet.cell(row=9, column=5).value = "Comentario de asesor"
        sheet.cell(row=9, column=6).value = "Constructora ABC SAC"
        sheet.cell(row=9, column=7).value = "2026-07-05"
        sheet.cell(row=9, column=8).value = "EDIFICACIONES"
        sheet.cell(row=9, column=9).value = "COTIZACION ENVIADA"
        sheet.cell(row=9, column=10).value = "ENSAYO DE DENSIDAD"
        sheet.cell(row=9, column=11).value = "COT-123-26"
        sheet.cell(row=9, column=12).value = 17759.0
        sheet.cell(row=9, column=13).value = "VENTA"
        sheet.cell(row=9, column=14).value = "DEN"

        payload_buffer = BytesIO()
        workbook.save(payload_buffer)

        inserted = SeguimientoClienteComercialService.importar_excel(self.db, payload_buffer.getvalue(), creado_por="Test 14-Col Import")
        self.assertEqual(inserted, 1)

        imported = self.db.query(SeguimientoClienteComercial).order_by(SeguimientoClienteComercial.id.desc()).first()
        self.assertIsNotNone(imported)
        self.assertEqual(imported.razon_social, "Constructora ABC SAC")
        self.assertEqual(imported.comentarios_asesor, "Comentario de asesor")
        self.assertEqual(imported.numero_cotizacion, "COT-123-26")
        self.assertEqual(imported.costo_cotiz_sin_igv, "17759")
        self.assertEqual(imported.categoria_servicio, "Categoría 1 (DEN)")


if __name__ == "__main__":
    unittest.main()
