import io
import unittest
import openpyxl
from fastapi.testclient import TestClient

from app.modules.recepcion.excel import ExcelLogic
from app.main import app

class TestRecepcionImportExcel(unittest.TestCase):
    def setUp(self):
        self.excel_logic = ExcelLogic()
        self.client = TestClient(app)

    def _create_dummy_excel(self, doc_code: str, cliente_name: str = "CLIENTE PRUEBA"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RECEPCION"
        
        # Row 1: Document code anchor
        ws.cell(row=1, column=1, value=f"CÓDIGO: {doc_code}")
        
        # Key-value anchors
        ws.cell(row=5, column=1, value="RECEPCIÓN N°:")
        ws.cell(row=5, column=2, value="REC-2026-99")
        
        ws.cell(row=6, column=1, value="CLIENTE:")
        ws.cell(row=6, column=2, value=cliente_name)
        
        ws.cell(row=7, column=1, value="PROYECTO:")
        ws.cell(row=7, column=2, value="PROYECTO DE PRUEBA DE RECEPCION")
        
        # Header for samples table
        ws.cell(row=21, column=1, value="N°")
        ws.cell(row=21, column=2, value="IDENTIFICACIÓN DE MUESTRA")
        ws.cell(row=21, column=3, value="FECHA DE MOLDEO")
        
        # Data rows
        ws.cell(row=23, column=1, value=1)
        ws.cell(row=23, column=2, value="MUESTRA-01")
        ws.cell(row=23, column=3, value="2026-08-10")

        ws.cell(row=24, column=1, value=2)
        ws.cell(row=24, column=2, value="MUESTRA-02")
        ws.cell(row=24, column=3, value="2026-08-11")

        # Footer
        ws.cell(row=30, column=1, value="NOTA:")
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    def test_parsear_recepcion_concreto(self):
        content = self._create_dummy_excel("F-LEM-P-01.02")
        data = self.excel_logic.parsear_recepcion(content)
        self.assertEqual(data.get("tipo_recepcion"), "CONCRETO")
        self.assertEqual(data.get("cliente"), "CLIENTE PRUEBA")
        self.assertEqual(len(data.get("muestras", [])), 2)

    def test_parsear_recepcion_roca(self):
        content = self._create_dummy_excel("F-LEM-P-01.04 - ROCA")
        data = self.excel_logic.parsear_recepcion(content)
        self.assertEqual(data.get("tipo_recepcion"), "ROCA")
        self.assertEqual(len(data.get("muestras", [])), 2)

    def test_parsear_recepcion_albanileria(self):
        content = self._create_dummy_excel("F-LEM-P-01.05 - ALBAÑILERIA")
        data = self.excel_logic.parsear_recepcion(content)
        self.assertEqual(data.get("tipo_recepcion"), "ALBANILERIA")
        self.assertEqual(len(data.get("muestras", [])), 2)

    def test_parsear_recepcion_agua(self):
        content = self._create_dummy_excel("F-LEM-P-01.06 - AGUA")
        data = self.excel_logic.parsear_recepcion(content)
        self.assertEqual(data.get("tipo_recepcion"), "AGUA")
        self.assertEqual(len(data.get("muestras", [])), 2)

    def test_parsear_recepcion_suelo_agregado(self):
        content = self._create_dummy_excel("F-LEM-P-01.13 - SUELO Y AGREGADO")
        data = self.excel_logic.parsear_recepcion(content)
        self.assertEqual(data.get("tipo_recepcion"), "SUELO_AGREGADO")
        self.assertEqual(len(data.get("muestras", [])), 2)

    def test_endpoint_importar_excel(self):
        content = self._create_dummy_excel("F-LEM-P-01.02")
        response = self.client.post(
            "/api/recepcion/importar-excel",
            files={"file": ("test_recepcion.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        self.assertEqual(response.status_code, 200)
        res_json = response.json()
        self.assertIn("tipo_recepcion", res_json)
        self.assertEqual(res_json["tipo_recepcion"], "CONCRETO")

if __name__ == "__main__":
    unittest.main()
