from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.verificacion.excel import ExcelLogic


class TestVerificacionExcelSpacing(unittest.TestCase):
    def setUp(self):
        self.logic = ExcelLogic()
        self.template_path = PROJECT_ROOT / "app" / "templates" / "Template_Verificacion.xlsx"

    def _make_sample(self, item_numero: int) -> SimpleNamespace:
        return SimpleNamespace(
            item_numero=item_numero,
            codigo_lem=f"LEM-{item_numero:03d}",
            tipo_testigo="4in x 8in",
            diametro_1_mm=150.0,
            diametro_2_mm=150.1,
            tolerancia_porcentaje=2.0,
            aceptacion_diametro="CUMPLE",
            perpendicularidad_sup1=True,
            perpendicularidad_sup2=True,
            perpendicularidad_inf1=True,
            perpendicularidad_inf2=True,
            perpendicularidad_medida=True,
            planitud_superior_aceptacion="CUMPLE",
            planitud_inferior_aceptacion="CUMPLE",
            planitud_depresiones_aceptacion="CUMPLE",
            accion_realizar="CAPEO",
            conformidad=".",
            longitud_1_mm=200.0,
            longitud_2_mm=200.1,
            longitud_3_mm=200.2,
            masa_muestra_aire_g=3966.0,
            pesar="",
        )

    def _make_verificacion(self, sample_count: int) -> SimpleNamespace:
        return SimpleNamespace(
            numero_verificacion="N-001",
            verificado_por="ANDRES SANCHEZ",
            fecha_verificacion="2026-05-05",
            cliente="De Vicente Constructora",
            equipo_bernier="EQ1",
            equipo_lainas_1="EQ2",
            equipo_lainas_2="EQ3",
            equipo_escuadra="EQ4",
            equipo_balanza="EQ5",
            nota="-",
            muestras_verificadas=[self._make_sample(i) for i in range(1, sample_count + 1)],
        )

    def test_row_17_keeps_template_spacer_height_when_not_needed(self):
        verificacion = self._make_verificacion(sample_count=3)

        workbook_bytes = self.logic.generar_excel_verificacion(verificacion)
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
        sheet = workbook.active

        template_wb = load_workbook(self.template_path, data_only=False)
        template_sheet = template_wb.active

        self.assertEqual(sheet.row_dimensions[17].height, template_sheet.row_dimensions[17].height)
        self.assertEqual(sheet["B18"].value, "Còdigo equipo")
        self.assertEqual(sheet["A19"].value, "Nota")

    def test_eighth_sample_still_promotes_row_17_and_shifts_footer(self):
        verificacion = self._make_verificacion(sample_count=8)

        workbook_bytes = self.logic.generar_excel_verificacion(verificacion)
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
        sheet = workbook.active

        self.assertEqual(sheet["A17"].value, 8)
        self.assertEqual(sheet["B19"].value, "Còdigo equipo")
        self.assertEqual(sheet["A20"].value, "Nota")

    def test_main_headers_keep_wrap_text_for_long_titles(self):
        verificacion = self._make_verificacion(sample_count=3)

        workbook_bytes = self.logic.generar_excel_verificacion(verificacion)
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
        sheet = workbook.active

        self.assertTrue(sheet["B8"].alignment.wrap_text)
        self.assertTrue(sheet["C8"].alignment.wrap_text)
        self.assertTrue(sheet["H8"].alignment.wrap_text)
        self.assertTrue(sheet["R8"].alignment.wrap_text)
        self.assertTrue(sheet["U8"].alignment.wrap_text)


if __name__ == "__main__":
    unittest.main()
