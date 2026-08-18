from __future__ import annotations

import sys
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.database import Base
from app.modules.recepcion.models import MuestraConcreto, RecepcionMuestra
from app.modules.compresion.models import EnsayoCompresion, ItemCompresion
from app.modules.control_probetas.router import get_control_probetas, get_control_probetas_kpis


class TestControlProbetasRouter(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        self.Session = sessionmaker(bind=self.engine, autocommit=False, autoflush=False)
        Base.metadata.create_all(self.engine)
        self.db = self.Session()

        # Create Reception
        recepcion = RecepcionMuestra(
            numero_ot="OT-1234-26",
            numero_recepcion="1234-REC",
            cliente="Geofal Peru SAC",
            domicilio_legal="LIMA",
            ruc="20987654321",
            persona_contacto="ADMINISTRADOR",
            email="admin@geofal.com",
            telefono="987654321",
            solicitante="ADMINISTRADOR",
            domicilio_solicitante="LIMA",
            proyecto="Edificio Central",
            ubicacion="MIRAFLORES",
            fecha_recepcion=datetime(2026, 6, 1),
            emision_fisica=False,
            emision_digital=True,
            estado="PENDIENTE",
        )
        self.db.add(recepcion)
        self.db.flush()

        today_str = datetime.now().strftime("%Y/%m/%d")
        yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y/%m/%d")
        tomorrow_str = (datetime.now() + timedelta(days=1)).strftime("%Y/%m/%d")

        # Create Muestras
        # 1. Specimen Curing (tomorrow break date)
        self.muestra_curado = MuestraConcreto(
            recepcion_id=recepcion.id,
            item_numero=1,
            codigo_muestra="001",
            codigo_muestra_lem="7001-CO-26",
            identificacion_muestra="Cilindro 1",
            estructura="Zapata Z-1",
            fc_kg_cm2=210,
            fecha_moldeo="2026/06/01",
            edad=7,
            fecha_rotura=tomorrow_str,
            requiere_densidad=False,
            es_control_probetas=True,
        )
        
        # 2. Specimen Pending (today break date)
        self.muestra_pendiente = MuestraConcreto(
            recepcion_id=recepcion.id,
            item_numero=2,
            codigo_muestra="002",
            codigo_muestra_lem="7002-CO-26",
            identificacion_muestra="Cilindro 2",
            estructura="Zapata Z-1",
            fc_kg_cm2=210,
            fecha_moldeo="2026/06/01",
            edad=7,
            fecha_rotura=today_str,
            requiere_densidad=True,
            es_control_probetas=True,
        )

        # 3. Specimen Overdue (yesterday break date, no test results)
        self.muestra_vencido = MuestraConcreto(
            recepcion_id=recepcion.id,
            item_numero=3,
            codigo_muestra="003",
            codigo_muestra_lem="7003-CO-26",
            identificacion_muestra="Cilindro 3",
            estructura="Columna C-1",
            fc_kg_cm2=280,
            fecha_moldeo="2026/05/28",
            edad=7,
            fecha_rotura=yesterday_str,
            requiere_densidad=False,
            es_control_probetas=True,
        )

        # 4. Specimen Crushed (tested with results)
        self.muestra_ensayado = MuestraConcreto(
            recepcion_id=recepcion.id,
            item_numero=4,
            codigo_muestra="004",
            codigo_muestra_lem="7004-CO-26",
            identificacion_muestra="Cilindro 4",
            estructura="Viga V-1",
            fc_kg_cm2=280,
            fecha_moldeo="2026/05/28",
            edad=7,
            fecha_rotura=yesterday_str,
            requiere_densidad=False,
            es_control_probetas=True,
        )

        self.db.add(self.muestra_curado)
        self.db.add(self.muestra_pendiente)
        self.db.add(self.muestra_vencido)
        self.db.add(self.muestra_ensayado)
        self.db.flush()

        # Add Compression test records for sample 4
        ensayo = EnsayoCompresion(
            recepcion_id=recepcion.id,
            numero_ot="OT-1234-26",
            numero_recepcion="1234-REC",
            estado="COMPLETADO",
            fecha_creacion=datetime(2026, 6, 5),
        )
        self.db.add(ensayo)
        self.db.flush()

        item_comp = ItemCompresion(
            ensayo_id=ensayo.id,
            item=4,
            codigo_lem="7004-CO-26",
            carga_maxima=255.4,
            tipo_fractura="1",
            fecha_ensayo=datetime.now(),
        )
        self.db.add(item_comp)
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_kpis_calculation(self):
        kpis = get_control_probetas_kpis(db=self.db)
        self.assertEqual(kpis.total, 4)
        self.assertEqual(kpis.curado, 1)      # 1 tomorrow
        self.assertEqual(kpis.pendiente, 1)   # 1 today
        self.assertEqual(kpis.vencido, 1)     # 1 yesterday (no results)
        self.assertEqual(kpis.ensayado, 1)    # 1 yesterday (with results)

    def test_get_control_probetas_list_no_filters(self):
        response = get_control_probetas(page=1, page_size=10, db=self.db)
        self.assertEqual(response.total, 4)
        self.assertEqual(len(response.items), 4)

        # Check mapping status values
        status_map = {item.muestra_id: item.estado_probeta for item in response.items}
        self.assertEqual(status_map[self.muestra_curado.id], "curado")
        self.assertEqual(status_map[self.muestra_pendiente.id], "pendiente")
        self.assertEqual(status_map[self.muestra_vencido.id], "vencido")
        self.assertEqual(status_map[self.muestra_ensayado.id], "ensayado")

    def test_filter_by_status(self):
        response = get_control_probetas(page=1, page_size=10, estado="pendiente", db=self.db)
        self.assertEqual(response.total, 1)
        self.assertEqual(response.items[0].muestra_id, self.muestra_pendiente.id)

    def test_filter_by_search_text(self):
        # Search for client name
        response = get_control_probetas(page=1, page_size=10, search="Geofal", db=self.db)
        self.assertEqual(response.total, 4)

        # Search for unique identification
        response = get_control_probetas(page=1, page_size=10, search="Cilindro 3", db=self.db)
        self.assertEqual(response.total, 1)
        self.assertEqual(response.items[0].muestra_id, self.muestra_vencido.id)

    def test_update_status_entrega_entregado(self):
        from app.modules.control_probetas.router import update_probeta
        from unittest.mock import MagicMock
        req = MagicMock()
        req.state.user = {}
        req.headers = {}
        # Mark the overdue specimen as ENTREGADO
        updated = update_probeta(
            muestra_id=self.muestra_vencido.id,
            payload={"status_entrega": "ENTREGADO", "status_ensayo": "ENSAYADO", "fecha_entrega": "2026/08/05"},
            request=req,
            db=self.db
        )
        self.assertEqual(updated.estado_probeta, "ensayado")
        self.assertEqual(updated.status_ensayo, "ENTREGADO")

        kpis = get_control_probetas_kpis(db=self.db)
        self.assertEqual(kpis.vencido, 0)
        self.assertEqual(kpis.ensayado, 2)

    def test_update_status_entregado_auto_fecha(self):
        from app.modules.control_probetas.router import update_probeta, LIMA_TZ
        from datetime import datetime
        from unittest.mock import MagicMock
        req = MagicMock()
        req.state.user = {}
        req.headers = {}
        
        # When setting status to ENTREGADO without passing fecha_entrega
        updated = update_probeta(
            muestra_id=self.muestra_pendiente.id,
            payload={"status": "ENTREGADO"},
            request=req,
            db=self.db
        )
        today_str = datetime.now(LIMA_TZ).strftime("%Y/%m/%d")
        self.assertEqual(updated.status_ensayo, "ENTREGADO")
        self.assertEqual(updated.fecha_entrega, today_str)

        # When resetting back to FALTA
        reset = update_probeta(
            muestra_id=self.muestra_pendiente.id,
            payload={"status": "FALTA"},
            request=req,
            db=self.db
        )
        self.assertEqual(reset.status_ensayo, "FALTA")
        self.assertEqual(reset.fecha_entrega, "-")

    def test_ot_excel_concreto_row_33_injection(self):
        import openpyxl, io
        from app.modules.ot.models import OrdenTrabajo
        from app.modules.ot.excel import generar_excel_ot_concreto

        ot = OrdenTrabajo(
            numero_ot="OT-1953",
            numero_recepcion="1953-26",
            cliente="CLIENTE DE PRUEBA",
            proyecto="PROYECTO DE PRUEBA",
            fecha_recepcion="2026/08/15",
            ot_aperturada_por="BETZABETH ZARABIA",
            ot_designada_a="DEYVI INFANZÓN",
            items=[{"item": 1, "codigo_muestra": "M-1", "elemento": "4 in x 8 in", "fecha_rotura": "2026/08/15"}]
        )
        buf = generar_excel_ot_concreto(ot)
        wb = openpyxl.load_workbook(buf)
        ws = wb["MYP"]
        self.assertEqual(ws["C33"].value, "BETZABETH ZARABIA")
        self.assertTrue("OT DESIGNADA A:" in (ws["F33"].value or ""))
        self.assertEqual(ws["I33"].value, "DEYVI INFANZÓN")

    def test_ot_excel_concreto_dynamic_row_expansion_15_items(self):
        import openpyxl
        from app.modules.ot.models import OrdenTrabajo
        from app.modules.ot.excel import generar_excel_ot_concreto

        items_15 = [
            {
                "item": i,
                "codigo_muestra": f"PROB-{i:02d}",
                "elemento": "4 in x 8 in",
                "fecha_rotura": "2026/08/20",
                "densidad": "SI",
                "edad": 7,
                "fc_kg_cm2": 210,
            }
            for i in range(1, 16)
        ]
        ot = OrdenTrabajo(
            numero_ot="OT-1924-26",
            numero_recepcion="1920-26",
            cliente="CLIENTE DE PRUEBA 15",
            proyecto="PROYECTO DE PRUEBA 15",
            fecha_recepcion="2026/08/15",
            ot_aperturada_por="BETZABETH ZARABIA",
            ot_designada_a="DEYVI INFANZÓN",
            items=items_15,
        )
        buf = generar_excel_ot_concreto(ot)
        wb = openpyxl.load_workbook(buf)
        ws = wb["MYP"]

        # Check all 15 items in rows 9 to 23
        for i in range(1, 16):
            r = 8 + i
            self.assertEqual(ws[f"A{r}"].value, i)
            self.assertEqual(ws[f"B{r}"].value, f"PROB-{i:02d}")
            self.assertEqual(ws[f"C{r}"].value, "COMPRESION PROBETAS ASTM C39/C39M")
            self.assertEqual(ws[f"F{r}"].value, "4 in x 8 in")
            self.assertEqual(ws[f"G{r}"].value, "2026/08/20")
            self.assertEqual(ws[f"H{r}"].value, "SI")
            self.assertEqual(ws[f"I{r}"].value, 7)
            self.assertEqual(ws[f"J{r}"].value, 210)

        # Check shifted footer (24 + 3 = 27, 33 + 3 = 36)
        self.assertEqual(ws["C27"].value, "2026/08/15")
        self.assertEqual(ws["F27"].value, "2026/08/20")
        self.assertEqual(ws["J27"].value, "2026/08/20")
        self.assertEqual(ws["C36"].value, "BETZABETH ZARABIA")
        self.assertTrue("OT DESIGNADA A:" in (ws["F36"].value or ""))
        self.assertEqual(ws["I36"].value, "DEYVI INFANZÓN")

    def test_evaluate_ot_estado_complete_and_incomplete(self):
        from app.modules.ot.models import OrdenTrabajo
        from app.modules.ot.router import _evaluate_ot_estado

        ot_complete = OrdenTrabajo(
            numero_ot="OT-1960-26",
            numero_recepcion="1960-26",
            cliente="De Vicente Constructora",
            proyecto="LIMA SUR UCS",
            fecha_recepcion="2026-08-17",
            ot_aperturada_por="BETZABETH ZARABIA",
            ot_designada_a="DEYVI INFANZON",
            items=[{"item": 1, "elemento": "VIGA", "codigo_muestra": "14678-CO-26"}],
            estado="PENDIENTE",
        )
        self.assertEqual(_evaluate_ot_estado(ot_complete), "EMITIDO")

        ot_incomplete = OrdenTrabajo(
            numero_ot="OT-1961-26",
            numero_recepcion="1961-26",
            cliente="De Vicente Constructora",
            proyecto="LIMA SUR UCS",
            fecha_recepcion="2026-08-17",
            ot_aperturada_por="BETZABETH ZARABIA",
            ot_designada_a="DEYVI INFANZON",
            items=[{"item": 1, "elemento": "-", "codigo_muestra": "14678-CO-26"}],
            estado="PENDIENTE",
        )
        self.assertEqual(_evaluate_ot_estado(ot_incomplete), "PENDIENTE")


if __name__ == "__main__":
    unittest.main()
