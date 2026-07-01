import io
import sys
from pathlib import Path
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.abra.excel import TEMPLATE_PATH, generate_abra_excel
from app.modules.abra.schemas import AbraRequest

def test_abra_excel_template_integrity_and_generation():
    assert Path(TEMPLATE_PATH).exists(), f"Template not found at: {TEMPLATE_PATH}"

    payload = AbraRequest(
        muestra="147-SU-26",
        numero_ot="1000-26",
        fecha_ensayo="2026/05/07",
        realizado_por="D.I.C",
        masa_muestra_inicial_g=5000.0,
        masa_muestra_inicial_seca_g=4980.0,
        masa_muestra_inicial_seca_constante_g=4980.0,
        requiere_lavado="SI",
        tmn='3/4"',
        masa_12_esferas_g=5000.0,
        gradacion_1_tamiz_g=[1250, 1250, 1250, 1250, 0, 0],
        gradacion_2_tamiz_g=[0]*6,
        gradacion_3_tamiz_g=[0]*6,
        item_a_masa_original_g=[5000.0, 0, 0],
        item_b_masa_retenida_tamiz_12_g=[4000.0, 0, 0],
        item_c_masa_lavada_seca_retenida_g=[3950.0, 0, 0],
        item_d_masa_lavada_seca_constante_g=[3950.0, 0, 0],
        item_e_diferencia_masa_g=[1050.0, 0, 0],
        item_f_desgaste_pct=[21.0, 0, 0],
        item_perdida_lavado_pct=[1.0, 0, 0],
        horno_codigo="EQP-0150",
        maquina_los_angeles_codigo="EQP-0043",
        balanza_1g_codigo="EQP-0054",
        malla_no_12_codigo="INS-0144",
        malla_no_4_codigo="INS-0053",
        observaciones="Test de abrasión menores",
        revisado_por="REVISOR TEST",
        revisado_fecha="2026/05/08",
        aprobado_por="APROBADOR TEST",
        aprobado_fecha="2026/05/09"
    )

    # Generate the excel
    excel_bytes = generate_abra_excel(payload)
    assert isinstance(excel_bytes, bytes)

    # Load and check sheet names/structures
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=False)
    # The new template sheet name is typically "FORMATO" or similar
    assert len(wb.sheetnames) > 0
    ws = wb.worksheets[0]
    
    # Check that basic header cells are filled
    assert ws["C11"].value == "147-SU-26"
    assert ws["D11"].value == "1000-26"
    assert ws["F11"].value == "2026/05/07"
