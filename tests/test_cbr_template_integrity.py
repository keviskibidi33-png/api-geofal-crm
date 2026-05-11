from __future__ import annotations

import sys
import unittest
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.cbr.excel import generate_cbr_excel
from app.modules.cbr.schemas import CBRRequest


class TestCBRTemplateIntegrity(unittest.TestCase):
    @staticmethod
    def _build_payload() -> CBRRequest:
        return CBRRequest.model_validate(
            {
                "muestra": "587-SU-26",
                "numero_ot": "OT-589-26",
                "fecha_ensayo": "25/04/26",
                "realizado_por": "ANDRES SANCHEZ",
                "sobretamano_porcentaje": 5.5,
                "masa_grava_adicionada_g": 1200.0,
                "condicion_muestra_saturado": "SI",
                "condicion_muestra_sin_saturar": "NO",
                "maxima_densidad_seca": 2.11,
                "optimo_contenido_humedad": 11.5,
                "temperatura_inicial_c": 20.1,
                "temperatura_final_c": 22.3,
                "tamano_maximo_visual_in": "3/4",
                "descripcion_muestra_astm": "CBR ASTM demo",
                "golpes_por_especimen": [56, 25, 10],
                "codigo_molde_por_especimen": ["M-01", "M-02", "M-03"],
                "temperatura_inicio_c_por_columna": [20, 20, 20, 20, 20, 20],
                "temperatura_final_c_por_columna": [22, 22, 22, 22, 22, 22],
                "masa_molde_suelo_g_por_columna": [1000, 1001, 1002, 1003, 1004, 1005],
                "codigo_tara_por_columna": ["T1", "T2", "T3", "T4", "T5", "T6"],
                "masa_tara_g_por_columna": [10, 10, 10, 10, 10, 10],
                "masa_suelo_humedo_tara_g_por_columna": [20, 21, 22, 23, 24, 25],
                "masa_suelo_seco_tara_g_por_columna": [15, 16, 17, 18, 19, 20],
                "masa_suelo_seco_tara_constante_g_por_columna": [14, 15, 16, 17, 18, 19],
                "lecturas_penetracion": [
                    {"tension_standard": 0.0, "lectura_dial_esp_01": 0, "lectura_dial_esp_02": 0, "lectura_dial_esp_03": 0},
                    {"tension_standard": 0.025, "lectura_dial_esp_01": 1.1, "lectura_dial_esp_02": 1.2, "lectura_dial_esp_03": 1.3},
                    {"tension_standard": 0.05, "lectura_dial_esp_01": 1.4, "lectura_dial_esp_02": 1.5, "lectura_dial_esp_03": 1.6},
                ],
                "hinchamiento": [
                    {"fecha": "25/04/26", "hora": "08:00", "esp_01": 0.1, "esp_02": 0.2, "esp_03": 0.3},
                    {"fecha": "25/04/26", "hora": "09:00", "esp_01": 0.1, "esp_02": 0.2, "esp_03": 0.3},
                    {"fecha": "25/04/26", "hora": "10:00", "esp_01": 0.1, "esp_02": 0.2, "esp_03": 0.3},
                    {"fecha": "25/04/26", "hora": "11:00", "esp_01": 0.1, "esp_02": 0.2, "esp_03": 0.3},
                    {"fecha": "25/04/26", "hora": "12:00", "esp_01": 0.1, "esp_02": 0.2, "esp_03": 0.3},
                ],
                "profundidad_hendidura_mm_por_celda": [0.12, 0.34, 0.56],
                "equipo_cbr": "CBR-01",
                "equipo_dial_deformacion": "DIAL-01",
                "equipo_dial_expansion": "EXP-01",
                "equipo_horno_110": "HORNO-01",
                "equipo_pison": "PISON-01",
                "equipo_balanza_1g": "BAL-1G",
                "equipo_balanza_01g": "BAL-0.1G",
                "observaciones": "Ensayo de prueba",
                "revisado_por": "FABIAN LA ROSA",
                "revisado_fecha": "25/04/26",
                "aprobado_por": "IRMA COAQUIRA",
                "aprobado_fecha": "25/04/26",
            }
        )

    def test_generate_cbr_excel_is_self_contained(self):
        payload = self._build_payload()
        content = generate_cbr_excel(payload)

        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertIn("cbr graficos", workbook.sheetnames)
        self.assertEqual(workbook["cbr graficos"].sheet_state, "hidden")

        with zipfile.ZipFile(BytesIO(content), "r") as archive:
            names = set(archive.namelist())
            self.assertEqual([name for name in names if name.startswith("xl/externalLinks/")], [])
            self.assertNotIn("xl/calcChain.xml", names)

            workbook_xml = archive.read("xl/workbook.xml").decode("utf-8", errors="ignore")
            self.assertNotIn("<externalReferences>", workbook_xml)
            self.assertNotIn("#REF!", workbook_xml)
            self.assertNotIn("[3]cbr graficos", workbook_xml)

            chart_names = sorted(name for name in names if name.startswith("xl/charts/chart") and name.endswith(".xml"))
            self.assertEqual(len(chart_names), 7)
            for chart_name in chart_names:
                chart_xml = archive.read(chart_name).decode("utf-8", errors="ignore")
                self.assertNotIn("#REF!", chart_xml)
                self.assertNotIn("[3]cbr graficos", chart_xml)


if __name__ == "__main__":
    unittest.main()
