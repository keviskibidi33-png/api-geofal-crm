import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.common.excel_xml import find_template_path
from app.modules.sul_magnesio.excel import TEMPLATE_FILE, generate_sul_magnesio_excel
from app.modules.sul_magnesio.schemas import SulMagnesioRequest


def _build_payload(**overrides) -> SulMagnesioRequest:
    payload = {
        "muestra": "22401-26",
        "numero_ot": "313233-26",
        "fecha_ensayo": "23/03/26",
        "realizado_por": "JORGE",
        "revisado_por": "FABIAN LA ROSA",
        "revisado_fecha": "23/03/26",
        "aprobado_por": "IRMA COAQUIRA",
        "aprobado_fecha": "23/03/26",
        "fino_rows": [
            {
                "gradacion_pct": 5,
                "masa_fraccion_ensayo_g": 100,
                "masa_material_retenido_g": 96,
            },
            {
                "gradacion_pct": 10,
                "masa_fraccion_ensayo_g": 200,
                "masa_material_retenido_g": 192,
            },
        ],
        "grueso_rows": [
            {
                "gradacion_pct": 25,
                "masa_individual_tamiz_g": 500,
                "masa_fraccion_ensayo_g": 100,
                "masa_material_retenido_g": 98,
            },
            {
                "gradacion_pct": 75,
                "masa_individual_tamiz_g": 900,
                "masa_fraccion_ensayo_g": 200,
                "masa_material_retenido_g": 194,
            },
        ],
        "cualitativo_rows": [
            {
                "total_particulas": 20,
                "rajadas_num": 2,
                "desmoronadas_num": 1,
                "fracturadas_num": 1,
                "astilladas_num": 0,
            }
        ],
        "horno_codigo": "HOR-110",
        "balanza_01_codigo": "BAL-0001",
    }
    payload.update(overrides)
    return SulMagnesioRequest(**payload)


def test_generate_sul_magnesio_excel_inyecta_totales_en_celdas_finales():
    payload = _build_payload()

    workbook = load_workbook(io.BytesIO(generate_sul_magnesio_excel(payload)), data_only=False)
    sheet = workbook["DURAB MAGNESIO"]

    assert sheet["G25"].value == "Σ Porcentaje de perdida"
    assert sheet["I25"].value == payload.fino_total_pct
    assert sheet["H25"].value is None

    assert sheet["H38"].value == "Σ Porcentaje de perdida"
    assert sheet["J38"].value == payload.grueso_total_pct
    assert sheet["I38"].value is None


def test_generate_sul_magnesio_excel_preserva_integridad_del_template():
    payload = _build_payload()

    template_sheet = load_workbook(find_template_path(TEMPLATE_FILE), data_only=False)["DURAB MAGNESIO"]
    generated_sheet = load_workbook(io.BytesIO(generate_sul_magnesio_excel(payload)), data_only=False)["DURAB MAGNESIO"]

    assert template_sheet.merged_cells.ranges == generated_sheet.merged_cells.ranges
