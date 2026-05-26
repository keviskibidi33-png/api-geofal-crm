from __future__ import annotations

import sys
import tempfile
import unittest
import zipfile
from collections import Counter
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.programacion.excel import export_programacion_xlsx


class TestProgramacionExcelIntegrity(unittest.TestCase):
    def test_lab_export_keeps_sheet_rows_unique_and_within_excel_limit(self):
        from app.modules.common.excel_xml import find_template_path
        template_path = find_template_path("Template_Programacion.xlsx")
        items = [
            {
                "item_numero": 1,
                "recep_numero": "REC-1",
                "ot": "OT-1",
                "codigo_muestra": "M-1",
                "fecha_recepcion": "2026-05-11",
                "fecha_inicio": "2026-05-11",
                "fecha_entrega_estimada": "2026-05-12",
                "cliente_nombre": "Cliente A",
                "descripcion_servicio": "Servicio A",
                "proyecto": "Proyecto A",
                "entrega_real": "",
                "estado_trabajo": "PENDIENTE",
                "cotizacion_lab": "COT-1",
                "autorizacion_lab": "",
                "nota_lab": "",
                "dias_atraso_lab": 0,
                "motivo_dias_atraso_lab": "",
                "evidencia_envio_recepcion": "",
                "envio_informes": "",
            },
            {
                "item_numero": 2,
                "recep_numero": "REC-2",
                "ot": "OT-2",
                "codigo_muestra": "M-2",
                "fecha_recepcion": "2026-05-11",
                "fecha_inicio": "2026-05-11",
                "fecha_entrega_estimada": "2026-05-12",
                "cliente_nombre": "Cliente B",
                "descripcion_servicio": "Servicio B",
                "proyecto": "Proyecto B",
                "entrega_real": "",
                "estado_trabajo": "PENDIENTE",
                "cotizacion_lab": "COT-2",
                "autorizacion_lab": "",
                "nota_lab": "",
                "dias_atraso_lab": 0,
                "motivo_dias_atraso_lab": "",
                "evidencia_envio_recepcion": "",
                "envio_informes": "",
            },
        ]

        output = export_programacion_xlsx(str(template_path), items)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(output.getvalue())
            tmp_path = Path(tmp.name)

        try:
            with zipfile.ZipFile(tmp_path, "r") as archive:
                self.assertIsNone(archive.testzip())
                sheet_root = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))

            ns = sheet_root.nsmap.get(None)
            row_numbers = [
                int(row.get("r"))
                for row in sheet_root.findall(f".//{{{ns}}}row")
                if row.get("r") and row.get("r").isdigit()
            ]
            counts = Counter(row_numbers)

            self.assertTrue(row_numbers)
            self.assertEqual(len(row_numbers), len(counts))
            self.assertLessEqual(max(row_numbers), 1048576)

            wb = load_workbook(tmp_path)
            self.assertIn("LABORATORIO", wb.sheetnames)
            self.assertEqual(wb["LABORATORIO"]["A9"].value, "1")
            self.assertEqual(wb["LABORATORIO"]["A10"].value, "2")
        finally:
            tmp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
