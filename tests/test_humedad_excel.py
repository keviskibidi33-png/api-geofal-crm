import io
import sys
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.humedad.excel import TEMPLATE_PATH as HUMEDAD_TEMPLATE_PATH
from app.modules.humedad.excel import generate_humedad_excel
from app.modules.humedad.schemas import HumedadRequest

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _build_humedad_payload() -> HumedadRequest:
    return HumedadRequest(
        muestra="147-SU-26",
        numero_ot="1000-26",
        fecha_ensayo="2026/05/07",
        realizado_por="D.I.C",
        condicion_masa_menor="SI",
        condicion_capas="NO",
        condicion_temperatura="NO",
        condicion_excluido="NO",
        descripcion_material_excluido=None,
        tipo_muestra="SUELO",
        condicion_muestra="ALTERADO",
        tamano_maximo_particula="3/4\"",
        forma_particula="SUBREDONDEADA",
        metodo_prueba="A",
        masa_recipiente_muestra_humeda=150,
        masa_recipiente_muestra_seca=120,
        masa_recipiente_muestra_seca_constante=120,
        masa_recipiente=20,
        numero_ensayo=1,
        recipiente_numero="R1",
        equipo_balanza_01="BAL-01",
        equipo_balanza_001="BAL-001",
        equipo_horno="HOR-110",
        observaciones="Ensayo de prueba",
        revisado_por="REVISOR H",
        revisado_fecha="2026/05/08",
        aprobado_por="APROBADOR H",
        aprobado_fecha="2026/05/09",
    )


def test_humedad_template_replaced_and_generation_forces_recalc():
    workbook = load_workbook(HUMEDAD_TEMPLATE_PATH, data_only=False)

    assert "Resumen" in workbook.sheetnames
    assert "informe ASTM" in workbook.sheetnames
    assert workbook["informe ASTM"]["C5"].value == "=+Resumen!P2"

    generated = load_workbook(io.BytesIO(generate_humedad_excel(_build_humedad_payload())), data_only=False)

    assert generated["informe NTP"]["D11"].value == "147-SU-26"
    assert generated["informe NTP"]["E11"].value == "1000-26"
    assert generated["informe NTP"]["G11"].value == "2026/05/07"
    assert generated["informe NTP"]["I11"].value == "D.I.C"
    assert generated["informe NTP"]["I37"].value == 30
    assert generated["informe NTP"]["I38"].value == 100
    assert generated["informe NTP"]["I39"].value == 30
    assert generated["Resumen"]["C55"].value == "Revisado:\n\nREVISOR H\n\nFecha: 2026/05/08"
    assert generated["Resumen"]["G55"].value == "Aprobado:\n\nAPROBADOR H\n\nFecha: 2026/05/09"

    with zipfile.ZipFile(io.BytesIO(generate_humedad_excel(_build_humedad_payload())), "r") as archive:
        names = set(archive.namelist())
        assert "xl/calcChain.xml" not in names
        assert all(not name.startswith("xl/externalLinks/") for name in names)
        assert "xl/workbook.xml" in names

        workbook_root = etree.fromstring(archive.read("xl/workbook.xml"))
        ns = {"m": NS_MAIN}
        workbook_pr = workbook_root.find("m:workbookPr", ns)
        assert workbook_pr is not None
        assert workbook_pr.get("updateLinks") == "never"
        assert workbook_root.find("m:externalReferences", ns) is None
        defined_names = workbook_root.find("m:definedNames", ns)
        assert defined_names is not None
        for defined_name in defined_names.findall("m:definedName", ns):
            text = (defined_name.text or "").strip()
            assert "#REF!" not in text
            assert "[" not in text
        calc_pr = workbook_root.find("m:calcPr", ns)
        assert calc_pr is not None
        assert calc_pr.get("calcMode") == "auto"
        assert calc_pr.get("fullCalcOnLoad") == "1"
        assert calc_pr.get("forceFullCalc") == "1"

        rels_root = etree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = [rel.get("Target", "") for rel in rels_root]
        assert all(not target.endswith("calcChain.xml") for target in rel_targets)
        assert all("externalLinks/" not in target for target in rel_targets)

        content_types_root = etree.fromstring(archive.read("[Content_Types].xml"))
        overrides = [override.get("PartName", "") for override in content_types_root]
        assert "/xl/calcChain.xml" not in overrides
        assert all(not part_name.startswith("/xl/externalLinks/") for part_name in overrides)
