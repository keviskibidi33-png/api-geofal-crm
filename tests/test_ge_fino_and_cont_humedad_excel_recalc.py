import io
import sys
import zipfile
import unittest
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.cont_humedad.excel import TEMPLATE_PATH as CONT_HUMEDAD_TEMPLATE_PATH
from app.modules.cont_humedad.excel import generate_cont_humedad_excel
from app.modules.cont_humedad.schemas import ContHumedadRequest
from app.modules.ge_fino.excel import TEMPLATE_PATH as GE_FINO_TEMPLATE_PATH
from app.modules.ge_fino.excel import generate_ge_fino_excel
from app.modules.ge_fino.schemas import GeFinoRequest

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _assert_recalc_and_links_cleaned(archive: zipfile.ZipFile) -> None:
    names = set(archive.namelist())

    assert "xl/calcChain.xml" not in names
    assert all(not name.startswith("xl/externalLinks/") for name in names)

    workbook_root = etree.fromstring(archive.read("xl/workbook.xml"))
    ns = {"m": NS_MAIN}
    workbook_pr = workbook_root.find("m:workbookPr", ns)
    assert workbook_pr is not None
    assert workbook_pr.get("updateLinks") == "never"
    assert workbook_root.find("m:externalReferences", ns) is None

    calc_pr = workbook_root.find("m:calcPr", ns)
    assert calc_pr is not None
    assert calc_pr.get("calcMode") == "auto"
    assert calc_pr.get("fullCalcOnLoad") == "1"
    assert calc_pr.get("forceFullCalc") == "1"

    rels_root = etree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = [rel.get("Target", "") for rel in rels_root]
    assert all(not target.endswith("calcChain.xml") for target in rel_targets)
    assert all("externalLinks/" not in target for target in rel_targets)

    content_types_root = etree.fromstring(archive.read("[Content_Types].xml"))
    overrides = [override.get("PartName", "") for override in content_types_root]
    assert "/xl/calcChain.xml" not in overrides
    assert all(not part_name.startswith("/xl/externalLinks/") for part_name in overrides)


class TestGeFinoAndContHumedadExcelRecalc(unittest.TestCase):
    def test_ge_fino_generation_forces_recalc_and_preserves_report_cells(self):
        assert Path(GE_FINO_TEMPLATE_PATH).exists()

        payload = GeFinoRequest(
            muestra="147-SU-26",
            numero_ot="1000-26",
            fecha_ensayo="2026/05/07",
            realizado_por="D.I.C",
            masa_humeda_g=500,
            masa_seca_g=400,
            masa_seca_constante_g=390,
            fecha_hora_inmersion="2026/05/07 10:00",
            fecha_hora_salida_inmersion="2026/05/07 12:00",
            temp_picnometro_contenido_c=20,
            temp_durante_calibracion_c=22,
            valor_s_g=500,
            valor_c_g=0,
            valor_b_g=0,
            valor_d_g="-",
            valor_e_g=0,
            valor_f_g=0,
            valor_g_g=0,
            valor_a_g=500,
            densidad_relativa_od=1.0,
            densidad_relativa_ssd=1.1,
            densidad_relativa_aparente=1.2,
            absorcion_pct=3.3,
            seco_horno_110_si_no="SI",
            observaciones="demo",
            revisado_por="FABIAN LA ROSA",
            revisado_fecha="2026/05/08",
            aprobado_por="IRMA COAQUIRA",
            aprobado_fecha="2026/05/09",
        )

        assert payload.muestra == "147-AG-26"

        workbook = load_workbook(io.BytesIO(generate_ge_fino_excel(payload)), data_only=False)

        assert workbook["FORMATO"]["F11"].value == "D.I.C"
        assert workbook["DATOS ENSAYO"]["K6"].value == "D.I.C"
        import datetime
        assert workbook["Incertidumbre"]["B113"].value == "FABIAN LA ROSA"
        val_b115 = workbook["Incertidumbre"]["B115"].value
        if isinstance(val_b115, datetime.datetime):
            assert val_b115.date() == datetime.date(2022, 11, 4)
        else:
            assert str(val_b115).startswith("2022-11-04")
        assert workbook["Incertidumbre"]["G113"].value == "IRMA COAQUIRA"
        val_g115 = workbook["Incertidumbre"]["G115"].value
        if isinstance(val_g115, datetime.datetime):
            assert val_g115.date() == datetime.date(2022, 11, 5)
        else:
            assert str(val_g115).startswith("2022-11-05")

        with zipfile.ZipFile(io.BytesIO(generate_ge_fino_excel(payload)), "r") as archive:
            _assert_recalc_and_links_cleaned(archive)

    def test_cont_humedad_generation_forces_recalc_and_preserves_report_cells(self):
        assert Path(CONT_HUMEDAD_TEMPLATE_PATH).exists()

        payload = ContHumedadRequest(
            muestra="147-SU-26",
            numero_ot="1000-26",
            fecha_ensayo="2026/05/07",
            realizado_por="D.I.C",
            condicion_masa_menor="SI",
            condicion_capas="NO",
            condicion_temperatura="NO",
            condicion_excluido="NO",
            descripcion_material_excluido=None,
            tipo_muestra="SUELO",
            condicion_muestra="ALTERADO",
            tamano_maximo_muestra_visual_in='3/4"',
            forma_particula="ANGULAR",
            metodo_prueba="A",
            masa_recipiente_muestra_humedo_g=100,
            masa_recipiente_muestra_seco_g=80,
            masa_recipiente_muestra_seco_constante_g=75,
            masa_agua_g=20,
            masa_recipiente_g=10,
            masa_muestra_seco_g=65,
            contenido_humedad_pct=30,
            numero_ensayo=1,
            recipiente_numero="ASA",
            balanza_01g_codigo="BAL-01",
            balanza_001g_codigo="BAL-001",
            horno_110c_codigo="HOR-110",
            revisado_por="FABIAN LA ROSA",
            revisado_fecha="2026/05/08",
            aprobado_por="IRMA COAQUIRA",
            aprobado_fecha="2026/05/09",
        )

        workbook = load_workbook(io.BytesIO(generate_cont_humedad_excel(payload)), data_only=False)

        assert workbook["FORMATO"]["K11"].value == "D.I.C"
        assert workbook["Datos ensayo"]["G8"].value == "D.I.C"
        
        import datetime
        assert workbook["Incertidumbre"]["B59"].value == "FABIAN LA ROSA"
        val_b61 = workbook["Incertidumbre"]["B61"].value
        if isinstance(val_b61, datetime.datetime):
            assert val_b61.date() == datetime.date(2022, 11, 4)
        else:
            assert str(val_b61).startswith("2022-11-04")
        assert workbook["Incertidumbre"]["G59"].value == "IRMA COAQUIRA"
        val_g61 = workbook["Incertidumbre"]["G61"].value
        if isinstance(val_g61, datetime.datetime):
            assert val_g61.date() == datetime.date(2022, 11, 5)
        else:
            assert str(val_g61).startswith("2022-11-05")

        with zipfile.ZipFile(io.BytesIO(generate_cont_humedad_excel(payload)), "r") as archive:
            _assert_recalc_and_links_cleaned(archive)
