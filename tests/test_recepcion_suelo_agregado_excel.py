import io
import json
import openpyxl
from datetime import datetime
from app.modules.recepcion.excel import ExcelLogic
from app.modules.recepcion.models import RecepcionMuestra, MuestraConcreto

def test_generar_excel_suelo_agregado_fidelity():
    excel_logic = ExcelLogic()

    recepcion = RecepcionMuestra(
        numero_recepcion="100-26",
        numero_ot="OT-100-26",
        numero_cotizacion="COT-500-26",
        tipo_recepcion="SUELO_AGREGADO",
        cliente="CONSTRUCTORA DEL SUR S.A.C.",
        domicilio_legal="AV. AREQUIPA 123",
        ruc="20123456789",
        persona_contacto="ING. MARIO LOPEZ",
        email="mlopez@delsur.pe",
        telefono="999888777",
        solicitante="SOLICITANTE TEST",
        domicilio_solicitante="DOMICILIO SOLICITANTE TEST",
        proyecto="MEJORAMIENTO CARRETERA",
        ubicacion="HUARI, ANCASH",
        fecha_recepcion=datetime(2026, 8, 20),
        fecha_estimada_culminacion=datetime(2026, 8, 28),
        emision_digital=True,
        emision_fisica=False,
        entregado_por="ING. MARIO LOPEZ",
        recibido_por="BETZABETH SARAVIA",
        observaciones="Muestras selladas en sacos.",
    )

    m1 = MuestraConcreto(
        item_numero=1,
        codigo_muestra_lem="1500-SU-26",
        identificacion_muestra="M-01 C-01 (0.00 - 1.50 M)",
        procedencia="CALICATA 01",
        cantera="CANTERA CENTRAL",
        cantidad="50 KG",
        ensayos_json=json.dumps([
            {"codigo": "SU24", "descripcion": "ANÁLISIS GRANULOMÉTRICO POR TAMIZADO EN SUELOS", "norma": "ASTM D6913/D6913M-17"},
            {"codigo": "SU23", "descripcion": "LÍMITES DE ATTERBERG (LL, LP)", "norma": "ASTM D4318-17ε1"},
            {"codigo": "SU20", "descripcion": "CONTENIDO DE HUMEDAD", "norma": "ASTM D2216-19"}
        ])
    )

    m2 = MuestraConcreto(
        item_numero=2,
        codigo_muestra_lem="1501-SU-26",
        identificacion_muestra="M-02 C-02 (0.00 - 2.00 M)",
        procedencia="CALICATA 02",
        cantera="CANTERA NORTE",
        cantidad="60 KG",
        ensayos_json=json.dumps([
            {"codigo": "SU24", "descripcion": "ANÁLISIS GRANULOMÉTRICO POR TAMIZADO EN SUELOS", "norma": "ASTM D6913/D6913M-17"},
            {"codigo": "SU22", "descripcion": "PESO ESPECÍFICO Y ABSORCIÓN", "norma": "ASTM C127"}
        ])
    )

    recepcion.muestras = [m1, m2]

    excel_bytes = excel_logic.generar_excel_recepcion(recepcion)
    assert excel_bytes is not None
    assert len(excel_bytes) > 5000

    # Verificar preservación 100% intacta de imágenes y dibujos
    import zipfile
    with zipfile.ZipFile(io.BytesIO(excel_bytes), 'r') as z_out:
        namelist = z_out.namelist()
        assert any('media' in name for name in namelist), "Falta el directorio media (logotipos/imagenes)"
        assert any('drawing' in name for name in namelist), "Falta el directorio drawings"

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb['RECEP. SU-AG']

    # Assert Headers
    assert ws['D6'].value == "100-26"
    assert ws['D7'].value == "COT-500-26"
    assert ws['M6'].value == "X"
    assert ws['M7'].value is None or ws['M7'].value == ""

    # Assert Customer
    assert ws['D10'].value == "CONSTRUCTORA DEL SUR S.A.C."
    assert ws['D12'].value == "20123456789"
    assert ws['D13'].value == "ING. MARIO LOPEZ"
    assert ws['D18'].value == "MEJORAMIENTO CARRETERA"

    # Assert Sample 1 (rows 22-25, h=max(4,3)=4)
    assert ws['A22'].value == 1
    assert ws['B22'].value == "1500-SU-26"
    assert ws['F22'].value == "M-01 C-01 (0.00 - 1.50 M)"
    assert ws['F23'].value == "CALICATA 01"
    assert ws['F24'].value == "CANTERA CENTRAL"
    assert ws['F25'].value == "50 KG"
    assert ws['H22'].value == "SU24"
    assert ws['H23'].value == "SU23"
    assert ws['H24'].value == "SU20"

    # Row 26 = separator (1 empty row)
    assert ws['A26'].value is None
    assert ws['H26'].value is None

    # Assert Sample 2 (correlative: rows 27-30, h=max(4,2)=4)
    assert ws['A27'].value == 2
    assert ws['B27'].value == "1501-SU-26"
    assert ws['F27'].value == "M-02 C-02 (0.00 - 2.00 M)"
    assert ws['F28'].value == "CALICATA 02"
    assert ws['F29'].value == "CANTERA NORTE"
    assert ws['F30'].value == "60 KG"
    assert ws['H27'].value == "SU24"
    assert ws['H28'].value == "SU22"

    # Row 31 = separator
    assert ws['A31'].value is None

    # Assert Footer (intact in template at rows 42 and 44)
    assert ws['C42'].value == "Muestras selladas en sacos."
    assert ws['C44'].value == "ING. MARIO LOPEZ"
    assert ws['K44'].value == "BETZABETH SARAVIA"

    print("Test generar_excel_suelo_agregado (2 samples correlative) passed successfully!")

    # Test 1 sample
    recepcion.muestras = [m1]
    excel_bytes_1 = excel_logic.generar_excel_recepcion(recepcion)
    wb_1 = openpyxl.load_workbook(io.BytesIO(excel_bytes_1), data_only=True)
    ws_1 = wb_1['RECEP. SU-AG']
    assert ws_1['A22'].value == 1
    assert ws_1['B22'].value == "1500-SU-26"
    assert ws_1['C42'].value == "Muestras selladas en sacos."
    assert ws_1['C44'].value == "ING. MARIO LOPEZ"
    print("Test generar_excel_suelo_agregado (1 sample) passed successfully!")

    # Test 3 samples (m1: 22-25, sep 26 | m2: 27-30, sep 31 | m3: 32-35, sep 36)
    m3 = MuestraConcreto(
        item_numero=3,
        codigo_muestra_lem="1502-SU-26",
        identificacion_muestra="M-03 C-03 (2.00 - 3.00 M)",
        procedencia="CALICATA 03",
        cantera="CANTERA SUR",
        cantidad="70 KG",
        ensayos_json=json.dumps([
            {"codigo": "SU20", "descripcion": "CONTENIDO DE HUMEDAD", "norma": "ASTM D2216-19"}
        ])
    )
    recepcion.muestras = [m1, m2, m3]
    excel_bytes_3 = excel_logic.generar_excel_recepcion(recepcion)
    wb_3 = openpyxl.load_workbook(io.BytesIO(excel_bytes_3), data_only=True)
    ws_3 = wb_3['RECEP. SU-AG']
    assert ws_3['A22'].value == 1
    assert ws_3['A27'].value == 2
    assert ws_3['A32'].value == 3
    assert ws_3['B32'].value == "1502-SU-26"
    assert ws_3['F32'].value == "M-03 C-03 (2.00 - 3.00 M)"
    assert ws_3['C42'].value == "Muestras selladas en sacos."
    assert ws_3['C44'].value == "ING. MARIO LOPEZ"
    # Test 20 samples (maximum capacity verification)
    muestras_20 = []
    for i in range(1, 21):
        m_i = MuestraConcreto(
            item_numero=i,
            codigo_muestra_lem=f"15{i:02d}-SU-26",
            identificacion_muestra=f"CALICATA C-{i:02d} (MUESTRA {i:02d})",
            procedencia=f"TRAMO KM {i*2}+000",
            cantera=f"CANTERA SAN JORGE #{i}",
            cantidad=f"{50 + i} KG",
            ensayos_json=json.dumps([
                {"codigo": f"SU{20 + (i % 5)}", "descripcion": f"ENSAYO ESPECIAL {i}", "norma": f"NTP 339.{100 + i}"},
                {"codigo": "SU20", "descripcion": "CONTENIDO DE HUMEDAD", "norma": "ASTM D2216-19"}
            ])
        )
        muestras_20.append(m_i)

    recepcion.muestras = muestras_20
    excel_bytes_20 = excel_logic.generar_excel_recepcion(recepcion)
    assert excel_bytes_20 is not None
    assert len(excel_bytes_20) > 10000

    wb_20 = openpyxl.load_workbook(io.BytesIO(excel_bytes_20), data_only=True)
    ws_20 = wb_20['RECEP. SU-AG']

    # Each sample has 2 assays, so height = max(4, 2) = 4, + 1 separator = 5 rows per sample
    # Total sample rows = 20 * 5 = 100 rows (from row 22 to 121)
    # extra_rows = 100 - 20 = 80
    expected_obs_row = 42 + 80   # 122
    expected_resp_row = 44 + 80  # 124

    # Validate each of the 20 samples
    for i in range(20):
        sample_start_r = 22 + (i * 5)
        sample_num = i + 1
        assert ws_20[f'A{sample_start_r}'].value == sample_num
        assert ws_20[f'B{sample_start_r}'].value == f"15{sample_num:02d}-SU-26"
        assert ws_20[f'C{sample_start_r}'].value == "MUESTRA:"
        assert ws_20[f'F{sample_start_r}'].value == f"CALICATA C-{sample_num:02d} (MUESTRA {sample_num:02d})"
        assert ws_20[f'C{sample_start_r+1}'].value == "PROCEDENCIA:"
        assert ws_20[f'F{sample_start_r+1}'].value == f"TRAMO KM {sample_num*2}+000"
        assert ws_20[f'C{sample_start_r+2}'].value == "CANTERA:"
        assert ws_20[f'F{sample_start_r+2}'].value == f"CANTERA SAN JORGE #{sample_num}"
        assert ws_20[f'C{sample_start_r+3}'].value == "CANTIDAD (KG):"
        assert ws_20[f'F{sample_start_r+3}'].value == f"{50 + sample_num} KG"
        
        # Check assays
        assert ws_20[f'H{sample_start_r}'].value == f"SU{20 + (sample_num % 5)}"
        assert ws_20[f'I{sample_start_r}'].value == f"ENSAYO ESPECIAL {sample_num}"
        assert ws_20[f'N{sample_start_r}'].value == f"NTP 339.{100 + sample_num}"
        assert ws_20[f'H{sample_start_r+1}'].value == "SU20"

        # Check separator row
        sep_row = sample_start_r + 4
        assert ws_20[f'A{sep_row}'].value is None
        assert ws_20[f'H{sep_row}'].value is None

    # Validate Footer at shifted position
    assert ws_20[f'A{expected_obs_row}'].value == "Observaciones:"
    assert ws_20[f'C{expected_obs_row}'].value == "Muestras selladas en sacos."
    assert ws_20[f'C{expected_resp_row}'].value == "ING. MARIO LOPEZ"
    assert ws_20[f'K{expected_resp_row}'].value == "BETZABETH SARAVIA"

    print("Test generar_excel_suelo_agregado (20 samples capacity) passed successfully!")

if __name__ == "__main__":
    test_generar_excel_suelo_agregado_fidelity()
