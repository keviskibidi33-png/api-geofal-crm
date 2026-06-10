from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.compresion_no_confinada.excel import generate_compresion_no_confinada_excel
from app.modules.compresion_no_confinada.schemas import CompresionNoConfinadaRequest


class TestCompresionNoConfinadaExcel(unittest.TestCase):
    def _build_payload(self) -> CompresionNoConfinadaRequest:
        return CompresionNoConfinadaRequest.model_validate(
            {
                "muestra": "123-26",
                "numero_ot": "456-26",
                "fecha_ensayo": "2026/06/10",
                "realizado_por": "ANDRES SANCHEZ",
                "revisado_por": "FABIAN LA ROSA",
                "aprobado_por": "IRMA COAQUIRA",
                "tara_numero": "T-01",
                "tara_suelo_humedo_g": "120,5",
                "tara_suelo_seco_g": "112,25",
                "peso_tara_g": "10,25",
                "diametro_cm": ["5,00", "5,10", "5,20"],
                "altura_cm": ["10,00", "10,10", "10,20"],
                "peso_gr": ["250,5", "251,5", "252,5"],
                "lectura_carga_kg": [f"{100 + idx},5" for idx in range(24)],
                "deformacion_tiempo": [f"00:{idx:02d}" for idx in range(24)],
                "deformacion_pulg_001": [round(idx * 0.1, 3) for idx in range(24)],
                "deformacion_mm": [round(idx * 0.25, 3) for idx in range(24)],
                "observaciones": "Ensayo con informe multihoja",
            }
        )

    def test_generate_compresion_no_confinada_excel_populates_both_sheets(self):
        payload = self._build_payload()
        excel_bytes = generate_compresion_no_confinada_excel(payload)
        workbook = load_workbook(BytesIO(excel_bytes), data_only=False)

        self.assertEqual(workbook.sheetnames, ["CNC", "CNC (2)"])
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self.assertEqual(sheet.sheet_state, "visible")
            self.assertEqual(sheet["B11"].value, payload.muestra)
            self.assertEqual(sheet["D11"].value, payload.numero_ot)
            self.assertEqual(sheet["E11"].value, payload.fecha_ensayo)
            self.assertEqual(sheet["F11"].value, payload.realizado_por)
            self.assertEqual(sheet["B52"].value, payload.observaciones)


if __name__ == "__main__":
    unittest.main()
