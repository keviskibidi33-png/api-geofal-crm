import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.common.excel_xml import find_template_path
from app.modules.sales_solubles.excel import generate_sales_solubles_excel
from app.modules.sales_solubles.schemas import SalesSolublesRequest, SalesSolublesCapsula

TEMPLATE_FILE = "1- Info N° 001-26 SU13 Sales-1 Ult.xlsx"


def _build_payload() -> SalesSolublesRequest:
    return SalesSolublesRequest(
        muestra="4567-26",
        numero_ot="1234-26",
        fecha_ensayo="2026/06/26",
        realizado_por="TEST OPERATOR",
        volumen_agua_ml=400.0,
        peso_suelo_g=120.0,
        volumen_solucion_tomada_ml=60.0,
        capsulas=[
            SalesSolublesCapsula(
                capsula_numero="C-01",
                peso_capsula_g=41.1,
                peso_capsula_sales_g=41.2,
                peso_sales_g=0.1,
                contenido_sales_ppm=1000.0
            ),
            SalesSolublesCapsula(
                capsula_numero="C-02",
                peso_capsula_g=42.1,
                peso_capsula_sales_g=42.3,
                peso_sales_g=0.2,
                contenido_sales_ppm=2000.0
            )
        ]
    )


def test_generate_sales_solubles_excel_completes_successfully():
    payload = _build_payload()
    excel_bytes = generate_sales_solubles_excel(payload)
    assert len(excel_bytes) > 0

    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    assert "FORMATO" in wb.sheetnames
    assert "SUELO" in wb.sheetnames
    assert "DATOS SUELO" in wb.sheetnames

    formato_sheet = wb["FORMATO"]
    assert formato_sheet["B10"].value == "4567-26"
    assert formato_sheet["D10"].value == "1234-26"
    assert formato_sheet["E10"].value == "2026/06/26"
    assert formato_sheet["G10"].value == "TEST OPERATOR"
    assert formato_sheet["G22"].value == 400.0
    assert formato_sheet["G23"].value == 120.0
    assert formato_sheet["G24"].value == 60.0


def test_sales_solubles_formula_integrity():
    payload = _build_payload()
    excel_bytes = generate_sales_solubles_excel(payload)

    wb_formulas = load_workbook(io.BytesIO(excel_bytes), data_only=False)
    suelo_sheet = wb_formulas["SUELO"]
    assert suelo_sheet["C11"].value == "=+FORMATO!M2"
    assert suelo_sheet["I33"].value == "=((I31*(I26/I28))/I28)*1000000"
