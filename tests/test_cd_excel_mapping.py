from __future__ import annotations

import sys
import unittest
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.cd.excel import generate_cd_excel
from app.modules.cd.schemas import CDRequest


DEF_VALUES = [
    0,
    3,
    6,
    12,
    18,
    30,
    45,
    60,
    75,
    90,
    105,
    120,
    150,
    180,
    210,
    240,
    270,
    300,
    360,
    420,
    480,
    540,
    600,
    660,
    720,
]


class TestCDExcelMapping(unittest.TestCase):
    maxDiff = None

    @staticmethod
    def _build_payload() -> CDRequest:
        return CDRequest.model_validate(
            {
                "muestra": "587-26",
                "numero_ot": "OT-589-26",
                "fecha_ensayo": "19/03/2026",
                "peso_kg": [1.11, 2.22, 3.33],
                "esf_normal": [0.5, 1.5, 2.5],
                "def_horizontal": DEF_VALUES,
                "carga_kg_1": [round(10 + idx * 0.5, 2) for idx in range(len(DEF_VALUES))],
                "carga_kg_2": [round(20 + idx * 0.5, 2) for idx in range(len(DEF_VALUES))],
                "carga_kg_3": [round(30 + idx * 0.5, 2) for idx in range(len(DEF_VALUES))],
                "humedad_puntos": [
                    {
                        "recipiente_numero": "R1",
                        "peso_recipiente_g": 10.1,
                        "peso_recipiente_suelo_humedo_g": 30.3,
                        "peso_recipiente_suelo_seco_g": 25.2,
                        "peso_agua_g": 5.1,
                        "peso_suelo_g": 15.1,
                        "contenido_humedad_pct": 33.77,
                    },
                    {
                        "recipiente_numero": "R2",
                        "peso_recipiente_g": 11.1,
                        "peso_recipiente_suelo_humedo_g": 31.3,
                        "peso_recipiente_suelo_seco_g": 26.2,
                        "peso_agua_g": 5.1,
                        "peso_suelo_g": 15.1,
                        "contenido_humedad_pct": 33.78,
                    },
                    {
                        "recipiente_numero": "R3",
                        "peso_recipiente_g": 12.1,
                        "peso_recipiente_suelo_humedo_g": 32.3,
                        "peso_recipiente_suelo_seco_g": 27.2,
                        "peso_agua_g": 5.1,
                        "peso_suelo_g": 15.1,
                        "contenido_humedad_pct": 33.79,
                    },
                ],
                "hora_1": ["08:00", "08:05", "08:10", "08:15", "08:20"],
                "deform_1": [1, 2, 3, 4, 5],
                "hora_2": ["09:00", "09:05", "09:10", "09:15", "09:20"],
                "deform_2": [6, 7, 8, 9, 10],
                "hora_3": ["10:00", "10:05", "10:10", "10:15", "10:20"],
                "deform_3": [11, 12, 13, 14, 15],
                "realizado_por": "ANDRES SANCHEZ",
                "revisado_por": "FABIAN LA ROSA",
                "aprobado_por": "IRMA COAQUIRA",
            }
        )

    @staticmethod
    def _workbook_from_payload(payload: CDRequest):
        content = generate_cd_excel(payload)
        workbook = load_workbook(BytesIO(content), data_only=False)
        return content, workbook

    def test_generate_cd_excel_injects_full_mapping(self):
        payload = self._build_payload()
        _, workbook = self._workbook_from_payload(payload)
        sheet = workbook[workbook.sheetnames[0]]

        self.assertEqual(sheet["B11"].value, payload.muestra)
        self.assertEqual(sheet["D11"].value, payload.numero_ot)
        self.assertEqual(sheet["F11"].value, payload.fecha_ensayo)

        for idx, cell in enumerate(("B16", "D16", "F16")):
            with self.subTest(section="peso_kg", cell=cell):
                self.assertEqual(sheet[cell].value, payload.peso_kg[idx])

        for idx, cell in enumerate(("B17", "D17", "F17")):
            with self.subTest(section="esf_normal", cell=cell):
                self.assertEqual(sheet[cell].value, payload.esf_normal[idx])

        carga_columns = ("C", "E", "G")
        def_columns = ("B", "D", "F")
        carga_sets = [payload.carga_kg_1, payload.carga_kg_2, payload.carga_kg_3]
        for offset, row_num in enumerate(range(20, 45)):
            expected_def = payload.def_horizontal[offset]
            for col in def_columns:
                with self.subTest(section="def_horizontal", cell=f"{col}{row_num}"):
                    self.assertEqual(sheet[f"{col}{row_num}"].value, expected_def)

            for idx, col in enumerate(carga_columns):
                with self.subTest(section="carga", cell=f"{col}{row_num}"):
                    self.assertEqual(sheet[f"{col}{row_num}"].value, carga_sets[idx][offset])

        humedad_columns = ("E", "F", "G")
        humedad_rows = {
            "recipiente_numero": 46,
            "peso_recipiente_g": 47,
            "peso_recipiente_suelo_humedo_g": 48,
            "peso_recipiente_suelo_seco_g": 49,
            "peso_agua_g": 50,
            "peso_suelo_g": 51,
            "contenido_humedad_pct": 52,
        }
        for idx, point in enumerate(payload.humedad_puntos):
            for field_name, row_num in humedad_rows.items():
                cell = f"{humedad_columns[idx]}{row_num}"
                with self.subTest(section="humedad", cell=cell):
                    self.assertEqual(sheet[cell].value, getattr(point, field_name))

        hora_columns = (("B", "C"), ("D", "E"), ("F", "G"))
        hora_sets = [payload.hora_1, payload.hora_2, payload.hora_3]
        deform_sets = [payload.deform_1, payload.deform_2, payload.deform_3]
        for offset, row_num in enumerate(range(55, 60)):
            for idx, (hora_col, deform_col) in enumerate(hora_columns):
                with self.subTest(section="hora", cell=f"{hora_col}{row_num}"):
                    self.assertEqual(sheet[f"{hora_col}{row_num}"].value, hora_sets[idx][offset])
                with self.subTest(section="deform", cell=f"{deform_col}{row_num}"):
                    self.assertEqual(sheet[f"{deform_col}{row_num}"].value, deform_sets[idx][offset])

        self.assertEqual(sheet["B62"].value, payload.realizado_por)
        self.assertEqual(sheet["D62"].value, payload.revisado_por)
        self.assertEqual(sheet["F62"].value, payload.aprobado_por)

    def test_generate_cd_excel_preserves_template_assets(self):
        payload = self._build_payload()
        content, workbook = self._workbook_from_payload(payload)
        self.assertEqual(workbook.sheetnames, ["PUS-PUC", "Hoja1"])

        with zipfile.ZipFile(BytesIO(content), "r") as archive:
            names = set(archive.namelist())

        self.assertIn("xl/worksheets/sheet1.xml", names)
        self.assertIn("xl/drawings/drawing1.xml", names)
        self.assertIn("xl/drawings/drawing2.xml", names)
        self.assertIn("xl/media/image1.jpeg", names)
        self.assertIn("xl/media/image2.png", names)


if __name__ == "__main__":
    unittest.main()
