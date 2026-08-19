import io
import sys
from pathlib import Path
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.ot.models import OrdenTrabajo
from app.modules.ot.excel import generar_excel_ot_su_ag

def test_generar_excel_ot_su_ag_standard():
    ot = OrdenTrabajo(
        numero_ot="193-26",
        numero_recepcion="193-26",
        cliente="CONSORCIO VIAL ANDINO S.A.C.",
        proyecto="MEJORAMIENTO VIA HUAROCHIRI",
        fecha_recepcion="2026/02/10",
        inicio_programado="2026/02/10",
        fin_programado="2026/02/18",
        observaciones="ENTREGAR INFORME EN FORMATO DIGITAL Y FÍSICO",
        ot_aperturada_por="BETZABETH SARAVIA",
        ot_designada_a="ERICK JULIAN",
        items=[
            {
                "item": 1,
                "codigo_muestra": "3386",
                "codigo_ensayo": "SU24",
                "descripcion": "ANÁLISIS GRANULOMÉTRICO POR TAMIZADO EN SUELOS",
                "norma": "ASTM D6913",
                "cantidad": 1,
            },
            {
                "item": 2,
                "codigo_muestra": "3386",
                "codigo_ensayo": "SU20",
                "descripcion": "CONTENIDO DE HUMEDAD DE SUELOS",
                "norma": "ASTM D2216",
                "cantidad": 1,
            },
            {
                "item": 3,
                "codigo_muestra": "3387",
                "codigo_ensayo": "SU24",
                "descripcion": "ANÁLISIS GRANULOMÉTRICO POR TAMIZADO EN SUELOS",
                "norma": "ASTM D6913",
                "cantidad": 2,
            },
        ],
    )

    buf = generar_excel_ot_su_ag(ot)
    assert buf is not None
    buf.seek(0)

    wb = openpyxl.load_workbook(buf)
    assert "HOJA 1 (2)" in wb.sheetnames
    ws = wb["HOJA 1 (2)"]

    # Verify Header
    assert ws["C6"].value == "OT-193-26"
    assert ws["G6"].value == "193-26"

    # Verify Items
    assert ws["A9"].value == 1
    assert str(ws["B9"].value) == "3386"
    assert ws["D9"].value == "SU24"
    assert "GRANULOMÉTRICO" in str(ws["E9"].value).upper()
    assert ws["H9"].value == "ASTM D6913"
    assert ws["I9"].value == 1

    assert ws["A10"].value == 2
    assert ws["D10"].value == "SU20"
    assert ws["H10"].value == "ASTM D2216"

    assert ws["A11"].value == 3
    assert str(ws["B11"].value) == "3387"
    assert ws["I11"].value == 2

    # Verify Footer
    assert ws["E32"].value == "2026/02/10"
    assert ws["G32"].value == "2026/02/10"
    assert ws["I32"].value == "2026/02/18"
    assert "DIGITAL" in str(ws["B33"].value).upper()
    assert ws["D36"].value == "BETZABETH SARAVIA"
    assert ws["H36"].value == "ERICK JULIAN"

    print("Test generar_excel_ot_su_ag passed successfully!")

if __name__ == "__main__":
    test_generar_excel_ot_su_ag_standard()
