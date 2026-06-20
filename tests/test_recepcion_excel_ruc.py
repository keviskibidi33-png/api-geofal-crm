from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook
from lxml import etree

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.modules.recepcion.excel import ExcelLogic
from app.modules.recepcion.excel import _detect_recepcion_layout


class TestRecepcionExcelRuc(unittest.TestCase):
    def setUp(self):
        self.logic = ExcelLogic()

    def _parse_workbook(self, workbook: Workbook) -> dict:
        buffer = io.BytesIO()
        workbook.save(buffer)
        return self.logic.parsear_recepcion(buffer.getvalue())

    def test_discards_address_when_ruc_fallback_points_to_domicilio(self):
        workbook = Workbook()
        sheet = workbook.active

        sheet["B5"] = "RECEPCIÓN N°"
        sheet["D5"] = "REC-100-26"
        sheet["B6"] = "COTIZACIÓN N°"
        sheet["D6"] = "COT-100"
        sheet["B7"] = "OT N°"
        sheet["D7"] = "OT-100"
        sheet["B10"] = "CLIENTE:"
        sheet["D10"] = "CLIENTE TEST"
        sheet["B11"] = "DOMICILIO LEGAL:"
        sheet["D11"] = "AV. COMANDANTE ESPINAR 860 - MIRAFLORES"
        sheet["D12"] = "AV. COMANDANTE ESPINAR 860 - MIRAFLORES"

        parsed = self._parse_workbook(workbook)

        self.assertEqual(parsed["domicilio_legal"], "AV. COMANDANTE ESPINAR 860 - MIRAFLORES")
        self.assertEqual(parsed["ruc"], "")

    def test_keeps_numeric_ruc_when_fallback_cell_is_valid(self):
        workbook = Workbook()
        sheet = workbook.active

        sheet["B5"] = "RECEPCIÓN N°"
        sheet["D5"] = "REC-101-26"
        sheet["B6"] = "COTIZACIÓN N°"
        sheet["D6"] = "COT-101"
        sheet["B7"] = "OT N°"
        sheet["D7"] = "OT-101"
        sheet["B10"] = "CLIENTE:"
        sheet["D10"] = "CLIENTE TEST"
        sheet["B11"] = "DOMICILIO LEGAL:"
        sheet["D11"] = "AV. JAVIER PRADO 123 - SAN ISIDRO"
        sheet["D12"] = "20505212739"

        parsed = self._parse_workbook(workbook)

        self.assertEqual(parsed["ruc"], "20505212739")

    def test_repeated_document_numbers_must_match(self):
        workbook = Workbook()
        sheet = workbook.active

        sheet["B5"] = "RECEPCIÓN N°"
        sheet["D5"] = "REC-123-26"
        sheet["B6"] = "COTIZACIÓN N°"
        sheet["D6"] = "COT-555"
        sheet["B7"] = "OT N°"
        sheet["D7"] = "OT-777"

        # Repeat the same numbers elsewhere in the document body
        sheet["B20"] = "RECEPCIÓN N°"
        sheet["D20"] = "REC-123-26"
        sheet["B21"] = "COTIZACIÓN N°"
        sheet["D21"] = "COT-555"
        sheet["B22"] = "OT N°"
        sheet["D22"] = "OT-777"

        parsed = self._parse_workbook(workbook)

        self.assertEqual(parsed["numero_recepcion"], "REC-123-26")
        self.assertEqual(parsed["numero_cotizacion"], "COT-555")
        self.assertEqual(parsed["numero_ot"], "OT-777")

    def test_repeated_document_numbers_raise_when_values_differ(self):
        workbook = Workbook()
        sheet = workbook.active

        sheet["B5"] = "RECEPCIÓN N°"
        sheet["D5"] = "REC-123-26"
        sheet["B6"] = "COTIZACIÓN N°"
        sheet["D6"] = "COT-555"
        sheet["B7"] = "OT N°"
        sheet["D7"] = "OT-777"

        sheet["B20"] = "OT N°"
        sheet["D20"] = "OT-778"

        with self.assertRaisesRegex(ValueError, "Inconsistencia en OT N°"):
            self._parse_workbook(workbook)

    def test_missing_cotizacion_is_allowed_but_recepcion_and_ot_parse(self):
        workbook = Workbook()
        sheet = workbook.active

        sheet["B5"] = "RECEPCIÓN N°"
        sheet["D5"] = "REC-200-26"
        sheet["B7"] = "OT N°"
        sheet["D7"] = "OT-200"

        parsed = self._parse_workbook(workbook)

        self.assertEqual(parsed["numero_recepcion"], "REC-200-26")
        self.assertEqual(parsed["numero_cotizacion"], "")
        self.assertEqual(parsed["numero_ot"], "OT-200")

    def test_detect_recepcion_layout_uses_anchor_positions(self):
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        sheet = etree.fromstring(
            f"""
            <worksheet xmlns="{ns}">
              <sheetData>
                <row r="21">
                  <c r="A21" t="s"><v>0</v></c>
                </row>
                <row r="43">
                  <c r="B43" t="s"><v>1</v></c>
                </row>
              </sheetData>
              <mergeCells count="42">
                <mergeCell ref="A1:B1"/>
              </mergeCells>
            </worksheet>
            """.strip().encode("utf-8")
        )
        shared_strings = ["N°", "NOTA:"]

        layout = _detect_recepcion_layout(sheet.find(f".//{{{ns}}}sheetData"), shared_strings, ns)

        self.assertEqual(layout["table_header_row"], 21)
        self.assertEqual(layout["data_start_row"], 23)
        self.assertEqual(layout["footer_row"], 43)
        self.assertEqual(layout["available_rows_before_footer"], 20)
        self.assertEqual(layout["variant"], "canonical")


if __name__ == "__main__":
    unittest.main()
