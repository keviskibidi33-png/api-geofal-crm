import io
import sys
import zipfile
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.planas.excel import TEMPLATE_PATH, generate_planas_excel
from app.modules.planas.schemas import PlanasRequest, PlanasGradacionRow, PlanasMetodoRow


def test_planas_excel_template_integrity_and_generation():
    assert Path(TEMPLATE_PATH).exists(), f"Template not found at: {TEMPLATE_PATH}"

    payload = PlanasRequest(
        muestra="587-SU-26",
        numero_ot="1000-26",
        fecha_ensayo="2026/06/16",
        realizado_por="OPERADOR TEST",
        relacion_dimensional="1:3",
        metodo_ensayo="A",
        tamiz_requerido="3/8 in.",
        masa_inicial_g=2500.0,
        masa_seca_g=2480.0,
        masa_seca_constante_g=2480.0,
        gradacion_rows=[
            PlanasGradacionRow(
                pasa_tamiz="2 in.",
                retenido_tamiz="1 1/2 in.",
                masa_retenido_original_g=100.0,
                porcentaje_retenido=4.0,
                criterio_acepta=False,
                numero_particulas_aprox_100=0,
                masa_retenido_g=0.0
            ),
            PlanasGradacionRow(
                pasa_tamiz="1 1/2 in.",
                retenido_tamiz="1 in.",
                masa_retenido_original_g=500.0,
                porcentaje_retenido=20.0,
                criterio_acepta=True,
                numero_particulas_aprox_100=100,
                masa_retenido_g=490.0
            )
        ],
        metodo_rows=[
            PlanasMetodoRow(
                retenido_tamiz="1 1/2 in.",
                grupo1_numero_particulas=10,
                grupo1_masa_g=50.0,
                grupo2_numero_particulas=5,
                grupo2_masa_g=25.0,
                grupo3_numero_particulas=2,
                grupo3_masa_g=10.0,
                grupo4_numero_particulas=83,
                grupo4_masa_g=405.0
            )
        ],
        dispositivo_calibre_codigo="EQP-0044",
        balanza_01g_codigo="EQP-0046",
        horno_codigo="EQP-0150",
        nota="Nota de prueba",
        revisado_por="REVISOR TEST",
        revisado_fecha="2026/06/17",
        aprobado_por="APROBADOR TEST",
        aprobado_fecha="2026/06/18"
    )

    # Generate the excel
    excel_bytes = generate_planas_excel(payload)
    assert isinstance(excel_bytes, bytes)

    # Load and check sheets
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    assert "FORMATO" in wb.sheetnames
    assert "INCERTIDUMBRE" in wb.sheetnames

    ws_formato = wb["FORMATO"]
    assert ws_formato["D11"].value == "587-SU-26"
    assert ws_formato["G11"].value == "1000-26"
    assert ws_formato["J11"].value == "2026/06/16"
    assert ws_formato["M11"].value == "OPERADOR TEST"

    # Check relation checkbox
    assert ws_formato["C19"].value == "X"
    assert ws_formato["B19"].value == ""
    assert ws_formato["D19"].value == ""

    # Check uncertainty sheet
    ws_incert = wb["INCERTIDUMBRE"]
    assert ws_incert["C126"].value == "REVISOR TEST"
    assert ws_incert["H126"].value == "APROBADOR TEST"

    # Zip check
    with zipfile.ZipFile(io.BytesIO(excel_bytes), "r") as archive:
        names = archive.namelist()
        assert "xl/calcChain.xml" not in names
