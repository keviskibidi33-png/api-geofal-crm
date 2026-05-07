import io
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from lxml import etree

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.equi_arena.excel import TEMPLATE_PATH, generate_equi_arena_excel
from app.modules.equi_arena.schemas import EquiArenaRequest

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _build_payload(**overrides) -> EquiArenaRequest:
    payload = {
        "muestra": "1234-SU-26",
        "numero_ot": "4567-26",
        "fecha_ensayo": "21/03/26",
        "realizado_por": "OPERADOR",
        "tipo_muestra": "SUELO",
        "metodo_agitacion": "MANUAL",
        "preparacion_muestra": "PROCEDIMIENTO A",
        "temperatura_solucion_c": 23,
        "masa_4_medidas_g": 85.5,
        "cronometro_entrada_saturacion_hmin": ["08:00", "08:10", "08:20"],
        "cronometro_salida_saturacion_hmin": ["08:10", "08:20", "08:30"],
        "tiempo_saturacion_min": [10, 10, 10],
        "tiempo_agitacion_seg": [45, 45, 45],
        "cronometro_entrada_decantacion_hmin": ["08:15", "08:25", "08:35"],
        "cronometro_salida_decantacion_hmin": ["08:35", "08:45", "08:55"],
        "tiempo_decantacion_min": [20, 20, 20],
        "lectura_arcilla_in": [10, 10, 10],
        "lectura_arena_in": [5, 5, 5],
        "equivalente_arena_promedio_pct": 999,
    }
    payload.update(overrides)
    return EquiArenaRequest(**payload)


def _font_color_signature(cell) -> tuple[str, object, object | None] | None:
    color = cell.font.color
    if color is None or color.type is None:
        return None

    if color.type == "rgb":
        return ("rgb", color.rgb, None)
    if color.type == "theme":
        return ("theme", color.theme, color.tint)
    if color.type == "indexed":
        return ("indexed", color.indexed, None)
    if color.type == "auto":
        return ("auto", color.auto, None)
    return (str(color.type), None, None)


def test_equi_arena_request_ignora_promedio_del_cliente():
    payload = _build_payload()

    assert payload.equivalente_arena_promedio_pct == 50.0


def test_generate_equi_arena_excel_escribe_resultados():
    payload = _build_payload()

    workbook = load_workbook(io.BytesIO(generate_equi_arena_excel(payload)), data_only=False)
    sheet = workbook.active

    assert sheet["B17"].value == "SUELO"
    assert sheet["F17"].value == "MANUAL"
    assert sheet["F21"].value == "PROCEDIMIENTO A"

    assert sheet["H26"].value == "08:00"
    assert sheet["I26"].value == "08:10"
    assert sheet["J26"].value == "08:20"
    assert sheet["H27"].value == "08:10"
    assert sheet["I27"].value == "08:20"
    assert sheet["J27"].value == "08:30"
    assert sheet["H30"].value == "08:15"
    assert sheet["I30"].value == "08:25"
    assert sheet["J30"].value == "08:35"
    assert sheet["H31"].value == "08:35"
    assert sheet["I31"].value == "08:45"
    assert sheet["J31"].value == "08:55"

    assert sheet["H35"].value == 50
    assert sheet["I35"].value == 50
    assert sheet["J35"].value == 50
    assert sheet["H36"].value == 50
    assert sheet["H35"].protection.locked is True
    assert sheet["I35"].protection.locked is True
    assert sheet["J35"].protection.locked is True
    assert sheet["H36"].protection.locked is True
    assert sheet["B11"].protection.locked is False
    assert sheet["F17"].protection.locked is False

    with zipfile.ZipFile(io.BytesIO(generate_equi_arena_excel(payload)), "r") as workbook_zip:
        sheet_root = etree.fromstring(workbook_zip.read("xl/worksheets/sheet1.xml"))
        assert sheet_root.find(f".//{{{NS_MAIN}}}sheetProtection") is not None


def test_generate_equi_arena_excel_no_deja_calc_chain_colgante():
    payload = _build_payload()

    with zipfile.ZipFile(io.BytesIO(generate_equi_arena_excel(payload)), "r") as workbook_zip:
        assert "xl/calcChain.xml" not in workbook_zip.namelist()

        rels_root = etree.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        rel_targets = [rel.get("Target", "") for rel in rels_root]
        assert all(not target.endswith("calcChain.xml") for target in rel_targets)

        content_types_root = etree.fromstring(workbook_zip.read("[Content_Types].xml"))
        overrides = [override.get("PartName", "") for override in content_types_root]
        assert "/xl/calcChain.xml" not in overrides


def test_generate_equi_arena_excel_preserva_colores_de_fuente_del_template():
    payload = _build_payload()

    template_sheet = load_workbook(TEMPLATE_PATH, data_only=False).active
    generated_sheet = load_workbook(io.BytesIO(generate_equi_arena_excel(payload)), data_only=False).active

    colored_cells = [
        (cell.coordinate, signature)
        for row in template_sheet.iter_rows()
        for cell in row
        for signature in [_font_color_signature(cell)]
        if signature is not None
    ]

    assert colored_cells, "El template de EquiArena no tiene celdas con color de fuente para validar."

    for coordinate, expected_signature in colored_cells:
        assert _font_color_signature(generated_sheet[coordinate]) == expected_signature
