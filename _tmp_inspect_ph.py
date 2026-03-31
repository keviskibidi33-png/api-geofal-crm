from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')
ws = wb.active

print('=== HEADER DATA CELLS ===')
print(f'B11 (muestra): {ws["B11"].value}')
print(f'D11 (OT): {ws["D11"].value}')
print(f'E11 (fecha): {ws["E11"].value}')
print(f'G11 (realizado): {ws["G11"].value}')

print('\n=== CONDICIONES SECADO ===')
print(f'D17 (secado aire): {ws["D17"].value}')
print(f'D18 (secado horno 60C): {ws["D18"].value}')

print('\n=== RESULTADOS ===')
print(f'D24 (temperatura): {ws["D24"].value}')
print(f'D25 (ph): {ws["D25"].value}')

print('\n=== EQUIPOS ===')
print(f'E36 (horno): {ws["E36"].value}')
print(f'E37 (balanza): {ws["E37"].value}')
print(f'E38 (phmetro): {ws["E38"].value}')

print('\n=== FOOTER ===')
print(f'B48 (revisado): {ws["B48"].value}')
print(f'E48 (aprobado): {ws["E48"].value}')

print('\n=== OBSERVACIONES ===')
print(f'A32 (observaciones): {ws["A32"].value}')
