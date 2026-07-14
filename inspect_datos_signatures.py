import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb["Datos"]

print("=== Datos sheet Q10:R15 ===")
for r in range(10, 16):
    print(f"Row {r}: Q{r}={repr(ws.cell(row=r, column=17).value)} | R{r}={repr(ws.cell(row=r, column=18).value)}")
