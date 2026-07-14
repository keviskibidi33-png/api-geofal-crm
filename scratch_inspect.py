import openpyxl
import sys

# Reconfigure stdout to use utf-8
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx", data_only=False)

print("SHEET NAMES:")
print(wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n--- SHEET: {sheetname} ({ws.max_row} rows, {ws.max_column} cols) ---")
    
    # We want to find cells with text or formulas
    for row in range(1, min(ws.max_row + 1, 100)): # check first 100 rows
        for col in range(1, min(ws.max_column + 1, 26)): # check first 26 columns (A to Z)
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None:
                cell_coord = f"{openpyxl.utils.get_column_letter(col)}{row}"
                # If it's a formula, print it. If it is a label, print it.
                if isinstance(val, str) and val.startswith("="):
                    print(f"  Formula at {cell_coord}: {val}")
                elif isinstance(val, str) and any(x in val.lower() for x in ["ot", "muestra", "fecha", "humedad", "densidad", "peso", "volumen", "arena", "cono", "operador", "realizado", "revisado", "aprobado", "progresiva", "capa", "cota", "lado", "mximo", "mínimo", "óptimo"]):
                    print(f"  Label/Val at {cell_coord}: {val!r}")
                elif not isinstance(val, str):
                    print(f"  Value at {cell_coord}: {val}")
