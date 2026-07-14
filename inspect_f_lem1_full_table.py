import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb["F LEM1"]

print("=== Search in F LEM1 ===")
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value or "")
        if "compacta" in val.lower() or "peso" in val.lower() or "humedad" in val.lower() or "densidad" in val.lower():
            print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: value={repr(cell.value)}")
