import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb["F LEM2"]

print("=== Search 'criterio' in F LEM2 ===")
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value or "")
        if "criterio" in val.lower() or "aceptación" in val.lower():
            print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: value={repr(cell.value)}")
            # print surrounding cells
            row_str = f"Row {r} cells: "
            for col in range(1, 12):
                v = ws.cell(row=r, column=col).value
                row_str += f"{openpyxl.utils.get_column_letter(col)}{r}: {repr(v)} | "
            print(row_str)
