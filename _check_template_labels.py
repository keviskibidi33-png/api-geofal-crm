from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')
ws = wb.worksheets[1]

print("=== TEMPLATE LABELS IN CONDICIONES AREA ===\n")
print("Row 17 (labels):")
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    val = ws[f'{col}17'].value
    if val:
        print(f"  {col}17: {val}")

print("\nRow 18 (labels):")
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    val = ws[f'{col}18'].value
    if val:
        print(f"  {col}18: {val}")
