"""Test PH Excel generation with sample data"""
import sys
sys.path.insert(0, 'c:/Users/Lenovo/Documents/crmnew/api-geofal-crm')

from app.modules.ph.schemas import PHRequest
from app.modules.ph.excel import generate_ph_excel

# Sample payload
payload = PHRequest(
    muestra="1234-AG-26",
    numero_ot="5678-26",
    fecha_ensayo="15/03/26",
    realizado_por="Juan Perez",
    condicion_secado_aire="SI",
    condicion_secado_horno="NO",
    temperatura_ensayo_c=25.5,
    ph_resultado=7.2,
    equipo_horno_codigo="EQP-0049",
    equipo_balanza_001_codigo="EQP-0046",
    equipo_ph_metro_codigo="EQP-0050",
    observaciones="Prueba de generación",
    revisado_por="FABIAN LA ROSA",
    revisado_fecha="16/03/26",
    aprobado_por="IRMA COAQUIRA",
    aprobado_fecha="17/03/26"
)

print("Generating Excel with payload...")
print(f"Muestra: {payload.muestra}")
print(f"Temperatura: {payload.temperatura_ensayo_c}")
print(f"PH: {payload.ph_resultado}")

try:
    excel_bytes = generate_ph_excel(payload)
    print(f"\n✓ Excel generated successfully: {len(excel_bytes)} bytes")
    
    # Save to file
    with open("_test_ph_output.xlsx", "wb") as f:
        f.write(excel_bytes)
    print("✓ Saved to _test_ph_output.xlsx")
    
    # Verify cells using openpyxl
    from openpyxl import load_workbook
    wb = load_workbook("_test_ph_output.xlsx")
    ws = wb.active
    
    print("\n=== VERIFICATION ===")
    print(f"B11 (muestra): {ws['B11'].value}")
    print(f"D11 (OT): {ws['D11'].value}")
    print(f"E11 (fecha): {ws['E11'].value}")
    print(f"G11 (realizado): {ws['G11'].value}")
    print(f"F17 (secado aire): {ws['F17'].value}")
    print(f"F18 (secado horno): {ws['F18'].value}")
    print(f"F24 (temperatura): {ws['F24'].value}")
    print(f"F25 (ph): {ws['F25'].value}")
    print(f"E36 (horno): {ws['E36'].value}")
    print(f"E37 (balanza): {ws['E37'].value}")
    print(f"E38 (phmetro): {ws['E38'].value}")
    
except Exception as e:
    print(f"\n✗ Error: {e}")
    import traceback
    traceback.print_exc()
