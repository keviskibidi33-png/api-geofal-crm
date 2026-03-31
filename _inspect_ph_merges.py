from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')

print("=== ALL SHEETS ===")
for sheet_name in wb.sheetnames:
    print(f"  {sheet_name}")

# Get sheet by index (sheet2 = index 1)
ws = wb.worksheets[1]
print(f"\n=== WORKING WITH SHEET: {ws.title} ===")

print("\n=== MERGED CELLS IN CONDICIONES AREA (rows 16-18) ===")
for mc in ws.merged_cells:
    min_row, min_col, max_row, max_col = mc.bounds
    if min_row >= 16 and max_row <= 18:
        start = ws.cell(min_row, min_col).coordinate
        end = ws.cell(max_row, max_col).coordinate
        print(f"{start}:{end} (rows {min_row}-{max_row})")

print("\n=== MERGED CELLS IN RESULTADOS AREA (rows 23-25) ===")
for mc in ws.merged_cells:
    min_row, min_col, max_row, max_col = mc.bounds
    if min_row >= 23 and max_row <= 25:
        start = ws.cell(min_row, min_col).coordinate
        end = ws.cell(max_row, max_col).coordinate
        print(f"{start}:{end} (rows {min_row}-{max_row})")

print("\n=== CELL VALUES IN ROWS 17-18 ===")
for row in [17, 18]:
    for col in ['C', 'D', 'E', 'F', 'G']:
        val = ws[f'{col}{row}'].value
        if val:
            print(f"{col}{row}: {val}")

print("\n=== CELL VALUES IN ROWS 24-25 ===")
for row in [24, 25]:
    for col in ['C', 'D', 'E', 'F', 'G']:
        val = ws[f'{col}{row}'].value
        if val:
            print(f"{col}{row}: {val}")
