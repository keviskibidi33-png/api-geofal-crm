import openpyxl

def detail_cells(filepath, sheetname):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb[sheetname]
    print(f"\n================= {filepath} -> {sheetname} =================")
    for r in range(20, 45):
        row_str = []
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            ref = f"{openpyxl.utils.get_column_letter(c)}{r}"
            val = cell.value
            if val is not None or cell.coordinate in ws.merged_cells:
                row_str.append(f"{ref}: {repr(val)}")
            elif cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != '00000000':
                # even if None, print it if it has fill/border/style
                row_str.append(f"{ref}: None (styled)")
        if row_str:
            print(f"Row {r}: " + " | ".join(row_str))

detail_cells('app/templates/Subir auditoria/1- Inf N° 001-26 SU14 Cloruros-1.xlsx', 'FORMATO')
detail_cells('app/templates/Subir auditoria/1- Inf N° 001-26 SU03 PH-1..xlsx', 'FORMATO')
