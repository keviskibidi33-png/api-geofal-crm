import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx")
print("Sheet Names:", wb.sheetnames)

# Let's inspect "Datos" sheet cells in detail
print("\n=== DATOS SHEET (first 45 rows, all cols) ===")
ws = wb["Datos"]
for r in range(1, 45):
    row_vals = []
    for c in range(1, 27):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_vals.append(f"{col_letter}{r}: {repr(val)}")
    if row_vals:
        print(f"Row {r}: " + " | ".join(row_vals))

print("\n=== F LEM1 SHEET (first 40 rows) ===")
if "F LEM1" in wb.sheetnames:
    ws = wb["F LEM1"]
    for r in range(1, 40):
        row_vals = []
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                row_vals.append(f"{col_letter}{r}: {repr(val)}")
        if row_vals:
            print(f"Row {r}: " + " | ".join(row_vals))

print("\n=== F LEM2 SHEET (first 40 rows) ===")
if "F LEM2" in wb.sheetnames:
    ws = wb["F LEM2"]
    for r in range(1, 40):
        row_vals = []
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                row_vals.append(f"{col_letter}{r}: {repr(val)}")
        if row_vals:
            print(f"Row {r}: " + " | ".join(row_vals))
