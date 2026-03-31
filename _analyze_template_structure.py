from lxml import etree
import zipfile
from openpyxl import load_workbook

# Check cell structure
wb = load_workbook('app/templates/Template_PH.xlsx')
ws = wb.worksheets[1]

print("=== CELL VALUES IN TEMPLATE ===")
print("\nRow 16 (header?):")
for col in ['C', 'D', 'E', 'F', 'G']:
    val = ws[f'{col}16'].value
    if val:
        print(f"  {col}16: {val}")

print("\nRow 17:")
for col in ['C', 'D', 'E', 'F', 'G']:
    val = ws[f'{col}17'].value
    if val:
        print(f"  {col}17: {val}")

print("\nRow 18:")
for col in ['C', 'D', 'E', 'F', 'G']:
    val = ws[f'{col}18'].value
    if val:
        print(f"  {col}18: {val}")

# Check shapes
with zipfile.ZipFile('app/templates/Template_PH.xlsx', 'r') as z:
    drawing_xml = z.read('xl/drawings/drawing2.xml')

root = etree.fromstring(drawing_xml)
ns = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
}

print("\n=== SHAPES IN ROWS 16-18 ===")
for anchor in root.findall('.//xdr:twoCellAnchor', ns):
    from_elem = anchor.find('.//xdr:from', ns)
    if from_elem is None:
        continue
    
    from_row = int(from_elem.find('xdr:row', ns).text)
    from_col = int(from_elem.find('xdr:col', ns).text)
    
    to_elem = anchor.find('.//xdr:to', ns)
    to_col = int(to_elem.find('xdr:col', ns).text) if to_elem is not None else None
    
    if from_row == 16:
        txBody = anchor.find('.//xdr:txBody', ns)
        texts = []
        if txBody is not None:
            for p in txBody.findall('.//a:p', ns):
                for t in p.findall('.//a:t', ns):
                    if t.text:
                        texts.append(t.text)
        
        print(f"\nShape from col {from_col} to col {to_col}, row {from_row}")
        print(f"  Current text: {texts if texts else 'Empty'}")
        
        # Map column numbers to letters
        col_letters = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G', 7: 'H'}
        print(f"  Covers cells: {col_letters.get(from_col, '?')} to {col_letters.get(to_col, '?')}")
