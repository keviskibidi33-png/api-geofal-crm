import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb["F LEM1"]

print("=== F LEM1 Rows 25 to 30 ===")
for r in range(25, 31):
    row_str = f"Row {r}: "
    for c in range(1, 11):
        cell = ws.cell(row=r, column=c)
        col_letter = openpyxl.utils.get_column_letter(c)
        row_str += f"{col_letter}{r}:{repr(cell.value)} | "
    print(row_str)
