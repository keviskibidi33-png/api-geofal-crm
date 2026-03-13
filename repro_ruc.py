
import os
import sys
from openpyxl import Workbook

# Mock the ExcelLogic class to use the actual implementation from excel.py
# We achieve this by importing the module, but we need to ensure path is correct
sys.path.append(os.getcwd())
from app.modules.recepcion.excel import ExcelLogic

def create_mock_excel_ruc_false_positive():
    wb = Workbook()
    ws = wb.active
    
    # Simulate a header row that contains "ESTRUCTURA" which has "RUC" inside it
    # And a separate "RUC" field
    
    # Row 10: Metadata
    ws['B10'] = "CLIENTE:"
    ws['C10'] = "MY CLIENT"
    
    ws['E10'] = "RUC:"
    ws['F10'] = "20505212739"
    
    # Row 20: Table Header
    ws['B20'] = "ITEM"
    ws['C20'] = "DESCRIPCION"
    ws['D20'] = "ESTRUCTURA"  # This contains "RUC" substring!
    ws['E20'] = "F'C (KG/CM2)"
    
    filename = "test_ruc_fox.xlsx"
    wb.save(filename)
    return filename

def test_ruc_extraction():
    print("--- TEST RUC FALSE POSITIVE ---")
    filename = create_mock_excel_ruc_false_positive()
    
    with open(filename, "rb") as f:
        content = f.read()
        
    logic = ExcelLogic()
    data = logic.parsear_recepcion(content)
    
    ruc = data.get("ruc")
    print(f"Extracted RUC: '{ruc}'")
    
    if ruc == "F'C (KG/CM2)":
        print(">> FAIL: RUC matched ESTRUCTURA header!")
        sys.exit(1)
    elif ruc == "20505212739":
        print(">> PASS: RUC Correctly Extracted")
    else:
        print(f">> FAIL: RUC extracted unknown value: {ruc}")
        sys.exit(1)
        
    # clean up
    try:
        os.remove(filename)
    except: pass

if __name__ == "__main__":
    test_ruc_extraction()
