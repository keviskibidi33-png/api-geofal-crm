import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx")
ws = wb["Datos"]

print("=== CELL BY CELL MAPPING FOR 'Datos' SHEET ===")
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            coord = f"{col_letter}{r}"
            if isinstance(val, str) and val.startswith("="):
                print(f"  {coord} (Formula): {val}")
            else:
                # If it's a value and not a label (usually numbers or simple strings that aren't titles)
                if not isinstance(val, str) or (len(val) < 50 and not val.isupper()):
                    print(f"  {coord} (Value): {val!r}")
