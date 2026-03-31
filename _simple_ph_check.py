from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')
ws = wb.worksheets[1]  # PH sheet

print("=== ROW 17 ===")
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
    val = ws[f'{col_letter}17'].value
    print(f"{col_letter}17: {val}")

print("\n=== ROW 18 ===")
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
    val = ws[f'{col_letter}18'].value
    print(f"{col_letter}18: {val}")

print("\n=== ROW 24 ===")
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
    val = ws[f'{col_letter}24'].value
    print(f"{col_letter}24: {val}")

print("\n=== ROW 25 ===")
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
    val = ws[f'{col_letter}25'].value
    print(f"{col_letter}25: {val}")

print("\n=== MERGED CELLS (all) ===")
for mc in ws.merged_cells:
    print(mc)
