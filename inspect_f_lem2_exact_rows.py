import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb["F LEM2"]

print("=== F LEM2 Rows 35 to 45 (All Cells) ===")
for r in range(35, ws.max_row + 1):
    row_str = f"Row {r}: "
    for c in range(1, 16):
        cell = ws.cell(row=r, column=c)
        col_letter = openpyxl.utils.get_column_letter(c)
        if cell.value is not None:
            row_str += f"{col_letter}{r}:{repr(cell.value)} | "
    print(row_str)
