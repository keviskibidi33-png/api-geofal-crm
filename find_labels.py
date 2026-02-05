import openpyxl
from pathlib import Path

LABELS_TO_FIND = [
    "CLIENTE:", "R.U.C", "CONTACTO:", "TELÉFONO DE CONTACTO:", "CORREO:",
    "PROYECTO", "UBICACIÓN", "PERSONAL COMERCIAL", "TELÉFONO DE COMERCIAL",
    "FECHA SOLICITUD", "FECHA DE EMISIÓN:", "COTIZACIÓN",
    "CÓDIGO", "DESCRIPCIÓN", "NORMA", "ACREDITADO", "COSTO UNITARIO", "CANTIDAD", "COSTO PARCIAL"
]

def find_labels_in_sheet(template_path, sheet_name):
    print(f"\nSearching labels in: {sheet_name}")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        print("Sheet not found")
        return
    
    ws = wb[sheet_name]
    found = {}
    for r in range(1, 100):
        for c in range(1, 30):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str):
                val_upper = val.strip().upper()
                for l in LABELS_TO_FIND:
                    if l.upper() in val_upper:
                        found[l] = f"{openpyxl.utils.cell.get_column_letter(c)}{r}"
    
    for l, addr in found.items():
        print(f"'{l}' found at {addr}")

if __name__ == "__main__":
    template = Path("c:/Users/Lenovo/Documents/crmnew/api-geofal-crm/app/templates/Temp_Cotizacion.xlsx")
    if template.exists():
        find_labels_in_sheet(template, "MORT2")
    else:
        print("Template not found")
