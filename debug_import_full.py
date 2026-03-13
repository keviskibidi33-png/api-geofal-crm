
import os
import sys
from openpyxl import Workbook
from datetime import datetime

# Setup path
sys.path.append(os.getcwd())
from app.modules.recepcion.excel import ExcelLogic

def create_mock_excel_user_scenario():
    wb = Workbook()
    ws = wb.active
    
    # Simulate Header Layout based on screenshots
    # Row 6: Recepcion N
    ws['A6'] = "RECEPCIÓN N°:"
    ws['B6'] = "193-26"
    
    # Row 8-14: Client Info
    ws['B10'] = "CLIENTE:"
    ws['C10'] = ":" # Simulate the issue: cell has just colon?
    ws['D10'] = "MY ACTUAL CLIENT" # Value to the right
    
    ws['E10'] = "RUC:"
    ws['F10'] = "20505212739"
    
    ws['B11'] = "DOMICILIO LEGAL:"
    ws['C11'] = "AV. JAVIER PRADO..."
    
    ws['B13'] = "PERSONA CONTACTO:"
    ws['C13'] = "JUAN PEREZ"
    
    # Row 20-22: Table Header
    ws['B20'] = "ITEM"
    ws['C20'] = "CODIGO LEM"
    ws['D20'] = "CODIGO"
    ws['E20'] = "ESTRUCTURA"
    ws['F20'] = "F'C"
    ws['G20'] = "FECHA MOLDEO"
    
    # Row 23: Data
    ws['B21'] = 1
    ws['C21'] = "1483"
    ws['D21'] = "BD C62 (2X1)"
    ws['E21'] = "BANCODUCTO"
    ws['F21'] = "" # Empty
    ws['G21'] = datetime(2026, 2, 18)
    
    filename = "test_user_scenario.xlsx"
    wb.save(filename)
    return filename

def test_full_extraction():
    print("--- TEST USER SCENARIO ---")
    filename = create_mock_excel_user_scenario()
    
    with open(filename, "rb") as f:
        content = f.read()
        
    logic = ExcelLogic()
    data = logic.parsear_recepcion(content)
    
    print("\n[DATA EXTRACTED]")
    for k, v in data.items():
        if k == 'muestras':
            print(f"Muestras ({len(v)} found):")
            for m in v:
                print(f"  - {m}")
        else:
            print(f"{k}: '{v}'")

    # validation
    if data.get('cliente') == ":":
        print("\n>> FAIL: Client extracted as ':'")
    elif data.get('cliente') == "MY ACTUAL CLIENT":
        print("\n>> PASS: Client extracted correctly")
    else:
        print(f"\n>> FAIL: Client extracted unknown: '{data.get('cliente')}'")
        
    if len(data.get('muestras', [])) > 0:
        print(">> PASS: Muestras found")
    else:
        print(">> FAIL: No muestras found")

    try:
        os.remove(filename)
    except: pass

if __name__ == "__main__":
    test_full_extraction()
