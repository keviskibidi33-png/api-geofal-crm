import openpyxl
from pathlib import Path

def dump_sheet_structure(template_path, sheet_name):
    print(f"\n--- Structure for sheet: {sheet_name} ---")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        print("Sheet not found")
        return
    
    ws = wb[sheet_name]
    # Check header area
    for r in range(1, 15):
        row_vals = []
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            if val:
                row_vals.append(f"[{openpyxl.utils.cell.get_column_letter(c)}{r}]: '{val}'")
        if row_vals:
            print(" | ".join(row_vals))

if __name__ == "__main__":
    template = Path("c:/Users/Lenovo/Documents/crmnew/api-geofal-crm/app/templates/Temp_Cotizacion.xlsx")
    if template.exists():
        dump_sheet_structure(template, "MORT1")
        dump_sheet_structure(template, "MORT2")
    else:
        print("Template not found")
