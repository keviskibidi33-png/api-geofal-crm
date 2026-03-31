from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')
ws = wb.worksheets[1]

print("=== FULL CONDICIONES STRUCTURE ===\n")

# Check rows 16-19 with all columns
for row in range(16, 20):
    print(f"Row {row}:")
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        val = ws[f'{col}{row}'].value
        merged = any(f'{col}{row}' in str(mc) for mc in ws.merged_cells)
        print(f"  {col}{row}: {repr(val)} {'[MERGED]' if merged else ''}")
    print()
