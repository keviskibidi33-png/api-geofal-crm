import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb["F LEM1"]

print("=== COLUMNS L, M, N, O IN F LEM1 ===")
for r in range(2, 17):
    row_str = f"Row {r}: "
    for c in [12, 13, 14, 15]: # L, M, N, O
        cell = ws.cell(row=r, column=c)
        col_letter = openpyxl.utils.get_column_letter(c)
        if cell.value is not None:
            row_str += f"{col_letter}{r}: {repr(cell.value)} | "
    print(row_str)
