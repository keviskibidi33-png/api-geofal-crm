from __future__ import annotations

import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.modules.proctor.excel import POINT_COLS, SIEVE_ROWS, generate_proctor_excel
from app.modules.proctor.schemas import ProctorPuntoRow, ProctorRequest


OUTPUT_DIR = ROOT / "verificaciones" / "proctor_visual"
REPORT_FILE = OUTPUT_DIR / "report.txt"

custom_output = os.getenv("PROCTOR_VISUAL_OUTPUT_DIR", "").strip()
if custom_output:
    OUTPUT_DIR = Path(custom_output)
    REPORT_FILE = OUTPUT_DIR / "report.txt"

if os.getenv("PROCTOR_VISUAL_TIMESTAMPED_DIR", "").strip().lower() in {"1", "true", "yes"}:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUTPUT_DIR = OUTPUT_DIR.parent / f"{OUTPUT_DIR.name}_{stamp}"
    REPORT_FILE = OUTPUT_DIR / "report.txt"


@dataclass
class ScenarioResult:
    name: str
    file_path: Path
    payload_path: Path
    checked_cells: int
    mismatches: list[str]
    footer_ok: bool


def _round(value: float | None, decimals: int) -> float | None:
    if value is None:
        return None
    return round(value, decimals)


def _resolve_point(point: ProctorPuntoRow) -> tuple[float | None, float | None, float | None, float | None, float | None, float | None]:
    masa_compactado_c = point.masa_suelo_compactado_c
    if masa_compactado_c is None and point.masa_suelo_humedo_molde_a is not None and point.masa_molde_compactacion_b is not None:
        masa_compactado_c = _round(point.masa_suelo_humedo_molde_a - point.masa_molde_compactacion_b, 2)

    densidad_humeda_x = point.densidad_humeda_x
    if densidad_humeda_x is None and masa_compactado_c is not None and point.volumen_molde_compactacion_d not in (None, 0):
        densidad_humeda_x = _round(masa_compactado_c / point.volumen_molde_compactacion_d, 3)

    masa_agua_y = point.masa_agua_y
    if (
        masa_agua_y is None
        and point.masa_recipiente_suelo_humedo_e is not None
        and point.masa_recipiente_suelo_seco_3_f is not None
    ):
        masa_agua_y = _round(point.masa_recipiente_suelo_humedo_e - point.masa_recipiente_suelo_seco_3_f, 2)

    masa_suelo_seco_z = point.masa_suelo_seco_z
    if masa_suelo_seco_z is None and point.masa_recipiente_suelo_seco_3_f is not None and point.masa_recipiente_g is not None:
        masa_suelo_seco_z = _round(point.masa_recipiente_suelo_seco_3_f - point.masa_recipiente_g, 2)

    contenido_humedad_w = point.contenido_humedad_moldeo_w
    if contenido_humedad_w is None and masa_agua_y is not None and masa_suelo_seco_z not in (None, 0):
        contenido_humedad_w = _round((masa_agua_y / masa_suelo_seco_z) * 100, 2)

    densidad_seca = point.densidad_seca
    if densidad_seca is None and densidad_humeda_x is not None and contenido_humedad_w is not None:
        divisor = 1 + (contenido_humedad_w / 100)
        if divisor != 0:
            densidad_seca = _round(densidad_humeda_x / divisor, 3)

    return masa_compactado_c, densidad_humeda_x, masa_agua_y, masa_suelo_seco_z, contenido_humedad_w, densidad_seca


def _expected_cells(payload: ProctorRequest) -> dict[str, str | float | int]:
    cells: dict[str, str | float | int] = {}

    def add(ref: str, value: object) -> None:
        if value is None:
            return
        cells[ref] = value

    # Header
    add("B9", payload.muestra)
    add("C9", payload.numero_ot)
    add("F9", payload.fecha_ensayo)
    add("H9", payload.realizado_por)

    # Point columns
    for idx, col in enumerate(POINT_COLS):
        point = payload.puntos[idx]
        masa_compactado_c, densidad_humeda_x, masa_agua_y, masa_suelo_seco_z, contenido_humedad_w, densidad_seca = _resolve_point(point)

        add(f"{col}15", point.prueba_numero)
        add(f"{col}16", point.numero_capas)
        add(f"{col}17", point.numero_golpes)
        add(f"{col}18", point.masa_suelo_humedo_molde_a)
        add(f"{col}19", point.masa_molde_compactacion_b)
        add(f"{col}20", masa_compactado_c)
        add(f"{col}21", point.volumen_molde_compactacion_d)
        add(f"{col}22", densidad_humeda_x)

        add(f"{col}24", point.tara_numero)
        add(f"{col}25", point.masa_recipiente_suelo_humedo_e)
        add(f"{col}26", point.masa_recipiente_suelo_seco_1)
        add(f"{col}27", point.masa_recipiente_suelo_seco_2)
        add(f"{col}28", point.masa_recipiente_suelo_seco_3_f)
        add(f"{col}29", masa_agua_y)
        add(f"{col}30", point.masa_recipiente_g)
        add(f"{col}31", masa_suelo_seco_z)
        add(f"{col}32", contenido_humedad_w)
        add(f"{col}33", densidad_seca)

    # Description and conditions
    add("C35", payload.tipo_muestra)
    add("C36", payload.condicion_muestra)
    add("C37", payload.tamano_maximo_particula_in)
    add("C38", payload.forma_particula)
    add("C39", payload.clasificacion_sucs_visual)
    add("C41", payload.metodo_ensayo)
    add("C42", payload.metodo_preparacion)
    add("C43", payload.tipo_apisonador)
    add("C44", payload.contenido_humedad_natural_pct)
    add("C45", payload.excluyo_material_muestra)
    add("C46", payload.observaciones)

    # Sieve calculations
    sieve_mass = list(payload.tamiz_masa_retenida_g)
    sieve_pct = list(payload.tamiz_porcentaje_retenido)
    sieve_pct_acc = list(payload.tamiz_porcentaje_retenido_acumulado)

    total_index = len(sieve_mass) - 1
    if sieve_mass[total_index] is None and all(value is not None for value in sieve_mass[:total_index]):
        sieve_mass[total_index] = _round(sum(value for value in sieve_mass[:total_index] if value is not None), 2)

    total_mass = sieve_mass[total_index] if sieve_mass[total_index] not in (None, 0) else None
    if total_mass:
        running = 0.0
        for idx in range(total_index):
            value = sieve_mass[idx]
            if value is not None and sieve_pct[idx] is None:
                sieve_pct[idx] = _round((value / total_mass) * 100, 2)
            if sieve_pct[idx] is not None:
                running += sieve_pct[idx] or 0
                if sieve_pct_acc[idx] is None:
                    sieve_pct_acc[idx] = _round(running, 2)
        if sieve_pct[total_index] is None:
            sieve_pct[total_index] = 100.0
        if sieve_pct_acc[total_index] is None:
            sieve_pct_acc[total_index] = 100.0

    for idx, row_num in enumerate(SIEVE_ROWS):
        add(f"G{row_num}", sieve_mass[idx])
        add(f"H{row_num}", sieve_pct[idx])
        add(f"I{row_num}", sieve_pct_acc[idx])

    # Equipment
    add("H44", payload.tamiz_utilizado_metodo_codigo)
    add("H45", payload.balanza_1g_codigo)
    add("H46", payload.balanza_codigo)
    add("H47", payload.horno_110_codigo)
    add("H48", payload.molde_codigo)
    add("H49", payload.pison_codigo)

    return cells


def _compare_cells(xlsx_bytes: bytes, payload: ProctorRequest) -> tuple[int, list[str]]:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    ws = wb.active

    expected = _expected_cells(payload)
    mismatches: list[str] = []

    for cell_ref, expected_value in expected.items():
        got = ws[cell_ref].value
        if isinstance(expected_value, float):
            if got is None or abs(float(got) - expected_value) > 1e-6:
                mismatches.append(f"{cell_ref}: expected={expected_value} got={got}")
        else:
            if got != expected_value:
                mismatches.append(f"{cell_ref}: expected={expected_value!r} got={got!r}")

    return len(expected), mismatches


def _check_footer(xlsx_bytes: bytes, payload: ProctorRequest) -> bool:
    if not any([payload.revisado_por, payload.revisado_fecha, payload.aprobado_por, payload.aprobado_fecha]):
        return True

    with ZipFile(BytesIO(xlsx_bytes), "r") as zf:
        drawing_xml = zf.read("xl/drawings/drawing1.xml").decode("utf-8", errors="ignore")

    tokens = ["Revisado:", "Aprobado:"]
    if payload.revisado_por:
        tokens.append(payload.revisado_por)
    if payload.revisado_fecha:
        tokens.append(payload.revisado_fecha)
    if payload.aprobado_por:
        tokens.append(payload.aprobado_por)
    if payload.aprobado_fecha:
        tokens.append(payload.aprobado_fecha)

    return all(token in drawing_xml for token in tokens)


def _build_payloads() -> list[tuple[str, dict]]:
    base_points = [
        {
            "prueba_numero": 1,
            "numero_capas": 4,
            "numero_golpes": 45,
            "masa_suelo_humedo_molde_a": 4100,
            "masa_molde_compactacion_b": 2100,
            "volumen_molde_compactacion_d": 1000,
            "tara_numero": "T-01",
            "masa_recipiente_suelo_humedo_e": 111,
            "masa_recipiente_suelo_seco_1": 105,
            "masa_recipiente_suelo_seco_2": 101,
            "masa_recipiente_suelo_seco_3_f": 99,
            "masa_recipiente_g": 10,
        },
        {
            "prueba_numero": 2,
            "numero_capas": 5,
            "numero_golpes": 50,
            "masa_suelo_humedo_molde_a": 4200,
            "masa_molde_compactacion_b": 2200,
            "volumen_molde_compactacion_d": 1000,
            "tara_numero": "T-02",
            "masa_recipiente_suelo_humedo_e": 121,
            "masa_recipiente_suelo_seco_1": 114,
            "masa_recipiente_suelo_seco_2": 109,
            "masa_recipiente_suelo_seco_3_f": 104,
            "masa_recipiente_g": 11,
        },
        {
            "prueba_numero": 3,
            "numero_capas": 6,
            "numero_golpes": 55,
            "masa_suelo_humedo_molde_a": 4300,
            "masa_molde_compactacion_b": 2300,
            "volumen_molde_compactacion_d": 1000,
            "tara_numero": "T-03",
            "masa_recipiente_suelo_humedo_e": 131,
            "masa_recipiente_suelo_seco_1": 123,
            "masa_recipiente_suelo_seco_2": 117,
            "masa_recipiente_suelo_seco_3_f": 110,
            "masa_recipiente_g": 12,
        },
        {
            "prueba_numero": 4,
            "numero_capas": 7,
            "numero_golpes": 60,
            "masa_suelo_humedo_molde_a": 4400,
            "masa_molde_compactacion_b": 2400,
            "volumen_molde_compactacion_d": 1000,
            "tara_numero": "T-04",
            "masa_recipiente_suelo_humedo_e": 141,
            "masa_recipiente_suelo_seco_1": 132,
            "masa_recipiente_suelo_seco_2": 125,
            "masa_recipiente_suelo_seco_3_f": 116,
            "masa_recipiente_g": 13,
        },
        {
            "prueba_numero": 5,
            "numero_capas": 8,
            "numero_golpes": 65,
            "masa_suelo_humedo_molde_a": 4500,
            "masa_molde_compactacion_b": 2500,
            "volumen_molde_compactacion_d": 1000,
            "tara_numero": "T-05",
            "masa_recipiente_suelo_humedo_e": 151,
            "masa_recipiente_suelo_seco_1": 141,
            "masa_recipiente_suelo_seco_2": 133,
            "masa_recipiente_suelo_seco_3_f": 122,
            "masa_recipiente_g": 14,
        },
    ]

    base_common = {
        "muestra": "789",
        "numero_ot": "1234",
        "fecha_ensayo": "050226",
        "realizado_por": "QA LOCAL",
        "puntos": base_points,
        "tipo_muestra": "TIPO A",
        "condicion_muestra": "SECA",
        "tamano_maximo_particula_in": "3/4",
        "forma_particula": "ANGULAR",
        "clasificacion_sucs_visual": "GW",
        "metodo_ensayo": "A",
        "metodo_preparacion": "HUMEDO",
        "tipo_apisonador": "MANUAL",
        "contenido_humedad_natural_pct": 5.5,
        "excluyo_material_muestra": "NO",
        "tamiz_masa_retenida_g": [100, 80, 60, 40, None],
        "tamiz_porcentaje_retenido": [None, None, None, None, None],
        "tamiz_porcentaje_retenido_acumulado": [None, None, None, None, None],
        "tamiz_utilizado_metodo_codigo": "TM-01",
        "balanza_1g_codigo": "B1-01",
        "balanza_codigo": "B2-01",
        "horno_110_codigo": "H-110",
        "molde_codigo": "M-01",
        "pison_codigo": "P-01",
        "observaciones": "OBS LOCAL",
        "revisado_por": "REV LOCAL",
        "revisado_fecha": "060226",
        "aprobado_por": "APR LOCAL",
        "aprobado_fecha": "070226",
    }

    scenario_manual_override = json.loads(json.dumps(base_common))
    scenario_manual_override["muestra"] = "900"
    scenario_manual_override["numero_ot"] = "9876"
    scenario_manual_override["puntos"][0]["masa_suelo_compactado_c"] = 1999.5
    scenario_manual_override["puntos"][0]["densidad_humeda_x"] = 1.999
    scenario_manual_override["puntos"][0]["masa_agua_y"] = 10.5
    scenario_manual_override["puntos"][0]["masa_suelo_seco_z"] = 88.8
    scenario_manual_override["puntos"][0]["contenido_humedad_moldeo_w"] = 11.82
    scenario_manual_override["puntos"][0]["densidad_seca"] = 1.788
    scenario_manual_override["tamiz_porcentaje_retenido"] = [35.71, 28.57, 21.43, 14.29, 100.0]
    scenario_manual_override["tamiz_porcentaje_retenido_acumulado"] = [35.71, 64.28, 85.71, 100.0, 100.0]

    scenario_compacto = json.loads(json.dumps(base_common))
    scenario_compacto["muestra"] = " 777 su "
    scenario_compacto["numero_ot"] = "OT-2222"
    scenario_compacto["fecha_ensayo"] = "7/2"
    scenario_compacto["revisado_fecha"] = "080226"
    scenario_compacto["aprobado_fecha"] = "090226"
    scenario_compacto["tamiz_masa_retenida_g"] = [20, 30, 25, 25, None]
    scenario_compacto["tamiz_porcentaje_retenido"] = [None, None, None, None, None]
    scenario_compacto["tamiz_porcentaje_retenido_acumulado"] = [None, None, None, None, None]

    return [
        ("full_auto_calculated", base_common),
        ("manual_override_fields", scenario_manual_override),
        ("normalized_input_formats", scenario_compacto),
    ]


def run() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    results: list[ScenarioResult] = []

    for index, (name, raw_payload) in enumerate(_build_payloads(), start=1):
        payload = ProctorRequest.model_validate(raw_payload)
        excel_bytes = generate_proctor_excel(payload)

        file_path = OUTPUT_DIR / f"{index:02d}_{name}.xlsx"
        payload_path = OUTPUT_DIR / f"{index:02d}_{name}_payload.json"

        file_path.write_bytes(excel_bytes)
        payload_path.write_text(json.dumps(payload.model_dump(mode="json"), indent=2), encoding="utf-8")

        checked_cells, mismatches = _compare_cells(excel_bytes, payload)
        footer_ok = _check_footer(excel_bytes, payload)

        results.append(
            ScenarioResult(
                name=name,
                file_path=file_path,
                payload_path=payload_path,
                checked_cells=checked_cells,
                mismatches=mismatches,
                footer_ok=footer_ok,
            )
        )

    lines: list[str] = []
    lines.append("PROCTOR VISUAL FILL TESTS")
    lines.append("=" * 80)
    lines.append(f"Output directory: {OUTPUT_DIR}")
    lines.append("")

    has_errors = False
    for result in results:
        ok = (not result.mismatches) and result.footer_ok
        status = "OK" if ok else "FAIL"
        lines.append(f"[{status}] {result.name}")
        lines.append(f"  file: {result.file_path}")
        lines.append(f"  payload: {result.payload_path}")
        lines.append(f"  checked_cells: {result.checked_cells}")
        lines.append(f"  mismatches: {len(result.mismatches)}")
        lines.append(f"  footer_ok: {result.footer_ok}")
        if result.mismatches:
            has_errors = True
            lines.append("  mismatch_details:")
            for mismatch in result.mismatches:
                lines.append(f"    - {mismatch}")
        if not result.footer_ok:
            has_errors = True
            lines.append("  footer_details: missing expected footer tokens in drawing1.xml")
        lines.append("")

    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines))
    return 1 if has_errors else 0


if __name__ == "__main__":
    sys.exit(run())
