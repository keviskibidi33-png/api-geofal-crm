import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=False)
ws = wb["F LEM1"]

print("=== F LEM1 row 30-37 ===")
for r in range(30, 38):
    row_str = f"Row {r}: "
    for c in range(2, 11): # B to J
        cell = ws.cell(row=r, column=c)
        col_letter = openpyxl.utils.get_column_letter(c)
        if cell.value is not None:
            row_str += f"{col_letter}{r}: {repr(cell.value)} | "
    print(row_str)
