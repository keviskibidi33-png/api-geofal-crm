
import openpyxl
import os
import glob

def verify_file(filepath):
    print(f"\nChecking: {os.path.basename(filepath)}")
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # 1. Verify Density (P11)
        density_val = ws['P11'].value
        print(f"  [P11] Density: {density_val} (Expected: 'SI' or 'NO' or None)")
        
        # 2. Verify Item Row 14 (First item)
        # Columns: C=Estructura, D=f'c, E=Fecha Moldeo, F=Fecha Rotura, O=Tipo Fractura
        c14 = ws['C14'].value
        d14 = ws['D14'].value
        e14 = ws['E14'].value
        f14 = ws['F14'].value
        o14 = ws['O14'].value
        
        print(f"  [C14] Estructura    : {c14}")
        print(f"  [D14] f'c           : {d14}")
        print(f"  [E14] Fecha Moldeo  : {e14}")
        print(f"  [F14] Fecha Rotura  : {f14}")
        print(f"  [O14] Tipo Fractura : {o14}")

        # Basic Assertions
        if "Viga" in str(c14): print("  -> C14 OK")
        else: print("  -> C14 FAIL (Expected 'Viga...')")
        
        if d14 is not None and float(d14) > 0: print("  -> D14 OK")
        else: print("  -> D14 FAIL (Expected numeric > 0)")
        
        # 3. Verify Expansion & Spacer (if > 14 items)
        # Just check if last data row + 1 is empty or formatted
        # We know test_vlayout_15_items... has 15 items. 
        # Last item should be at 14 + 14 = 28. Spacer at 29.
        if "15_items" in filepath:
            # Item 15 is at row 14 + 14 = 28
            c28 = ws['C28'].value
            print(f"  [C28] Item 15 Estructura: {c28}")
            if "Viga 15" in str(c28): print("  -> C28 OK (Expansion Works)")
            else: print("  -> C28 FAIL (Expansion Broken)")
            
            # Spacer at 29
            c29 = ws['C29'].value
            print(f"  [C29] Spacer Row Content: '{c29}'")
            if c29 is None or c29 == "": print("  -> C29 OK (Spacer Empty)")
            else: print("  -> C29 FAIL (Spacer Not Empty)")

    except Exception as e:
        print(f"  -> ERROR reading file: {e}")

if __name__ == "__main__":
    files = glob.glob("test_vlayout_*.xlsx")
    for f in files:
        verify_file(f)
