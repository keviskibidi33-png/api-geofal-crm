from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')

print("=== SHEETS IN WORKBOOK ===")
for idx, sheet_name in enumerate(wb.sheetnames):
    print(f"{idx}: {sheet_name}")
    ws = wb[sheet_name]
    
    # Check if it has PH content
    if ws['A7'].value:
        print(f"  A7: {ws['A7'].value}")
    if ws['A13'].value:
        print(f"  A13: {ws['A13'].value}")
    if ws['B10'].value:
        print(f"  B10: {ws['B10'].value}")
    print()

print("\n=== ACTIVE SHEET ===")
print(f"Active: {wb.active.title}")
