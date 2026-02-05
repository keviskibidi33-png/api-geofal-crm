
import os
import sys
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load env
env_path = Path(__file__).resolve().parents[1] / ".env"
load_dotenv(env_path)

DATABASE_URL = os.getenv("QUOTES_DATABASE_URL")
if not DATABASE_URL:
    print("Error: QUOTES_DATABASE_URL not set")
    sys.exit(1)

VENDOR_NAME = "Silvia Peralta"

def generate_report():
    print(f"Connecting to database to fetch report for: {VENDOR_NAME}...")
    # ... (rest of connection logic is fine) ...
    engine = create_engine(DATABASE_URL)
    
    with engine.connect() as conn:
        # Query with readable columns
        query = text("""
            SELECT 
                to_char(fecha_emision, 'DD/MM/YYYY') as "Fecha Emisión",
                CONCAT(numero, '-', year) as "Cotización",
                cliente_nombre as "Cliente",
                proyecto as "Proyecto",
                estado as "Estado",
                coalesce(moneda, 'PEN') as "Moneda",
                coalesce(total, 0) as "Monto Total"
            FROM cotizaciones
            WHERE personal_comercial ILIKE :vendor
            ORDER BY fecha_emision DESC
        """)
        
        # Use pandas for easy handling
        df = pd.read_sql(query, conn, params={"vendor": f"%{VENDOR_NAME}%"})
    
    if df.empty:
        print(f"No records found for vendor: {VENDOR_NAME}")
        return

    print(f"Found {len(df)} records. Generating Excel...")
    
    # Calculate Summary
    total_quotes = len(df)
    total_amount_pen = df[df["Moneda"] == 'PEN']["Monto Total"].sum()
    try:
        total_amount_usd = df[df["Moneda"] == 'USD']["Monto Total"].sum()
    except KeyError:
        total_amount_usd = 0
    
    # Create Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Productividad"
    
    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    currency_fmt = '#,##0.00'
    
    # --- Title ---
    ws.merge_cells('A1:G1')
    ws['A1'] = f"REPORTE DE PRODUCTIVIDAD COMERCIAL - {datetime.now().year}"
    ws['A1'].font = Font(bold=True, size=16, color="003366")
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"VENDEDOR: {VENDOR_NAME.upper()}"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = center_align
    
    ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # --- Summary Section (Lines 5-7) ---
    ws['A5'] = "Resumen General"
    ws['A5'].font = Font(bold=True, underline="single")
    
    ws['A6'] = "Total Cotizaciones:"
    ws['B6'] = total_quotes
    
    ws['A7'] = "Monto Total (PEN):"
    ws['B7'] = total_amount_pen
    ws['B7'].number_format = currency_fmt
    
    if total_amount_usd > 0:
        ws['A8'] = "Monto Total (USD):"
        ws['B8'] = total_amount_usd
        ws['B8'].number_format = currency_fmt
    
    # --- Data Table (Line 10 onwards) ---
    start_row = 10
    columns = list(df.columns)
    
    # Helper to calculate column widths
    col_widths = {'A': 15, 'B': 15, 'C': 40, 'D': 40, 'E': 15, 'F': 10, 'G': 15}
    
    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_thin
        
        # Set simplistic width
        col_letter = get_column_letter(col_idx)
        if col_letter in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[col_letter]
            
    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_thin
            
            # Format Currency Column (Index 7 = G)
            if c_idx == 7: # Monto Total
                cell.number_format = currency_fmt
    
    filename = f"Reporte_Productividad_{VENDOR_NAME.replace(' ', '_')}.xlsx"
    wb.save(filename)
    print(f"Report saved successfully: {os.path.abspath(filename)}")

if __name__ == "__main__":
    generate_report()
