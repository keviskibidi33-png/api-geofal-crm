import io
import os
import zipfile
import copy
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from lxml import etree
import openpyxl  # Added for parsing
from .models import RecepcionMuestra, MuestraConcreto

NAMESPACES = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}

def _parse_cell_ref(ref: str) -> tuple[str, int]:
    col = ''.join(c for c in ref if c.isalpha())
    row = int(''.join(c for c in ref if c.isdigit()))
    return col, row

def _col_letter_to_num(col: str) -> int:
    num = 0
    for c in col.upper():
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num

def _num_to_col_letter(n: int) -> str:
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def _find_or_create_row(sheet_data: etree._Element, row_num: int, ns: str) -> etree._Element:
    for row in sheet_data.iterfind(f'{{{ns}}}row'):
        if row.get('r') == str(row_num):
            return row
    row = etree.SubElement(sheet_data, f'{{{ns}}}row')
    row.set('r', str(row_num))
    return row

def _set_cell_value_fast(row, ref, value, ns, is_number=False, get_string_idx=None):
    c = row.find(f'{{{ns}}}c[@r="{ref}"]')
    if c is None:
        c = etree.SubElement(row, f'{{{ns}}}c')
        c.set('r', ref)
    
    style = c.get('s')
    for child in list(c):
        c.remove(child)
    
    if value is None or value == '':
        if 't' in c.attrib: del c.attrib['t']
        if style: c.set('s', style)
        return

    if is_number:
        if 't' in c.attrib: del c.attrib['t']
        v = etree.SubElement(c, f'{{{ns}}}v')
        v.text = str(value)
    else:
        if get_string_idx:
            c.set('t', 's')
            v = etree.SubElement(c, f'{{{ns}}}v')
            v.text = str(get_string_idx(str(value)))
        else:
            c.set('t', 'inlineStr')
            is_elem = etree.SubElement(c, f'{{{ns}}}is')
            t = etree.SubElement(is_elem, f'{{{ns}}}t')
            t.text = str(value)
    
    if style:
        c.set('s', style)

def _duplicate_row_xml(sheet_data: etree._Element, source_row_num: int, target_row_num: int, ns: str):
    source_row = sheet_data.find(f'{{{ns}}}row[@r="{source_row_num}"]')
    if source_row is None:
        return
    
    new_row = copy.deepcopy(source_row)
    new_row.set('r', str(target_row_num))
    
    for cell in new_row.findall(f'{{{ns}}}c'):
        old_ref = cell.get('r')
        col, _ = _parse_cell_ref(old_ref)
        cell.set('r', f'{col}{target_row_num}')
        # Clear value but keep style
        for child in list(cell):
            if child.tag != f'{{{ns}}}v' or True: # Clear all kids to be safe, we will write later
                cell.remove(child)

    # Insert in order
    rows = sheet_data.findall(f'{{{ns}}}row')
    inserted = False
    for i, r in enumerate(rows):
        if int(r.get('r')) > target_row_num:
            r.addprevious(new_row)
            inserted = True
            break
    if not inserted:
        sheet_data.append(new_row)

def _shift_rows(sheet_data: etree._Element, from_row: int, shift: int, ns: str):
    if shift <= 0: return
    rows = list(sheet_data.findall(f'{{{ns}}}row'))
    rows.sort(key=lambda r: int(r.get('r')), reverse=True)
    for row in rows:
        row_num = int(row.get('r'))
        if row_num >= from_row:
            new_num = row_num + shift
            row.set('r', str(new_num))
            for cell in row.findall(f'{{{ns}}}c'):
                old_ref = cell.get('r')
                col, _ = _parse_cell_ref(old_ref)
                cell.set('r', f'{col}{new_num}')

def _shift_merged_cells(root: etree._Element, from_row: int, shift: int, ns: str):
    if shift <= 0: return
    merged_cells_node = root.find(f'{{{ns}}}mergeCells')
    if merged_cells_node is None: return
    for mc in merged_cells_node.findall(f'{{{ns}}}mergeCell'):
        ref = mc.get('ref')
        if not ref: continue
        parts = ref.split(':')
        new_parts = []
        changed = False
        for part in parts:
            c, r = _parse_cell_ref(part)
            if r >= from_row:
                new_parts.append(f"{c}{r + shift}")
                changed = True
            else:
                new_parts.append(part)
        if changed:
            mc.set('ref', ':'.join(new_parts))

def _duplicate_merged_cells(root: etree._Element, source_row_num: int, target_row_num: int, ns: str):
    merged_cells_node = root.find(f'{{{ns}}}mergeCells')
    if merged_cells_node is None: return
    
    new_merges = []
    for mc in merged_cells_node.findall(f'{{{ns}}}mergeCell'):
        ref = mc.get('ref')
        if not ref: continue
        parts = ref.split(':')
        if len(parts) != 2: continue
        
        c1, r1 = _parse_cell_ref(parts[0])
        c2, r2 = _parse_cell_ref(parts[1])
        
        # If the merge is exactly on the source row (e.g. B23:C23)
        if r1 == source_row_num and r2 == source_row_num:
            new_ref = f"{c1}{target_row_num}:{c2}{target_row_num}"
            new_merges.append(new_ref)
            
    for ref in new_merges:
        mc = etree.SubElement(merged_cells_node, f'{{{ns}}}mergeCell')
        mc.set('ref', ref)
    
    if new_merges:
        merged_cells_node.set('count', str(int(merged_cells_node.get('count', '0')) + len(new_merges)))

class ExcelLogic:
    def __init__(self, template_path: Optional[str] = None):
        if template_path:
            self.template_path = template_path
        else:
            # Standardized robust resolution
            filename = "Temp_Recepcion.xlsx"
            from pathlib import Path
            current_dir = Path(__file__).resolve().parent
            app_dir = current_dir.parents[1] # app/
            
            possible_paths = [
                app_dir / "templates" / filename,
                Path("/app/templates") / filename,
                current_dir.parents[2] / "app" / "templates" / filename,
            ]
            
            final_path = None
            for p in possible_paths:
                if p.exists():
                    final_path = p
                    break
            
            self.template_path = str(final_path or (app_dir / "templates" / filename))

    def generar_excel_recepcion(self, recepcion: RecepcionMuestra) -> bytes:
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template no encontrado en {self.template_path}")

        shared_strings = []
        ss_xml_original = None
        with zipfile.ZipFile(self.template_path, 'r') as z:
            if 'xl/sharedStrings.xml' in z.namelist():
                ss_xml_original = z.read('xl/sharedStrings.xml')
                ss_root = etree.fromstring(ss_xml_original)
                ns_ss = ss_root.nsmap.get(None, NAMESPACES['main'])
                sis = ss_root.findall(f'{{{ns_ss}}}si')
                for i, si in enumerate(sis):
                    t = si.find(f'{{{ns_ss}}}t')
                    if t is not None: shared_strings.append((t.text or "").strip())
                    else: shared_strings.append(''.join([x.text or '' for x in si.findall(f'.//{{{ns_ss}}}t')]).strip())

        ss_map = {text: i for i, text in enumerate(shared_strings)}
        
        def get_string_idx(text: str) -> int:
            text = str(text or "").strip()
            if text in ss_map: return ss_map[text]
            idx = len(shared_strings)
            shared_strings.append(text)
            ss_map[text] = idx
            return idx

        sheet_file = 'xl/worksheets/sheet1.xml'
        with zipfile.ZipFile(self.template_path, 'r') as z:
            sheet_xml = z.read(sheet_file)
        
        root = etree.fromstring(sheet_xml)
        ns = NAMESPACES['main']
        sheet_data = root.find(f'.//{{{ns}}}sheetData')

        # 1. Anchor Detection
        anchors = {} # label -> (col, row)
        for row_el in sheet_data.findall(f'{{{ns}}}row'):
            r_num = int(row_el.get('r'))
            for cell_el in row_el.findall(f'{{{ns}}}c'):
                val = ""
                if cell_el.get('t') == 's':
                    v_el = cell_el.find(f'{{{ns}}}v')
                    if v_el is not None:
                        try:
                            s_idx = int(v_el.text)
                            if 0 <= s_idx < len(shared_strings):
                                val = shared_strings[s_idx]
                        except: pass
                
                if val:
                    key = val.upper().strip()
                    c_name, _ = _parse_cell_ref(cell_el.get('r'))
                    if key not in anchors: # Take first occurrence
                        anchors[key] = (c_name, r_num)

        # 2. Dynamic Row Logic
        muestras = recepcion.muestras
        n_muestras = len(muestras)
        threshold = 18 # Rows 23 to 40 inclusive
        
        # Determine base coordinates using anchors or fallbacks
        row_n_label = anchors.get("N°", ("A", 21))[1]
        data_start_row = row_n_label + 2
        row_nota_label = anchors.get("NOTA:", ("B", 43))[1]
        
        if n_muestras > threshold:
            extra_rows = n_muestras - threshold
            # Shift from first row after the template table (Row 41 if threshold=18)
            shift_start = data_start_row + threshold 
            _shift_rows(sheet_data, shift_start, extra_rows, ns)
            _shift_merged_cells(root, shift_start, extra_rows, ns)
            # Duplicate template row for data
            # Row 23 is header-adjacent, Row 24 is inner (standard).
            inner_row_source = data_start_row + 1 
            for i in range(threshold, n_muestras):
                target_row = data_start_row + i
                _duplicate_row_xml(sheet_data, inner_row_source, target_row, ns)
                _duplicate_merged_cells(root, inner_row_source, target_row, ns)

        # Refresh cache for writing
        rows_cache = {r.get('r'): r for r in sheet_data.findall(f'{{{ns}}}row')}
        
        def write_cell(col, row_idx, value, is_num=False, is_footer=False):
            actual_row = row_idx
            if is_footer and n_muestras > threshold:
                actual_row += (n_muestras - threshold)
            
            row_el = rows_cache.get(str(actual_row))
            if row_el is None:
                row_el = _find_or_create_row(sheet_data, actual_row, ns)
                rows_cache[str(actual_row)] = row_el
            
            ref = f"{col}{actual_row}"
            _set_cell_value_fast(row_el, ref, value, ns, is_num, get_string_idx)

        # 3. Filling Data
        def format_dt(dt):
            if not dt: return "-"
            if isinstance(dt, (datetime, date)): return dt.strftime("%d/%m/%Y")
            return str(dt)

        # Header Section
        # Anchors: RECEPCIÓN N°, COTIZACIÓN N°, FECHA DE RECEPCIÓN, OT:, CLIENTE:, PROYECTO:, etc.
        def write_to_neighbor(label, value, is_num=False, offset_col=0):
            if label.upper() in anchors:
                c, r = anchors[label.upper()]
                target_col = _num_to_col_letter(_col_letter_to_num(c) + 1 + offset_col)
                write_cell(target_col, r, value, is_num)

        write_cell('D', anchors.get("RECEPCIÓN N°:", ("A", 6))[1], recepcion.numero_recepcion)
        write_to_neighbor("COTIZACIÓN N°:", recepcion.numero_cotizacion or "-", offset_col=2)
        write_to_neighbor("FECHA DE RECEPCIÓN:", format_dt(recepcion.fecha_recepcion))
        write_to_neighbor("OT N°:", recepcion.numero_ot)
        
        # Details (D/H offsets)
        write_cell('D', anchors.get("CLIENTE :", ("C", 10))[1], recepcion.cliente)
        write_cell('D', anchors.get("DOMICILIO LEGAL :", ("C", 11))[1], recepcion.domicilio_legal)
        write_cell('D', anchors.get("RUC :", ("C", 12))[1], recepcion.ruc)
        write_cell('D', anchors.get("PERSONA CONTACTO :", ("C", 13))[1], recepcion.persona_contacto)
        write_cell('D', anchors.get("E-MAIL :", ("C", 14))[1], recepcion.email)
        write_cell('H', anchors.get("TELÉFONO :", ("G", 14))[1], recepcion.telefono)
        
        write_cell('D', anchors.get("SOLICITANTE :", ("C", 16))[1], recepcion.solicitante)
        write_cell('D', anchors.get("SOLICITANTE :", ("C", 16))[1] + 1, recepcion.domicilio_solicitante)
        write_cell('D', anchors.get("PROYECTO :", ("C", 18))[1], recepcion.proyecto)
        write_cell('D', anchors.get("UBICACIÓN :", ("C", 19))[1], recepcion.ubicacion)

        # Samples Table
        for idx, m in enumerate(muestras):
            curr_row = data_start_row + idx
            write_cell('A', curr_row, idx + 1, is_num=True)
            write_cell('B', curr_row, getattr(m, 'codigo_muestra_lem', '') or '')
            # C is skipped (merged or empty in template)
            write_cell('D', curr_row, getattr(m, 'identificacion_muestra', '') or '')
            write_cell('E', curr_row, m.estructura)
            write_cell('F', curr_row, m.fc_kg_cm2, is_num=True)
            write_cell('G', curr_row, m.fecha_moldeo)
            write_cell('H', curr_row, m.hora_moldeo)
            write_cell('I', curr_row, m.edad, is_num=True)
            write_cell('J', curr_row, m.fecha_rotura)
            write_cell('K', curr_row, "SI" if m.requiere_densidad else "NO")

        # Footer
        footer_row = row_nota_label
        write_cell('D', footer_row, recepcion.observaciones or "", is_footer=True)
        
        # Checkboxes
        write_cell('B', footer_row + 3, "X" if recepcion.emision_fisica else "", is_footer=True)
        write_cell('B', footer_row + 4, "X" if recepcion.emision_digital else "", is_footer=True)

        # Fecha Culminacion
        write_cell('H', footer_row + 3, format_dt(recepcion.fecha_estimada_culminacion), is_footer=True)

        # Signatures
        write_cell('D', footer_row + 6, recepcion.entregado_por or "", is_footer=True)
        write_cell('I', footer_row + 6, recepcion.recibido_por or "", is_footer=True)

        # 4. Handle Drawings (Blue Line Shift)
        drawing_file = 'xl/drawings/drawing1.xml'
        modified_drawing_xml = None
        if n_muestras > threshold:
            shift = n_muestras - threshold
            with zipfile.ZipFile(self.template_path, 'r') as z:
                if drawing_file in z.namelist():
                    draw_xml = z.read(drawing_file)
                    d_root = etree.fromstring(draw_xml)
                    d_ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
                    
                    anchors_list = d_root.xpath('//xdr:twoCellAnchor | //xdr:oneCellAnchor', namespaces=d_ns)
                    for anchor in anchors_list:
                        frow = anchor.find('.//xdr:from/xdr:row', namespaces=d_ns)
                        trow = anchor.find('.//xdr:to/xdr:row', namespaces=d_ns)
                        
                        if frow is not None:
                            orig_row = int(frow.text)
                            if orig_row < 10:
                                continue
                                
                            if orig_row >= (row_nota_label - 1):
                                frow.text = str(orig_row + shift)
                                if trow is not None:
                                    trow.text = str(int(trow.text) + shift)
                        
                    modified_drawing_xml = etree.tostring(d_root, encoding='utf-8', xml_declaration=True)

        # 5. Serialize Sheet
        modified_sheet_xml = etree.tostring(root, encoding='utf-8', xml_declaration=True)

        # 6. Reconstruct Shared Strings (Cleanly)
        ss_root_new = etree.Element(f'{{{ns}}}sst', nsmap={None: ns})
        for text in shared_strings:
            si = etree.SubElement(ss_root_new, f'{{{ns}}}si')
            t = etree.SubElement(si, f'{{{ns}}}t')
            t.text = text
        ss_root_new.set('count', str(len(shared_strings)))
        ss_root_new.set('uniqueCount', str(len(shared_strings)))
        modified_ss_xml = etree.tostring(ss_root_new, encoding='utf-8', xml_declaration=True)

        # 7. Build Final ZIP
        output = io.BytesIO()
        with zipfile.ZipFile(self.template_path, 'r') as z_in:
            with zipfile.ZipFile(output, 'w', compression=zipfile.ZIP_DEFLATED) as z_out:
                for item in z_in.namelist():
                    if item == sheet_file:
                        z_out.writestr(item, modified_sheet_xml)
                    elif item == 'xl/sharedStrings.xml':
                        z_out.writestr(item, modified_ss_xml)
                    elif item == drawing_file and modified_drawing_xml is not None:
                        z_out.writestr(item, modified_drawing_xml)
                    else:
                        z_out.writestr(item, z_in.read(item))
        
        output.seek(0)
        return output.getvalue()

    def parsear_recepcion(self, content: bytes) -> dict:
        """
        Parsea el contenido de un Excel de Recepción (Formato con etiquetas).
        Retorna un diccionario con los datos extraídos.
        """
        if not openpyxl:
            raise Exception("openpyxl library is required for parsing")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active # Default to active sheet, assume it's the right one or logic from router can be adapted
        
        # Helper to find cell by value (Anchor parsing)
        # Scan first 100 rows, 20 cols
        anchors = {}
        anchors_list = {}
        
        # Define keywords to look for
        keywords = {
            "RECEPCIÓN N°": ["recepcion_n", "recepcion_no", "recepcion n"],
            "COTIZACIÓN N°": ["cotizacion_n", "cotizacion_no"],
            "OT N°": ["ot_n", "ot_no"],
            "CLIENTE": ["cliente"],
            "RUC": ["ruc"],
            "DOMICILIO LEGAL": ["domicilio_legal"],
            "PERSONA CONTACTO": ["contacto"],
            "E-MAIL": ["email", "e-mail"],
            "TELÉFONO": ["telefono"],
            "SOLICITANTE": ["solicitante"],
            "PROYECTO": ["proyecto"],
            "UBICACIÓN": ["ubicacion"],
            "FECHA DE RECEPCIÓN": ["fecha_recepcion"],
            "FECHA ESTIMADA DE CULMINACIÓN": ["fecha_culminacion"],
            "ENTREGADO POR": ["entregado_por"],
            "RECIBIDO POR": ["recibido_por"],
            "OBSERVACIONES": ["observaciones"],
            "EMISIÓN DE INFORMES": ["emision_informes"] # Checkbox area
        }

        # Scan for anchors
        import re
        max_row_scan = 100
        for row in ws.iter_rows(min_row=1, max_row=max_row_scan, max_col=20, values_only=False):
            for cell in row:
                if isinstance(cell.value, str):
                    # Fix: Use only first line for multiline cells (e.g. "Entregado por:\n(Cliente)")
                    first_line = str(cell.value).split('\n')[0].strip()
                    val_clean = first_line.upper().replace(":", " ").strip()
                    # Store anchor by exact match of keyword
                    for key in keywords:
                        clean_key = key.replace(":", "").strip()
                        
                        # Paranoid check: "RUC" should NEVER match "ESTRUCTURA"
                        if "RUC" in clean_key and "ESTRUCTURA" in val_clean:
                            continue

                        # Construct fuzzy regex for whole word match
                        # Only add \b if the end is alphanumeric. 
                        pattern = re.escape(clean_key)
                        if clean_key and clean_key[0].isalnum(): pattern = r'\b' + pattern
                        if clean_key and clean_key[-1].isalnum(): pattern = pattern + r'\b'

                        # Use regex for whole word match to avoid substring matches
                        if re.search(pattern, val_clean):
                             # Heuristic: If match is too long relative to key, it's likely a value, not a label
                             # Relaxed from +15 to +40 for long footer labels
                             if len(val_clean) > len(clean_key) + 40:
                                 continue
                                 
                             anchors[key] = (cell.column, cell.row) # Last found
                             if key not in anchors_list: anchors_list[key] = []
                             anchors_list[key].append((cell.column, cell.row))

        def get_val(r, c):
             try:
                 return ws.cell(row=r, column=c).value
             except: return None

        def safe_str(val):
            if val is None: return ""
            if isinstance(val, (datetime, date)): return val.strftime("%d/%m/%Y")
            return str(val).strip()

        data = {}

        # Fallback Coordinates (derived from "REC Nº 1000-26")
        # Can be adjusted if more formats appear.
        FALLBACK_COORDS = {
            "RECEPCIÓN N°": (4, 6),      # D6
            "COTIZACIÓN N°": (6, 6),     # F6
            "OT N°": (8, 6),             # H6
            "CLIENTE": (4, 10),          # D10 (Skip C10 which is ":")
            "RUC": (4, 12),              # D12
            "DOMICILIO LEGAL": (4, 11),  # D11
            "PERSONA CONTACTO": (4, 13), # D13
            "E-MAIL": (4, 14),           # D14
            "TELÉFONO": (8, 14),         # H14
            "SOLICITANTE": (4, 16),      # D16
            "PROYECTO": (4, 18),         # D18
            "UBICACIÓN": (4, 19),        # D19
            "FECHA DE RECEPCIÓN": (10, 6) # J6
        }

        # Extraction Helpers
        def extract_right(anchor_name, offset_col=1, search_limit=5):
            # 1. Try Dynamic Anchor
            val_found = ""
            if anchor_name in anchors:
                col, row = anchors[anchor_name]
                print(f"[DEBUG] Extracting right for {anchor_name} at ({row}, {col})")
                
                # Search next few columns for non-empty
                for i in range(search_limit):
                    current_col = col + offset_col + i
                    val = get_val(row, current_col)
                    s_val = safe_str(val)
                    
                    print(f"  -> Check ({row}, {current_col}): '{s_val}' (Raw: {val})")
                    
                    if val:
                        # Skip if it's purely punctuation or very short non-alphanumeric
                        if re.match(r'^[\W_]+$', s_val) or (s_val.strip().startswith(":") and len(s_val) < 5):
                           print("     [SKIP] Punctuation-only")
                           continue
                        
                        # NUCLEAR OPTION: If extraction is short and contains colon, KILL IT.
                        if ":" in s_val and len(s_val) < 5:
                            print(f"     [SKIP-NUCLEAR] Value '{s_val}' contains colon and is short.")
                            continue
                        
                        # Fallback for short stuff that might sneak through
                        if len(s_val) < 2 and not s_val.isalnum():
                           print("     [SKIP] Short non-alnum")
                           continue

                        print(f"     [MATCH] '{s_val}'")
                        val_found = s_val
                        break # Found it
            else:
                print(f"[DEBUG] Anchor {anchor_name} NOT FOUND")
            
            # 2. Hybrid Fallback: If dynamic failed, try absolute coordinate
            if not val_found and anchor_name in FALLBACK_COORDS:
                f_col, f_row = FALLBACK_COORDS[anchor_name]
                print(f"[DEBUG] Fallback used for {anchor_name} -> ({f_row}, {f_col})")
                val = get_val(f_row, f_col)
                s_val = safe_str(val)
                # Apply same cleaning rules?
                if s_val and not re.match(r'^[\W_]+$', s_val):
                    print(f"     [FALLBACK MATCH] '{s_val}'")
                    val_found = s_val

            return val_found

        def extract_below(anchor_name, offset_row=1):
             if anchor_name in anchors:
                col, row = anchors[anchor_name]
                val = get_val(row + offset_row, col)
                return safe_str(val)
             return ""

        # Map fields
        data['numero_recepcion'] = extract_right("RECEPCIÓN N°", 1)
        data['numero_cotizacion'] = extract_right("COTIZACIÓN N°", 1) # Sometimes col offset is 2 depending on merge
        if not data['numero_cotizacion']: data['numero_cotizacion'] = extract_right("COTIZACIÓN N°", 2)
        
        data['numero_ot'] = extract_right("OT N°", 1)
        # Fix: Handle float OT numbers like 340.26 → "340-26"
        ot_val = data.get('numero_ot', '')
        if ot_val and '.' in ot_val:
            parts = ot_val.split('.')
            if len(parts) == 2 and all(p.isdigit() for p in parts):
                data['numero_ot'] = f"{parts[0]}-{parts[1]}"
        
        data['fecha_recepcion'] = extract_right("FECHA DE RECEPCIÓN", 1)
        
        # Fix: If numero_recepcion is just a number, append year suffix from fecha_recepcion
        nr = data.get('numero_recepcion', '')
        if nr and nr.replace(' ', '').isdigit():
            fr = data.get('fecha_recepcion', '')
            if fr:
                # Try to extract year from fecha_recepcion (formats: DD/MM/YYYY or datetime)
                year_suffix = ''
                if '/' in fr:
                    parts_date = fr.split('/')
                    if len(parts_date) == 3 and len(parts_date[2]) >= 2:
                        year_suffix = parts_date[2][-2:]
                elif len(fr) >= 4 and fr[-4:].isdigit():
                    year_suffix = fr[-2:]
                if year_suffix:
                    data['numero_recepcion'] = f"{nr.strip()}-{year_suffix}"
        
        data['cliente'] = extract_right("CLIENTE", 1)
        data['ruc'] = extract_right("RUC", 1)
        # Handle duplicate Domicilio Legal
        # Logic: First one is Client, Second is Solicitante (if exists)
        dl_locs = anchors_list.get("DOMICILIO LEGAL", [])
        dl_locs.sort(key=lambda x: x[1]) # Sort by row
        
        if len(dl_locs) > 0:
            c_col, c_row = dl_locs[0]
            val = get_val(c_row, c_col + 1)
            if not val: val = get_val(c_row, c_col + 2)
            if not val: val = get_val(c_row, c_col + 3)  # Fix: value may be 3 cols away (after ':')
            data['domicilio_legal'] = safe_str(val)
            
        if len(dl_locs) > 1:
            s_col, s_row = dl_locs[1]
            val = get_val(s_row, s_col + 1)
            if not val: val = get_val(s_row, s_col + 2)
            if not val: val = get_val(s_row, s_col + 3)  # Fix: value may be 3 cols away
            data['domicilio_solicitante'] = safe_str(val)
        else:
            # Fallback: Get domicilio from row below CLIENTE / SOLICITANTE anchors
            # This handles common typo "Domicilo legal" which doesn't match "DOMICILIO LEGAL"
            if "CLIENTE" in anchors and not data.get('domicilio_legal'):
                cl_col, cl_row = anchors["CLIENTE"]
                # Domicilio is typically one row below the Client label
                for col_off in [3, 2, 1]:
                    val = get_val(cl_row + 1, cl_col + col_off)
                    val_str = safe_str(val)
                    if val_str and not re.match(r'^[\W_]+$', val_str) and len(val_str) > 5:
                        data['domicilio_legal'] = val_str
                        break
            
            if "SOLICITANTE" in anchors:
                s_col, s_row = anchors["SOLICITANTE"]
                # Domicilio solicitante is typically one row below the Solicitante label
                for col_off in [3, 2, 1]:
                    val = get_val(s_row + 1, s_col + col_off)
                    val_str = safe_str(val)
                    if val_str and not re.match(r'^[\W_]+$', val_str) and len(val_str) > 5:
                        data['domicilio_solicitante'] = val_str
                        break

        data['persona_contacto'] = extract_right("PERSONA CONTACTO", 1)
        data['email'] = extract_right("E-MAIL", 1)
        data['telefono'] = extract_right("TELÉFONO", 1)
        
        data['solicitante'] = extract_right("SOLICITANTE", 1)
        data['proyecto'] = extract_right("PROYECTO", 1)
        data['ubicacion'] = extract_right("UBICACIÓN", 1)
        
        # Dates and Signatures
        data['fecha_estimada_culminacion'] = extract_right("FECHA ESTIMADA DE CULMINACIÓN", 1)
        # Fix: fecha_estimada_culminacion often has the value BELOW the anchor, not to the right
        if not data['fecha_estimada_culminacion']:
            data['fecha_estimada_culminacion'] = extract_below("FECHA ESTIMADA DE CULMINACIÓN", 1)
        # Also try with shorter keyword variant  
        if not data['fecha_estimada_culminacion'] and "EMISIÓN DE INFORMES" in anchors:
            em_col, em_row = anchors["EMISIÓN DE INFORMES"]
            # Fecha estimated is typically at same row+1, further right
            for c_off in range(3, 8):
                val = get_val(em_row + 1, em_col + c_off)
                if val:
                    data['fecha_estimada_culminacion'] = safe_str(val)
                    break
        
        data['entregado_por'] = extract_below("ENTREGADO POR", 1)
        # Fix: Also try offset_row=2 since "Entregado por:\n(Cliente)" may have empty row between
        if not data['entregado_por']:
            data['entregado_por'] = extract_below("ENTREGADO POR", 2)
        # Fix: Filter out URL/website text that sometimes bleeds into footer fields
        ep = data.get('entregado_por', '')
        if ep and ('www.' in ep.lower() or 'http' in ep.lower() or '@' in ep or 'laboratorio' in ep.lower()):
            data['entregado_por'] = ''  # Discard, it's the footer, not a name
        data['recibido_por'] = extract_below("RECIBIDO POR", 1)
        if not data['recibido_por']:
            data['recibido_por'] = extract_below("RECIBIDO POR", 2)
        rp = data.get('recibido_por', '')
        if rp and ('www.' in rp.lower() or 'http' in rp.lower() or '@' in rp or 'laboratorio' in rp.lower()):
            data['recibido_por'] = ''  # Discard, it's the footer

        # Muestras
        # Find header row for table, looks for "ITEM" or "CÓDIGO LEM"
        header_row = 0
        col_map = {}
        
        # Scan for header row and map columns
        min_r = 15
        for idx, row in enumerate(ws.iter_rows(min_row=min_r, max_row=40, values_only=False)):
             current_row_idx = min_r + idx
             row_values = []
             col_map = {} # Reset per row
             
             for cell in row:
                 cval = str(cell.value).strip().upper()
                 row_values.append(cval)
                 
                 # Map columns if header candidate found
                 if "ITEM" in cval: 
                     col_map["ITEM"] = cell.column
                 if "CODIGO" in cval or "CÓDIGO" in cval or "LEM" in cval:
                     col_map["LEM"] = cell.column
                 if "DESCRIPCI" in cval or "IDENTIFICA" in cval:
                     col_map["IDENT"] = cell.column
                 if "ESTRUCTURA" in cval:
                     col_map["ESTRUCTURA"] = cell.column
                 if "F'C" in cval or "FC" in cval:
                     col_map["FC"] = cell.column
                 if "MOLDEO" in cval and "FECHA" in cval:
                     col_map["FECHA_MOLDEO"] = cell.column
                 if "MOLDEO" in cval and "HORA" in cval:
                     col_map["HORA_MOLDEO"] = cell.column
                 if "EDAD" in cval:
                     col_map["EDAD"] = cell.column
                 if "ROTURA" in cval:
                     col_map["FECHA_ROTURA"] = cell.column
                 if "DENSIDAD" in cval:
                     col_map["DENSIDAD"] = cell.column

             # Check if this row looks like a header (has critical fields)
             if "ITEM" in col_map or "LEM" in col_map:
                 header_row = current_row_idx
                 print(f"[DEBUG] Found Table Header at Row {header_row}. Map: {col_map}")
                 break
        
        muestras = []
        if header_row > 0:
            start_row = header_row + 1
            
            # Fix: Detect two-row headers (e.g. Row 21: N°/LEM, Row 22: Estructura/F'c/Edad)
            # Check if start_row still contains header-like text instead of data
            second_header_keywords = {"ESTRUCTURA", "CODIGO", "CÓDIGO", "EDAD", "F'C", "FC", "MOLDEO", "ROTURA", "DENSIDAD"}
            is_second_header = False
            sub_header_col_map = {}
            for check_cell in ws.iter_rows(min_row=start_row, max_row=start_row, max_col=20, values_only=False):
                for c in check_cell:
                    if isinstance(c.value, str):
                        v = str(c.value).strip().upper()
                        for kw in second_header_keywords:
                            if kw in v:
                                is_second_header = True
                                # Map columns from this sub-header row too
                                if "ESTRUCTURA" in v:
                                    sub_header_col_map["ESTRUCTURA"] = c.column
                                if "F'C" in v or ("FC" in v and "IDENTIFICAC" not in v):
                                    sub_header_col_map["FC"] = c.column
                                if "MOLDEO" in v and "FECHA" in v:
                                    sub_header_col_map["FECHA_MOLDEO"] = c.column
                                if "MOLDEO" in v and "HORA" in v:
                                    sub_header_col_map["HORA_MOLDEO"] = c.column
                                if "EDAD" in v:
                                    sub_header_col_map["EDAD"] = c.column
                                if "ROTURA" in v:
                                    sub_header_col_map["FECHA_ROTURA"] = c.column
                                if "DENSIDAD" in v:
                                    sub_header_col_map["DENSIDAD"] = c.column
                                if "CODIGO" in v or "CÓDIGO" in v:
                                    sub_header_col_map["IDENT"] = c.column
                                break
            
            if is_second_header:
                print(f"[DEBUG] Two-row header detected. Sub-header at Row {start_row}. Extra map: {sub_header_col_map}")
                # Merge sub-header columns into main col_map (sub-header takes priority)
                col_map.update(sub_header_col_map)
                start_row += 1  # Skip the second header row
            
            # Default fallback columns if not mapped (legacy behavior)
            # Layout assumption: B=2 (LEM), D=4 (Ident), E=5 (Estructura), F=6 (fc), G=7 (FM), H=8 (HM), I=9 (Edad), J=10 (FR), K=11 (Dens)
            c_lem = col_map.get("LEM", 2)
            c_ident = col_map.get("IDENT", 4)
            c_est = col_map.get("ESTRUCTURA", 5)
            c_fc = col_map.get("FC", 6)
            c_fm = col_map.get("FECHA_MOLDEO", 7)
            c_hm = col_map.get("HORA_MOLDEO", 8)
            c_edad = col_map.get("EDAD", 9)
            c_fr = col_map.get("FECHA_ROTURA", 10)
            c_dens = col_map.get("DENSIDAD", 11)

            # Iterate until empty
            for r in range(start_row, 300):
                lem = get_val(r, c_lem)
                ident = get_val(r, c_ident)
                
                # Stop conditions
                first_col_val = get_val(r, 1) or ""
                # Check for Footer Labels
                fcv_str = str(first_col_val).strip().upper()
                if "NOTA" in fcv_str or "OBSERVACIONES" in fcv_str: break
                
                # Also check if Column 2 (LEM) has "NOTA" or "OBSERVACIONES" (sometimes merged)
                lem_str = safe_str(lem).upper()
                if "NOTA" in lem_str or "OBSERVACIONES" in lem_str: break

                if not lem and not ident:
                    # Check next row to be sure (empty row tolerance)
                    if not get_val(r+1, c_ident): break
                    continue

                m = {
                    "codigo_muestra_lem": safe_str(lem),
                    "identificacion_muestra": safe_str(ident),
                    "estructura": safe_str(get_val(r, c_est)),
                    "fc_kg_cm2": 210, # Default
                    "fecha_moldeo": safe_str(get_val(r, c_fm)),
                    "hora_moldeo": safe_str(get_val(r, c_hm)),
                    "edad": 7, # Default
                    "fecha_rotura": safe_str(get_val(r, c_fr)),
                    "requiere_densidad": "SI" in safe_str(get_val(r, c_dens)).upper()
                }
                
                # Parse numbers safely
                try: 
                    v = get_val(r, c_fc)
                    if v: m["fc_kg_cm2"] = float(v)
                except: pass
                
                try:
                    v = get_val(r, c_edad)
                    if v: m["edad"] = int(v)
                except: pass
                
                muestras.append(m)
        else:
            print("[DEBUG] Table Header NOT FOUND")
        
        data['muestras'] = muestras
        return data
