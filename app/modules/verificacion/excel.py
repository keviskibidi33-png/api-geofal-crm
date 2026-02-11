"""
Servicio para generar archivos Excel de verificación de muestras cilíndricas
"""

import os
import io
import shutil
import logging
import zipfile
import tempfile as tmp
from datetime import datetime
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .models import VerificacionMuestras, MuestraVerificada

logger = logging.getLogger(__name__)

class ExcelLogic:
    """
    Lógica para generar archivos Excel de verificación.
    """
    
    COLUMNS = {
        'numero': 1, 'codigo_lem': 2, 'tipo_testigo': 3, 'diametro_1': 4, 'diametro_2': 5,
        'tolerancia_porcentaje': 6, 'aceptacion_diametro': 7, 'perpendicularidad_sup1': 8,
        'perpendicularidad_sup2': 9, 'perpendicularidad_inf1': 10, 'perpendicularidad_inf2': 11,
        'perpendicularidad_medida': 12, 'planitud_superior': 13, 'planitud_inferior': 14,
        'planitud_depresiones': 15, 'accion': 16, 'conformidad': 17, 'longitud_1': 18,
        'longitud_2': 19, 'longitud_3': 20, 'masa': 21, 'pesar': 22
    }
    
    def __init__(self):
        # Usamos el nuevo nombre del template oficial migrado del backend
        filename = "Template_Verificacion.xlsx"
        from pathlib import Path
        current_dir = Path(__file__).resolve().parent
        app_dir = current_dir.parents[1]
        self.template_path = str(app_dir / "templates" / filename)
        
        # Estilos base
        self.font_normal = Font(name='Arial', size=10)
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    def _copy_style(self, source_cell, target_cell):
        """Copia bordes, fuente y alineación de una celda a otra."""
        if source_cell.has_style:
            # FIX: No usar copy() en objetos que pueden ser StyleProxy (unhashable)
            # Los estilos de openpyxl son inmutables, así que la asignación directa es segura
            # si no planeamos modificarlos después.
            target_cell.font = source_cell.font
            target_cell.border = source_cell.border
            target_cell.fill = source_cell.fill
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection
            target_cell.alignment = source_cell.alignment

    def generar_excel_verificacion(self, verificacion: VerificacionMuestras) -> bytes:
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template no encontrado en {self.template_path}")

        # Cargar directamente desde el template
        wb = load_workbook(self.template_path)
        ws = wb.active
        
        # Mejorar la inserción de filas con replicación de estilos + GAP row
        num_muestras = len(verificacion.muestras_verificadas)
        
        # Calcular cuántas filas extra de datos necesitamos (más allá de las 8 base)
        rows_needed_for_data = max(0, num_muestras - 8)
        
        # Siempre insertamos al menos 1 fila extra para usarla como separador (GAP)
        # o las necesarias para los datos + el separador.
        # "los codigo de equipos deberia tener una linea vacia de diferencia"
        extra_rows_to_insert = rows_needed_for_data + 1
        
        if extra_rows_to_insert > 0:
            # CRITICAL FIX: OpenPyXL bug
            # Detected that merge A21:Q21 DOES NOT start shifting when inserting at row 18.
            # We must manually unmerge any range below row 18 to avoid collisions with new data.
            ranges_to_remove = []
            for merge in ws.merged_cells.ranges:
                if merge.min_row >= 18:
                    ranges_to_remove.append(merge)
            
            for m in ranges_to_remove:
                ws.unmerge_cells(str(m))

            # Insertar filas (Datos extra + Gap)
            ws.insert_rows(18, amount=extra_rows_to_insert)
            
            # Replicar estilos y ALTURA de la fila 10 (referencia)
            # Solo a las filas de DATOS, no a la fila GAP (última insertada)
            source_row_height = ws.row_dimensions[10].height
            
            # Rango de filas de datos nuevas: desde 18 hasta (18 + rows_needed_for_data - 1)
            # Si rows_needed_for_data es 0 (<= 8 items), este loop no corre, y solo queda el gap limpio.
            # FIX: El usuario reportó que el Item 8 (Fila 17) no tenía bordes.
            # Para asegurar consistencia, aplicamos el estilo a TODAS las filas de datos (10 hasta N),
            # no solo a las insertadas.
            last_data_row = 10 + num_muestras - 1
            for r in range(11, last_data_row + 1): # Empezamos en 11 porque 10 es el source
                # Copiar altura (si es insertada o si queremos forzar)
                if source_row_height:
                    ws.row_dimensions[r].height = source_row_height
                
                # Copiar estilos de celda (Bordes, etc.)
                for c in range(1, 23): # Columnas A a V
                    source = ws.cell(row=10, column=c)
                    target = ws.cell(row=r, column=c)
                    self._copy_style(source, target)
        
        # LLenar datos
        self._llenar_datos(ws, verificacion, offset_rows=extra_rows_to_insert)

        # FIX: Limpiar encabezado Masa/Pesar (Mantener fusionado U8:V9)
        # Asegurar que esté fusionado (reparar si el unmerge anterior rompió algo)
        try:
            ws.merge_cells("U8:V9")
        except:
            pass # Ya está fusionado
        
        header_masa_pesar = ws.cell(row=8, column=21)
        header_masa_pesar.value = "Masa muestra aire (g)"
        header_masa_pesar.alignment = self.align_center
        header_masa_pesar.font = ws.cell(row=8, column=1).font if ws.cell(row=8, column=1).font else header_masa_pesar.font

        for r in range(8, 10):
            for c in range(1, 23): # A a V
                cell = ws.cell(row=r, column=c)
                cell.border = self.border_thin
                # No forzamos alignment general aquí para respetar fusiones específicas
                if not cell.alignment:
                    cell.alignment = self.align_center

        # FIX: Mover el footer (Web...) a Columna J y descombinar
        # "agrega el de web y todo ello en columna J sin fusiones"
        # Buscamos la fila del footer. Originalmente es 28, pero se desplaza.
        # Buscamos en las ultimas filas. A veces max_row es engañoso si hay estilos vacios.
        # Buscamos en un rango seguro de 50 filas desde el final (o desde row 20 si es corto)
        max_r = ws.max_row
        start_search = max(20, max_r - 50)
        
        found_footer = False
        for r in range(max_r, start_search, -1): # Buscar desde abajo hacia arriba
            for c in range(1, 15): # Buscar en primeras columnas (A..N)
                cell = ws.cell(row=r, column=c)
                val = str(cell.value) if cell.value else ""
                # "Web: www.geofal.com.pe" o direccion "Av. Marañon"
                if "geofal.com.pe" in val or "Web:" in val or "Marañon" in val:
                    # Encontrado.
                    found_footer = True
                    # 1. Descombinar toda la fila si es necesario (o el rango detectado)
                    ranges_to_remove = []
                    for merge in ws.merged_cells.ranges:
                        if merge.min_row <= r <= merge.max_row:
                            ranges_to_remove.append(merge)
                    for m in ranges_to_remove:
                        ws.unmerge_cells(str(m))
                    
                    # 2. Mover contenido a Columna J (10)
                    # Primero limpiamos donde estaba
                    original_value = val
                    cell.value = None
                    
                    # Escribimos en J
                    target = ws.cell(row=r, column=10)
                    target.value = original_value
                    # Alineación izquierda para que se extienda hacia la derecha si es necesario
                    target.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Romper loops una vez encontrado y movido
                    break
            if found_footer:
                break
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        
        return output.getvalue()

    def _llenar_datos(self, ws, verificacion: VerificacionMuestras, offset_rows: int = 0):
        # Info General
        # Aumentamos el rango de búsqueda para encontrar etiquetas en el encabezado (hasta columna 30)
        self._buscar_y_llenar(ws, "VERIFICADO POR:", verificacion.verificado_por, col_offset=3, max_col=30)
        self._buscar_y_llenar(ws, "FECHA VERIFIC.:", verificacion.fecha_verificacion, max_col=30)
        self._buscar_y_llenar(ws, "CLIENTE:", verificacion.cliente, max_col=30)
        
        # Muestras
        start_row = 10
        
        for i, muestra in enumerate(verificacion.muestras_verificadas, 1):
            row = start_row + i - 1
            self._llenar_fila_muestra(ws, row, i, muestra)
            
        # Equipos y Nota desplazados
        # Si insertamos filas en 18, el bloque de equipos (orig 18) se movió esa cantidad.
        self._llenar_equipos(ws, verificacion, row=18 + offset_rows)
        
        if verificacion.nota:
            self._set_cell_value(ws, 19 + offset_rows, 2, verificacion.nota)

    def _set_cell_value(self, ws, row, col, value, apply_style=True):
        """
        Escribe un valor en una celda, manejando casos de celdas combinadas (MergedCells).
        Escribe siempre en la celda superior izquierda del rango combinado y aplica centrado.
        """
        from openpyxl.cell.cell import MergedCell
        
        # Validar coordenadas
        if row < 1 or col < 1:
            return

        cell = ws.cell(row=row, column=col)
        
        target_cell = cell
        if isinstance(cell, MergedCell):
            # Buscar el rango al que pertenece esta celda
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Escribir en la celda superior izquierda del rango
                    target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        
        target_cell.value = value
        if apply_style:
            # Asegurar centrado y ajuste de texto para que no se corte el contenido largo
            target_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _buscar_y_llenar(self, ws, etiqueta, valor, col_offset=1, max_row=15, max_col=30):
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value and etiqueta in str(cell.value).upper():
                    self._set_cell_value(ws, r, c + col_offset, valor)
                    return

    def _llenar_fila_muestra(self, ws, row, numero, muestra: MuestraVerificada):
        def _fmt(val):
            if val is True: return "CUMPLE"
            if val is False: return "NO CUMPLE"
            # Manejar strings que representan booleanos o el valor directo
            if str(val).upper() in ["TRUE", "CUMPLE", "✓"]: return "CUMPLE"
            if str(val).upper() in ["FALSE", "NO CUMPLE", "✗"]: return "NO CUMPLE"
            return str(val) if val is not None else ""

        self._set_cell_value(ws, row, self.COLUMNS['numero'], numero)
        self._set_cell_value(ws, row, self.COLUMNS['codigo_lem'], muestra.codigo_lem or "")
        self._set_cell_value(ws, row, self.COLUMNS['tipo_testigo'], muestra.tipo_testigo or "")
        self._set_cell_value(ws, row, self.COLUMNS['diametro_1'], muestra.diametro_1_mm)
        self._set_cell_value(ws, row, self.COLUMNS['diametro_2'], muestra.diametro_2_mm)
        self._set_cell_value(ws, row, self.COLUMNS['tolerancia_porcentaje'], (muestra.tolerancia_porcentaje / 100.0) if muestra.tolerancia_porcentaje is not None else 0)
        self._set_cell_value(ws, row, self.COLUMNS['aceptacion_diametro'], _fmt(muestra.aceptacion_diametro))
        
        self._set_cell_value(ws, row, self.COLUMNS['perpendicularidad_sup1'], _fmt(muestra.perpendicularidad_sup1))
        self._set_cell_value(ws, row, self.COLUMNS['perpendicularidad_sup2'], _fmt(muestra.perpendicularidad_sup2))
        self._set_cell_value(ws, row, self.COLUMNS['perpendicularidad_inf1'], _fmt(muestra.perpendicularidad_inf1))
        self._set_cell_value(ws, row, self.COLUMNS['perpendicularidad_inf2'], _fmt(muestra.perpendicularidad_inf2))
        self._set_cell_value(ws, row, self.COLUMNS['perpendicularidad_medida'], _fmt(muestra.perpendicularidad_medida))
        
        self._set_cell_value(ws, row, self.COLUMNS['planitud_superior'], _fmt(muestra.planitud_superior_aceptacion))
        self._set_cell_value(ws, row, self.COLUMNS['planitud_inferior'], _fmt(muestra.planitud_inferior_aceptacion))
        self._set_cell_value(ws, row, self.COLUMNS['planitud_depresiones'], _fmt(muestra.planitud_depresiones_aceptacion))
        
        self._set_cell_value(ws, row, self.COLUMNS['accion'], muestra.accion_realizar)
        self._set_cell_value(ws, row, self.COLUMNS['conformidad'], muestra.conformidad)
        
        self._set_cell_value(ws, row, self.COLUMNS['longitud_1'], muestra.longitud_1_mm)
        self._set_cell_value(ws, row, self.COLUMNS['longitud_2'], muestra.longitud_2_mm)
        self._set_cell_value(ws, row, self.COLUMNS['longitud_3'], muestra.longitud_3_mm)
        self._set_cell_value(ws, row, self.COLUMNS['masa'], muestra.masa_muestra_aire_g)
        self._set_cell_value(ws, row, self.COLUMNS['pesar'], muestra.pesar)

    def _llenar_equipos(self, ws, v: VerificacionMuestras, row=18):
        equipos = [
            (3, v.equipo_bernier), (5, v.equipo_lainas_1), (7, v.equipo_lainas_2),
            (9, v.equipo_escuadra), (11, v.equipo_balanza)
        ]
        for col, val in equipos:
            if val: self._set_cell_value(ws, row, col + 1, val)
