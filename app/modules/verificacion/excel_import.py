import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import io
import re
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

class ExcelImportParser:
    """
    Parsea archivos Excel de Verificación de Muestras Cilíndricas (Template_Verificacion).
    Sigue el formato inverso de la exportación (ExcelLogic).
    """

    def parse_excel(self, file_bytes: bytes) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet: Worksheet = wb.active

        # 1. Parsear Cabecera (Filas 1 a 9)
        # Buscamos valores clave en la cabecera por nombre
        cabecera = {}
        
        # Recorremos celdas superiores para identificar etiquetas
        for row in range(1, 10):
            for col in range(1, 25):
                val = sheet.cell(row=row, column=col).value
                if not val or not isinstance(val, str):
                    continue
                
                normalized = val.strip().upper()
                if "VERIFICADO POR" in normalized:
                    # El valor está corrido a la derecha
                    cabecera["verificado_por"] = sheet.cell(row=row, column=col+3).value
                elif "FECHA DE VERIFICACION" in normalized:
                    raw_date = sheet.cell(row=row, column=col+1).value
                    cabecera["fecha_verificacion"] = self._format_date(raw_date)
                elif "CLIENTE" in normalized:
                    cabecera["cliente"] = sheet.cell(row=row, column=col+1).value

        # Si no se encontraron por etiqueta, intentamos leer posiciones fijas del template original
        # En el template la fecha suele estar en la fila 8, cliente en fila 6, verificado por en fila 7/8
        if "fecha_verificacion" not in cabecera or not cabecera["fecha_verificacion"]:
            # Intentar celda fija si existe
            pass

        # 2. Parsear Muestras (Empieza en fila 10)
        # La tabla de datos va de la columna A (1) a la V (22).
        # Fila 10 a 16 son base, pero pueden ser más si se expandió.
        # Paramos cuando el N° de Item deje de ser numérico o cuando lleguemos a la fila de equipos (identificada por "EQUIPOS UTILIZADOS" o similar).
        muestras = []
        row = 10
        while True:
            item_val = sheet.cell(row=row, column=1).value
            
            # Condición de parada: fila vacía o no numérica en columna A
            if item_val is None:
                # Verificar si es fin de tabla o una fila vacía intermedia
                next_val = sheet.cell(row=row+1, column=1).value
                if next_val is None:
                    break
                row += 1
                continue
            
            # Si encontramos textos de equipos, paramos
            if isinstance(item_val, str) and ("EQUIPOS" in item_val.upper() or "BERNIER" in item_val.upper()):
                break
                
            try:
                item_numero = int(item_val)
            except (ValueError, TypeError):
                # Si no es un número de item, terminamos la tabla de muestras
                break

            # Leer columnas según mapeo de ExcelLogic
            codigo_lem = sheet.cell(row=row, column=2).value
            tipo_testigo = sheet.cell(row=row, column=3).value
            
            diametro_1 = sheet.cell(row=row, column=4).value
            diametro_2 = sheet.cell(row=row, column=5).value
            tolerancia = sheet.cell(row=row, column=6).value
            aceptacion = sheet.cell(row=row, column=7).value
            
            perpend_sup1 = sheet.cell(row=row, column=8).value
            perpend_sup2 = sheet.cell(row=row, column=9).value
            perpend_inf1 = sheet.cell(row=row, column=10).value
            perpend_inf2 = sheet.cell(row=row, column=11).value
            perpend_medida = sheet.cell(row=row, column=12).value
            
            planitud_sup = sheet.cell(row=row, column=13).value
            planitud_inf = sheet.cell(row=row, column=14).value
            planitud_dep = sheet.cell(row=row, column=15).value
            
            accion = sheet.cell(row=row, column=16).value
            conformidad = sheet.cell(row=row, column=17).value
            
            longitud_1 = sheet.cell(row=row, column=18).value
            longitud_2 = sheet.cell(row=row, column=19).value
            longitud_3 = sheet.cell(row=row, column=20).value
            
            masa = sheet.cell(row=row, column=21).value
            pesar = sheet.cell(row=row, column=22).value

            # Ignorar filas vacías
            if not codigo_lem and not diametro_1 and not longitud_1:
                row += 1
                continue

            muestra = {
                "item_numero": item_numero,
                "codigo_lem": str(codigo_lem or "").strip(),
                "tipo_testigo": str(tipo_testigo or "-").strip(),
                "diametro_1_mm": self._to_float(diametro_1),
                "diametro_2_mm": self._to_float(diametro_2),
                "tolerancia_porcentaje": self._to_float(tolerancia) * 100.0 if (self._to_float(tolerancia) is not None and self._to_float(tolerancia) <= 1.0) else self._to_float(tolerancia),
                "aceptacion_diametro": str(aceptacion or "").strip(),
                "perpendicularidad_sup1": self._to_bool(perpend_sup1),
                "perpendicularidad_sup2": self._to_bool(perpend_sup2),
                "perpendicularidad_inf1": self._to_bool(perpend_inf1),
                "perpendicularidad_inf2": self._to_bool(perpend_inf2),
                "perpendicularidad_medida": self._to_bool(perpend_medida),
                "planitud_superior_aceptacion": str(planitud_sup or "-").strip(),
                "planitud_inferior_aceptacion": str(planitud_inf or "-").strip(),
                "planitud_depresiones_aceptacion": str(planitud_dep or "-").strip(),
                "accion_realizar": str(accion or "-").strip(),
                "conformidad": str(conformidad or "-").strip(),
                "longitud_1_mm": self._to_float(longitud_1),
                "longitud_2_mm": self._to_float(longitud_2),
                "longitud_3_mm": self._to_float(longitud_3),
                "masa_muestra_aire_g": self._to_float(masa),
                "pesar": str(pesar or "").strip()
            }
            muestras.append(muestra)
            row += 1

        # 3. Parsear Equipos (Fila de Equipos posterior a las muestras)
        # Buscamos en las filas siguientes a la tabla de muestras
        equipos = {}
        for r in range(row, row + 10):
            # Recorrer columnas buscando las etiquetas de los equipos
            for col in range(1, 20):
                val = sheet.cell(row=r, column=col).value
                if not val or not isinstance(val, str):
                    continue
                
                normalized = val.strip().upper()
                if "BERNIER" in normalized:
                    equipos["equipo_bernier"] = sheet.cell(row=r, column=col+1).value
                elif "LAINAS" in normalized and "1" in normalized:
                    equipos["equipo_lainas_1"] = sheet.cell(row=r, column=col+1).value
                elif "LAINAS" in normalized and "2" in normalized:
                    equipos["equipo_lainas_2"] = sheet.cell(row=r, column=col+1).value
                elif "ESCUADRA" in normalized:
                    equipos["equipo_escuadra"] = sheet.cell(row=r, column=col+1).value
                elif "BALANZA" in normalized:
                    equipos["equipo_balanza"] = sheet.cell(row=r, column=col+1).value
                elif "NOTA" in normalized:
                    equipos["nota"] = sheet.cell(row=r, column=col+1).value

        # Construir y retornar el payload estructurado
        # Nota: los cálculos derivados se replican aquí para que la UI de importación
        # reciba los mismos valores automáticos que el flujo "nuevo".
        return {
            "verificado_por": cabecera.get("verificado_por") or "",
            "fecha_verificacion": cabecera.get("fecha_verificacion") or datetime.now().strftime("%Y-%m-%d"),
            "cliente": cabecera.get("cliente") or "",
            "equipo_bernier": equipos.get("equipo_bernier") or "-",
            "equipo_lainas_1": equipos.get("equipo_lainas_1") or "-",
            "equipo_lainas_2": equipos.get("equipo_lainas_2") or "-",
            "equipo_escuadra": equipos.get("equipo_escuadra") or "-",
            "equipo_balanza": equipos.get("equipo_balanza") or "-",
            "nota": equipos.get("nota") or "",
            "muestras_verificadas": self._apply_derived_calculations(muestras)
        }

    def _apply_derived_calculations(self, muestras: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        def to_bool(v: Any) -> bool:
            if isinstance(v, bool):
                return v
            if isinstance(v, str):
                return v.strip().lower() in ("cumple", "true", "1", "si", "sí", "v", "x")
            return False

        def calc_formula(d1: float, d2: float, tipo: str) -> tuple[float, bool]:
            if d1 == 0:
                raise ValueError("El diámetro 1 no puede ser 0")
            tol = abs(d1 - d2) / d1 * 100
            return round(tol, 2), tol <= 2.0

        def calc_patron(ps: bool, pi: bool, pd: bool) -> str:
            clave = f"{'C' if ps else 'N'}{'C' if pi else 'N'}{'C' if pd else 'N'}"
            return {
                'CCC': '-',
                'NCC': 'NEOPRENO SUPERIOR',
                'CNC': 'NEOPRENO INFERIOR',
                'NNC': 'NEOPRENO SUPERIOR E INFERIOR',
                'CCN': 'CAPEO SUPERIOR E INFERIOR',
                'NCN': 'CAPEO SUPERIOR',
                'CNN': 'CAPEO INFERIOR',
                'NNN': 'CAPEO SUPERIOR E INFERIOR',
            }.get(clave, '-')

        def calc_pesar(d1: Optional[float], d2: Optional[float], l1: Optional[float], l2: Optional[float], l3: Optional[float]) -> str:
            if not d1 or not d2 or not l1 or not l2:
                return ""
            avg_d = (d1 + d2) / 2
            longitudes = [v for v in (l1, l2, l3) if v and v > 0]
            if not longitudes or avg_d <= 0:
                return ""
            avg_l = sum(longitudes) / len(longitudes)
            return "PESAR" if (avg_l / avg_d) < 1.75 else "NO PESAR"

        enriched: List[Dict[str, Any]] = []
        for m in muestras:
            item = dict(m)
            d1 = item.get("diametro_1_mm")
            d2 = item.get("diametro_2_mm")
            if d1 is not None and d2 is not None:
                tol, cumple = calc_formula(float(d1), float(d2), str(item.get("tipo_testigo") or "20x10"))
                item["tolerancia_porcentaje"] = tol
                item["aceptacion_diametro"] = "Cumple" if cumple else "No cumple"

            item["pesar"] = calc_pesar(
                item.get("diametro_1_mm"),
                item.get("diametro_2_mm"),
                item.get("longitud_1_mm"),
                item.get("longitud_2_mm"),
                item.get("longitud_3_mm"),
            ) or item.get("pesar", "")

            manual_action = str(item.get("accion_realizar") or "").strip()
            if not manual_action or manual_action == "-":
                ps = to_bool(item.get("planitud_superior_aceptacion") or item.get("planitud_superior"))
                pi = to_bool(item.get("planitud_inferior_aceptacion") or item.get("planitud_inferior"))
                pd = to_bool(item.get("planitud_depresiones_aceptacion") or item.get("planitud_depresiones"))
                item["accion_realizar"] = calc_patron(ps, pi, pd)
            enriched.append(item)
        return enriched

    def _format_date(self, val: Any) -> Optional[str]:
        if not val:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        
        # Si es un string, normalizar
        text = str(val).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return text

    def _to_float(self, val: Any) -> Optional[float]:
        if val is None or val == "" or val == "-":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _to_bool(self, val: Any) -> Optional[bool]:
        if val is None or val == "" or val == "-":
            return None
        if isinstance(val, bool):
            return val
        text = str(val).strip().upper()
        if text in ("CUMPLE", "TRUE", "SI", "SÍ", "V", "1"):
            return True
        if text in ("NO CUMPLE", "FALSE", "NO", "X", "0"):
            return False
        return None
