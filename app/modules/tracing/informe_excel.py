"""
Generador de Excel de Resumen de Ensayo (Informe Final).
Consolida datos de Recepción + Verificación + Compresión en un solo reporte.
Usa openpyxl sobre el template Template_Informe.xlsx.
"""

import io
import os
import logging
from copy import copy
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)

# Ruta del template — multi-path search como los demás módulos
def _find_template():
    filename = "Template_Informe.xlsx"
    current_dir = Path(__file__).resolve().parent
    app_dir = current_dir.parents[1]  # app/

    possible_paths = [
        app_dir / "templates" / filename,                      # Standard: app/templates/
        Path("/app/templates") / filename,                     # Docker absolute fallback
        current_dir.parents[2] / "app" / "templates" / filename,  # Root/app/templates/
    ]

    for p in possible_paths:
        if p.exists():
            return str(p)

    # Fallback to standard
    return str(app_dir / "templates" / filename)

TEMPLATE_PATH = _find_template()

# Yellow fill for data cells
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
FONT_DATA = Font(name="Arial", size=8)
FONT_DATA_BOLD = Font(name="Arial", size=8, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Data row starts at row 16 in template (3 sample rows: 16, 17, 18)
DATA_START_ROW = 16
TEMPLATE_DATA_ROWS = 3  # Template has 3 yellow rows pre-formatted

# Column mapping for data items (row 16+)
#   A=1: Código LEM
#   B=2: Código cliente  
#   C=3: Diámetro 1
#   D=4: Diámetro 2
#   E=5: Longitud 1
#   F=6: Longitud 2
#   G=7: Longitud 3
#   H=8: Carga Máxima (kN)
#   I=9: Tipo fractura
#   J=10: Masa muestra aire (g)


def _format_date(val) -> str:
    """Formatea una fecha a DD/MM/YYYY."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, str):
        # If already formatted, return as-is
        if "/" in val:
            return val
        # Try ISO parse
        try:
            dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
            return dt.strftime("%d/%m/%Y")
        except Exception:
            return val
    return str(val)


def _copy_cell_style(source_cell, target_cell):
    """Copy style from source to target cell safely."""
    try:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
    except Exception:
        # Fallback: apply default data style
        target_cell.font = FONT_DATA
        target_cell.border = BORDER_THIN
        target_cell.fill = YELLOW_FILL
        target_cell.alignment = ALIGN_CENTER


def generate_informe_excel(data: dict) -> bytes:
    """
    Genera el Excel de Resumen de Ensayo a partir de datos consolidados.
    
    Args:
        data: dict con estructura:
            {
                "cliente": str,
                "direccion": str,
                "proyecto": str,
                "ubicacion": str,
                "recepcion_numero": str,
                "ot_numero": str,
                "estructura": str,
                "fc_kg_cm2": float,
                "fecha_recepcion": str,
                "fecha_moldeo": str,
                "fecha_ensayo_programado": str,
                "fecha_ensayo_real": str,
                "densidad": str/bool,
                "items": [
                    {
                        "codigo_lem": str,
                        "codigo_cliente": str,
                        "diametro_1": float,
                        "diametro_2": float,
                        "longitud_1": float,
                        "longitud_2": float,
                        "longitud_3": float,
                        "carga_maxima": float,
                        "tipo_fractura": str,
                        "masa_muestra_aire": float,
                    }, ...
                ]
            }
    
    Returns:
        bytes del archivo Excel generado.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template no encontrado: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    items = data.get("items", [])
    num_items = len(items)

    # --- 1. Insert extra rows if needed (more than 3 samples) ---
    if num_items > TEMPLATE_DATA_ROWS:
        extra_rows = num_items - TEMPLATE_DATA_ROWS
        # Insert rows after the last template data row
        insert_at = DATA_START_ROW + TEMPLATE_DATA_ROWS
        ws.insert_rows(insert_at, extra_rows)

        # Copy style from the last template row to new rows
        style_source_row = DATA_START_ROW + TEMPLATE_DATA_ROWS - 1  # Row 18 (last yellow row)
        for extra_idx in range(extra_rows):
            target_row = insert_at + extra_idx
            for col in range(1, 11):  # A to J
                source_cell = ws.cell(row=style_source_row, column=col)
                target_cell = ws.cell(row=target_row, column=col)
                _copy_cell_style(source_cell, target_cell)

    # --- 2. Fill header fields ---
    # CLIENTE (B5)
    ws["B5"] = data.get("cliente", "")
    # DIRECCIÓN (B6)
    ws["B6"] = data.get("direccion", "")
    # PROYECTO (B7)
    ws["B7"] = data.get("proyecto", "")
    # UBICACIÓN (B8)
    ws["B8"] = data.get("ubicacion", "")
    # RECEPCIÓN N° (J5)
    ws["J5"] = data.get("recepcion_numero", "")
    # OT N° (J6)
    ws["J6"] = data.get("ot_numero", "")
    # Estructura (B10)
    ws["B10"] = data.get("estructura", "")
    # F'c (B11)
    fc = data.get("fc_kg_cm2")
    ws["B11"] = f"{fc}" if fc else ""
    # Fecha Recepción (J9)
    ws["J9"] = _format_date(data.get("fecha_recepcion"))
    # Fecha Moldeo (J10)
    ws["J10"] = _format_date(data.get("fecha_moldeo"))
    # Fecha ensayo programado (I11 label + J11 value)
    # Sobreescribimos el label del template "Fecha Rotura" → "Fecha ensayo programado"
    ws["I11"] = "Fecha ensayo programado"
    ws["I11"].font = Font(name="Arial", size=8)
    ws["I11"].alignment = ALIGN_LEFT
    fecha_programado = data.get("fecha_ensayo_programado") or data.get("fecha_rotura", "")
    ws["J11"] = _format_date(fecha_programado)
    # Densidad (J13)
    densidad = data.get("densidad")
    if isinstance(densidad, bool):
        ws["J13"] = "Sí" if densidad else "No"
    else:
        ws["J13"] = str(densidad) if densidad else ""

    # --- 3. Fill data items ---
    for idx, item in enumerate(items):
        row = DATA_START_ROW + idx
        ws.cell(row=row, column=1, value=item.get("codigo_lem", ""))        # A: Código LEM
        ws.cell(row=row, column=2, value=item.get("codigo_cliente", ""))     # B: Código cliente
        ws.cell(row=row, column=3, value=item.get("diametro_1"))             # C: Diámetro 1
        ws.cell(row=row, column=4, value=item.get("diametro_2"))             # D: Diámetro 2
        ws.cell(row=row, column=5, value=item.get("longitud_1"))             # E: Longitud 1
        ws.cell(row=row, column=6, value=item.get("longitud_2"))             # F: Longitud 2
        ws.cell(row=row, column=7, value=item.get("longitud_3"))             # G: Longitud 3
        ws.cell(row=row, column=8, value=item.get("carga_maxima"))           # H: Carga Máxima
        ws.cell(row=row, column=9, value=item.get("tipo_fractura", ""))      # I: Tipo fractura
        ws.cell(row=row, column=10, value=item.get("masa_muestra_aire"))     # J: Masa muestra aire

    # --- 4. Footer: Realizado / Fecha ensayo / Revisado ---
    last_data_row = DATA_START_ROW + max(num_items, TEMPLATE_DATA_ROWS) - 1
    footer_row = last_data_row + 2  # Skip one empty row after data

    # Clear old note at B24 if it exists and is now in a different position
    try:
        if ws["B24"].value and "ESTO DATOS" in str(ws["B24"].value):
            ws["B24"] = None
    except Exception:
        pass

    # Realizado (columns A-C)
    cell_realizado = ws.cell(row=footer_row, column=1, value="Realizado:")
    cell_realizado.font = Font(name="Arial", size=9, bold=True)
    cell_realizado.alignment = ALIGN_LEFT
    # Line for signature
    ws.cell(row=footer_row + 1, column=1, value="_________________________")
    ws.cell(row=footer_row + 1, column=1).font = FONT_DATA
    ws.cell(row=footer_row + 1, column=1).alignment = ALIGN_CENTER

    # Fecha ensayo (columns D-G) — actual test date from compression
    cell_fecha = ws.cell(row=footer_row, column=4, value="Fecha ensayo:")
    cell_fecha.font = Font(name="Arial", size=9, bold=True)
    cell_fecha.alignment = ALIGN_LEFT
    fecha_ensayo_real = data.get("fecha_ensayo_real", "")
    if not fecha_ensayo_real and items:
        fecha_ensayo_real = items[0].get("fecha_ensayo", "")
    ws.cell(row=footer_row + 1, column=4, value=_format_date(fecha_ensayo_real))
    ws.cell(row=footer_row + 1, column=4).font = Font(name="Arial", size=9)
    ws.cell(row=footer_row + 1, column=4).alignment = ALIGN_CENTER

    # Revisado (columns H-J)
    cell_revisado = ws.cell(row=footer_row, column=8, value="Revisado:")
    cell_revisado.font = Font(name="Arial", size=9, bold=True)
    cell_revisado.alignment = ALIGN_LEFT
    # Line for signature
    ws.cell(row=footer_row + 1, column=8, value="_________________________")
    ws.cell(row=footer_row + 1, column=8).font = FONT_DATA
    ws.cell(row=footer_row + 1, column=8).alignment = ALIGN_CENTER

    # --- 5. Save to bytes ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
