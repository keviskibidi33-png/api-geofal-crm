from __future__ import annotations

import io
import os
from typing import List, Tuple, Optional
import openpyxl
from openpyxl import Workbook
from sqlalchemy import or_, func
from sqlalchemy.orm import Session

from .models import PublicidadGeofal
from .schemas import PublicidadGeofalCreate, PublicidadGeofalUpdate

class PublicidadGeofalService:
    @staticmethod
    def list_publicidad(
        db: Session,
        *,
        search: Optional[str] = None,
        limit: int = 500,
        offset: int = 0
    ) -> Tuple[int, List[PublicidadGeofal]]:
        """
        Retrieves paginated and filtered tracking records from the database.
        """
        query = db.query(PublicidadGeofal)
        
        # Apply search filter (search across razon_social, contacto, telefono, email)
        if search:
            search_pattern = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    PublicidadGeofal.razon_social_referencial.ilike(search_pattern),
                    PublicidadGeofal.contacto.ilike(search_pattern),
                    PublicidadGeofal.telefono.ilike(search_pattern),
                    PublicidadGeofal.telefono_2.ilike(search_pattern),
                    PublicidadGeofal.correo_referencial.ilike(search_pattern)
                )
            )
            
        # Order by: first by 'id_cliente' ascending
        query = query.order_by(
            PublicidadGeofal.id_cliente.asc().nullslast(),
            PublicidadGeofal.id.asc()
        )
        
        total = query.count()
        items = query.offset(max(0, offset)).limit(max(1, min(limit, 10000))).all()
        return total, items

    @staticmethod
    def get_publicidad(db: Session, id: int) -> Optional[PublicidadGeofal]:
        """
        Finds a single record by ID.
        """
        return db.query(PublicidadGeofal).filter(PublicidadGeofal.id == id).first()

    @staticmethod
    def create_publicidad(
        db: Session,
        *,
        data: PublicidadGeofalCreate,
        creado_por: Optional[str] = None
    ) -> PublicidadGeofal:
        """
        Creates a new tracking record.
        Auto-increments the 'id_cliente' field if not provided.
        """
        if data.id_cliente is None:
            max_id_cliente = db.query(func.max(PublicidadGeofal.id_cliente)).scalar() or 0
            id_cliente_val = max_id_cliente + 1
        else:
            id_cliente_val = data.id_cliente
            
        db_item = PublicidadGeofal(
            id_cliente=id_cliente_val,
            contacto=data.contacto,
            telefono=data.telefono,
            telefono_2=data.telefono_2,
            correo_referencial=data.correo_referencial,
            razon_social_referencial=data.razon_social_referencial,
            
            junio_asistente=data.junio_asistente,
            junio_asesor=data.junio_asesor,
            julio_asistente=data.julio_asistente,
            julio_asesor=data.julio_asesor,
            agosto_asistente=data.agosto_asistente,
            agosto_asesor=data.agosto_asesor,
            setiembre_asistente=data.setiembre_asistente,
            setiembre_asesor=data.setiembre_asesor,
            octubre_asistente=data.octubre_asistente,
            octubre_asesor=data.octubre_asesor,
            noviembre_asistente=data.noviembre_asistente,
            noviembre_asesor=data.noviembre_asesor,
            diciembre_asistente=data.diciembre_asistente,
            diciembre_asesor=data.diciembre_asesor,
            
            observacion_1=data.observacion_1,
            observacion_2=data.observacion_2,
            creado_por=creado_por
        )
        db.add(db_item)
        db.commit()
        db.refresh(db_item)
        return db_item

    @staticmethod
    def patch_publicidad(
        db: Session,
        id: int,
        data: dict
    ) -> Optional[PublicidadGeofal]:
        """
        Partially updates a record (used for inline grid edits).
        """
        db_item = PublicidadGeofalService.get_publicidad(db, id)
        if not db_item:
            return None
            
        # Define allowed fields for patching (excluding metadata)
        allowed_fields = {
            "id_cliente", "contacto", "telefono", "telefono_2",
            "correo_referencial", "razon_social_referencial",
            "junio_asistente", "junio_asesor",
            "julio_asistente", "julio_asesor",
            "agosto_asistente", "agosto_asesor",
            "setiembre_asistente", "setiembre_asesor",
            "octubre_asistente", "octubre_asesor",
            "noviembre_asistente", "noviembre_asesor",
            "diciembre_asistente", "diciembre_asesor",
            "observacion_1", "observacion_2"
        }
        
        for key, val in data.items():
            if key in allowed_fields:
                setattr(db_item, key, val)
                
        db.commit()
        db.refresh(db_item)
        return db_item

    @staticmethod
    def delete_publicidad(db: Session, id: int) -> bool:
        """
        Deletes a record.
        """
        db_item = PublicidadGeofalService.get_publicidad(db, id)
        if not db_item:
            return False
        db.delete(db_item)
        db.commit()
        return True

    @staticmethod
    def crear_plantilla_excel_defecto() -> io.BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hoja1"

        sheet.cell(row=1, column=1).value = "LISTA DE CLIENTES GEOFAL - PUBLICIDAD"
        sheet.cell(row=2, column=1).value = "ID CLIENTE"
        sheet.cell(row=2, column=2).value = "CONTACTO"
        sheet.cell(row=2, column=3).value = "TELÉFONO"
        sheet.cell(row=2, column=4).value = "TELÉFONO 2"
        sheet.cell(row=2, column=5).value = "CORREO REFERENCIAL"
        sheet.cell(row=2, column=6).value = "RAZON SOCIAL REFERENCIAL"
        
        months = ["JUNIO", "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        for idx, month in enumerate(months):
            col_start = 7 + (idx * 2)
            sheet.cell(row=2, column=col_start).value = month
            sheet.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+1)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def importar_excel(db: Session, file_content: bytes, creado_por: Optional[str] = None) -> int:
        """
        Imports database records from the Excel file contents.
        This deletes all existing records in 'publicidad_geofal' and inserts from row 3 onwards.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        
        sheet_name = 'Hoja1' if 'Hoja1' in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        # Clear existing table data to prevent duplicate keys
        db.query(PublicidadGeofal).delete()
        db.commit()
        
        inserted_count = 0
        
        def to_int(val) -> Optional[int]:
            if val is None:
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        def to_str(val) -> Optional[str]:
            if val is None:
                return None
            val_str = str(val).strip()
            if not val_str:
                return None
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return val_str

        # Parse rows starting from row 3
        for r in range(3, sheet.max_row + 1):
            id_cliente_val = sheet.cell(row=r, column=1).value
            contacto_val = sheet.cell(row=r, column=2).value
            telefono_val = sheet.cell(row=r, column=3).value
            telefono_2_val = sheet.cell(row=r, column=4).value
            correo_val = sheet.cell(row=r, column=5).value
            razon_val = sheet.cell(row=r, column=6).value
            
            # Skip if row is completely empty
            if not any([id_cliente_val, contacto_val, telefono_val, correo_val, razon_val]):
                continue
                
            db_item = PublicidadGeofal(
                id_cliente=to_int(id_cliente_val) or (inserted_count + 1),
                contacto=to_str(contacto_val),
                telefono=to_str(telefono_val),
                telefono_2=to_str(telefono_2_val),
                correo_referencial=to_str(correo_val),
                razon_social_referencial=to_str(razon_val),
                
                junio_asistente=to_str(sheet.cell(row=r, column=7).value),
                junio_asesor=to_str(sheet.cell(row=r, column=8).value),
                
                julio_asistente=to_str(sheet.cell(row=r, column=9).value),
                julio_asesor=to_str(sheet.cell(row=r, column=10).value),
                
                agosto_asistente=to_str(sheet.cell(row=r, column=11).value),
                agosto_asesor=to_str(sheet.cell(row=r, column=12).value),
                
                setiembre_asistente=to_str(sheet.cell(row=r, column=13).value),
                setiembre_asesor=to_str(sheet.cell(row=r, column=14).value),
                
                octubre_asistente=to_str(sheet.cell(row=r, column=15).value),
                octubre_asesor=to_str(sheet.cell(row=r, column=16).value),
                
                noviembre_asistente=to_str(sheet.cell(row=r, column=17).value),
                noviembre_asesor=to_str(sheet.cell(row=r, column=18).value),
                
                diciembre_asistente=to_str(sheet.cell(row=r, column=19).value),
                diciembre_asesor=to_str(sheet.cell(row=r, column=20).value),
                
                observacion_1=to_str(sheet.cell(row=r, column=21).value),
                observacion_2=to_str(sheet.cell(row=r, column=22).value),
                
                creado_por=creado_por
            )
            
            db.add(db_item)
            inserted_count += 1
            
            if inserted_count % 100 == 0:
                db.flush()
                
        db.commit()
        return inserted_count

    @staticmethod
    def exportar_excel(db: Session, template_path: str) -> io.BytesIO:
        """
        Loads the template file, populates columns A-V starting from row 3 with database records,
        and returns the result as a BytesIO file.
        """
        if not os.path.exists(template_path):
            workbook_io = PublicidadGeofalService.crear_plantilla_excel_defecto()
            workbook = openpyxl.load_workbook(workbook_io, data_only=False)
        else:
            workbook = openpyxl.load_workbook(template_path, data_only=False)
            
        sheet_name = 'Hoja1' if 'Hoja1' in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        
        # Get all records sorted by 'id_cliente' ascending
        records = db.query(PublicidadGeofal).order_by(
            PublicidadGeofal.id_cliente.asc().nullslast(),
            PublicidadGeofal.id.asc()
        ).all()
        
        # Clean existing template data rows (from row 3 onwards, up to sheet.max_row)
        for r in range(3, max(sheet.max_row + 1, len(records) + 10)):
            for col in range(1, 23):
                sheet.cell(row=r, column=col).value = None

        # Write data keeping styles
        for idx, rec in enumerate(records):
            r = 3 + idx
            sheet.cell(row=r, column=1).value = rec.id_cliente
            sheet.cell(row=r, column=2).value = rec.contacto
            sheet.cell(row=r, column=3).value = rec.telefono
            sheet.cell(row=r, column=4).value = rec.telefono_2
            sheet.cell(row=r, column=5).value = rec.correo_referencial
            sheet.cell(row=r, column=6).value = rec.razon_social_referencial
            
            sheet.cell(row=r, column=7).value = rec.junio_asistente
            sheet.cell(row=r, column=8).value = rec.junio_asesor
            
            sheet.cell(row=r, column=9).value = rec.julio_asistente
            sheet.cell(row=r, column=10).value = rec.julio_asesor
            
            sheet.cell(row=r, column=11).value = rec.agosto_asistente
            sheet.cell(row=r, column=12).value = rec.agosto_asesor
            
            sheet.cell(row=r, column=13).value = rec.setiembre_asistente
            sheet.cell(row=r, column=14).value = rec.setiembre_asesor
            
            sheet.cell(row=r, column=15).value = rec.octubre_asistente
            sheet.cell(row=r, column=16).value = rec.octubre_asesor
            
            sheet.cell(row=r, column=17).value = rec.noviembre_asistente
            sheet.cell(row=r, column=18).value = rec.noviembre_asesor
            
            sheet.cell(row=r, column=19).value = rec.diciembre_asistente
            sheet.cell(row=r, column=20).value = rec.diciembre_asesor
            
            sheet.cell(row=r, column=21).value = rec.observacion_1
            sheet.cell(row=r, column=22).value = rec.observacion_2

        # Save workbook to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
