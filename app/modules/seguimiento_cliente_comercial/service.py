from __future__ import annotations

import io
import os
import csv
import re
import unicodedata
from datetime import date, datetime
from typing import List, Tuple, Optional
import openpyxl
from openpyxl import Workbook
from sqlalchemy import or_, and_, func
from sqlalchemy.orm import Session

from .models import SeguimientoClienteComercial
from .schemas import SeguimientoClienteComercialCreate, SeguimientoClienteComercialUpdate, SeguimientoClienteComercialPatch

# Predefined catalogs from the LISTA sheet to merge with dynamic values
PREDEFINED_ASESORES = ["Silvia Peralta", "Juan Garcia"]
ADVISOR_ALIASES = {
    "SILVIA": "Silvia Peralta",
}
PREDEFINED_CONTACTOS = ["WHATSAPP", "LLAMADA", "CORREO", "EN PROSPECTO"]
PREDEFINED_RUBROS = ["LABORATORIO", "INGENIERÍA", "ALQUILER", "EN ESPERA"]
PREDEFINED_ESTADOS = [
    "EN ESPERA DE ATENCIÓN",
    "SE SOLICITÓ INFORMACIÓN",
    "EN ESPERA DE INFORMACIÓN",
    "NO ENVIÓ LA INFORMACIÓN",
    "DESCARTO EL SERVICIO",
    "COTIZACIÓN REALIZADA",
    "PROSPECTO",
    "CONTACTADO",
]

STATE_ALIASES = {
    "1 SOLICITUD INFORMACION": "SE SOLICITÓ INFORMACIÓN",
    "1. SOLICITUD INFORMACION": "SE SOLICITÓ INFORMACIÓN",
    "SE SOLICITO INFORMACION": "SE SOLICITÓ INFORMACIÓN",
    "2 PROCESANDO INFORMACION": "EN ESPERA DE INFORMACIÓN",
    "2. PROCESANDO INFORMACION": "EN ESPERA DE INFORMACIÓN",
    "INFORMACION RECIBIDA": "EN ESPERA DE INFORMACIÓN",
    "3 COTIZACION": "COTIZACIÓN REALIZADA",
    "3. COTIZACION": "COTIZACIÓN REALIZADA",
    "4 SEG COTIZACION": "COTIZACIÓN REALIZADA",
    "4. SEG. COTIZACION": "COTIZACIÓN REALIZADA",
    "DESCARTO EL SERVICIO": "DESCARTO EL SERVICIO",
    "DESCARTO EL SERVICIO ": "DESCARTO EL SERVICIO",
}

SEG_CLIENTE_HEADERS = [
    "N°",
    "FECHA CONTACTO",
    "PERSONA CONTACTO",
    "CELULAR",
    "EMAIL",
    "RAZÓN SOCIAL",
    "RUC",
    "ASESOR",
    "CONTACTO",
    "RUBRO",
    "ESTADO CLIENTE",
    "SERVICIO SOLICITADO",
    "F. ÚLTIMO CONTACTO",
    "OBSERVACIONES",
    "N° COTIZACIÓN",
    "ESTADO SEGUIMIENTO",
]

class SeguimientoClienteComercialService:
    @staticmethod
    def _normalize_catalog_key(value: object) -> str:
        raw_value = str(value or "").strip()
        if not raw_value:
            return ""
        normalized = unicodedata.normalize("NFKD", raw_value)
        ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
        return " ".join(ascii_value.upper().split())

    @staticmethod
    def _normalize_catalog_value(value: object, allowed_values: list[str], aliases: Optional[dict[str, str]] = None) -> Optional[str]:
        raw_value = str(value or "").strip()
        if not raw_value:
            return None

        normalized_key = SeguimientoClienteComercialService._normalize_catalog_key(raw_value)
        if aliases:
            alias_value = aliases.get(normalized_key)
            if alias_value:
                return alias_value

        for allowed_value in allowed_values:
            if SeguimientoClienteComercialService._normalize_catalog_key(allowed_value) == normalized_key:
                return allowed_value

        return raw_value

    @staticmethod
    def _normalize_tsv_header(value: object) -> str:
        raw_value = str(value or "").strip()
        if not raw_value:
            return ""
        normalized = unicodedata.normalize("NFKD", raw_value)
        ascii_value = normalized.encode("ascii", "ignore").decode("ascii").upper()
        ascii_value = re.sub(r"[^A-Z0-9]+", " ", ascii_value).strip()
        normalized = " ".join(ascii_value.split())
        header_aliases = {
            "N": "NO",
            "N DE COTIZACION": "N COTIZACION",
            "N COTIZACION": "N COTIZACION",
            "NUMERO CELULAR": "CELULAR",
            "NMERO CELULAR": "CELULAR",
            "TELEFONO": "CELULAR",
            "E MAIL": "EMAIL",
            "CORREO ELECTRONICO": "EMAIL",
            "RAZON SOCIAL": "RAZON SOCIAL",
            "F ULTIMO CONTACTO": "F ULTIMO CONTACTO",
            "FECHA ULTIMO CONTACTO": "F ULTIMO CONTACTO",
            "FECHA LTIMO CONTACTO": "F ULTIMO CONTACTO",
            "N DE COTIZACIN": "N COTIZACION",
            "F ULTIMO C": "F ULTIMO CONTACTO",
            "ESTADO DE SEGUIMIENTO": "ESTADO SEGUIMIENTO",
        }
        return header_aliases.get(normalized, normalized)

    @staticmethod
    def _parse_text_date(value: object) -> Optional[date]:
        raw_value = str(value or "").strip()
        if not raw_value:
            return None

        direct_date = SeguimientoClienteComercialService._parse_date_value(raw_value)
        if direct_date:
            return direct_date

        normalized = SeguimientoClienteComercialService._normalize_catalog_key(raw_value)
        month_map = {
            "ENE": 1,
            "FEB": 2,
            "MAR": 3,
            "ABR": 4,
            "MAY": 5,
            "JUN": 6,
            "JUL": 7,
            "AGO": 8,
            "SEP": 9,
            "SET": 9,
            "OCT": 10,
            "NOV": 11,
            "DIC": 12,
        }

        compact_value = normalized.replace(" ", "").replace(".", "")
        match = re.match(r"^(\d{1,2})[-./]?([A-Z]{3,4}|\d{1,2})[-./]?(\d{2,4})$", compact_value)
        if match:
            day_str, month_str, year_str = match.groups()
            if month_str.isdigit():
                month = int(month_str)
            else:
                month = month_map.get(month_str[:3])
            if month:
                year = int(year_str)
                if year < 100:
                    year += 2000 if year < 70 else 1900
                try:
                    return date(year, month, int(day_str))
                except ValueError:
                    return None

        return None

    @staticmethod
    def _parse_date_value(value: object) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            val_str = value.strip()
            if not val_str:
                return None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
        return None

    @staticmethod
    def _build_seguimiento_item_from_values(
        values: dict[str, object],
        *,
        creado_por: Optional[str] = None,
        allow_missing_no: bool = True,
    ) -> Optional[SeguimientoClienteComercial]:
        def to_str(val) -> Optional[str]:
            if val is None:
                return None
            val_str = str(val).strip()
            if not val_str:
                return None
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return val_str

        def to_int(val) -> Optional[int]:
            if val is None:
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        no_val = values.get("no")
        fecha_contacto_val = values.get("fecha_contacto")
        persona_contacto_val = values.get("persona_contacto")
        razon_social_val = values.get("razon_social")
        ruc_val = values.get("ruc")

        if not any([no_val, fecha_contacto_val, persona_contacto_val, razon_social_val, ruc_val]):
            return None

        return SeguimientoClienteComercial(
            no=to_int(no_val) if allow_missing_no or no_val is not None else None,
            fecha_contacto=SeguimientoClienteComercialService._parse_date_value(fecha_contacto_val) or SeguimientoClienteComercialService._parse_text_date(fecha_contacto_val),
            persona_contacto=to_str(persona_contacto_val),
            numero_celular=to_str(values.get("numero_celular")),
            email=to_str(values.get("email")),
            razon_social=to_str(razon_social_val),
            ruc=to_str(ruc_val),
            asesor=SeguimientoClienteComercialService._normalize_catalog_value(values.get("asesor"), PREDEFINED_ASESORES, ADVISOR_ALIASES),
            contacto=SeguimientoClienteComercialService._normalize_catalog_value(values.get("contacto"), PREDEFINED_CONTACTOS),
            rubro=SeguimientoClienteComercialService._normalize_catalog_value(values.get("rubro"), PREDEFINED_RUBROS),
            estado_cliente=SeguimientoClienteComercialService._normalize_catalog_value(values.get("estado_cliente"), PREDEFINED_ESTADOS, STATE_ALIASES),
            servicio_solicitado=to_str(values.get("servicio_solicitado")),
            fecha_ultimo_contacto=SeguimientoClienteComercialService._parse_date_value(values.get("fecha_ultimo_contacto")) or SeguimientoClienteComercialService._parse_text_date(values.get("fecha_ultimo_contacto")),
            observaciones=to_str(values.get("observaciones")),
            numero_cotizacion=to_str(values.get("numero_cotizacion")),
            estado_seguimiento=to_str(values.get("estado_seguimiento")),
            creado_por=creado_por,
        )

    @staticmethod
    def _import_from_text_tsv(db: Session, file_content: bytes, creado_por: Optional[str] = None) -> int:
        text = file_content.decode("utf-8-sig", errors="replace")
        
        # Detect delimiter dynamically
        first_lines = text.splitlines()[:5]
        delimiter = "\t"
        if first_lines:
            counts = {
                "\t": sum(line.count("\t") for line in first_lines),
                ";": sum(line.count(";") for line in first_lines),
                ",": sum(line.count(",") for line in first_lines),
            }
            best_delim = max(counts, key=counts.get)
            if counts[best_delim] > 0:
                delimiter = best_delim

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)

        if not rows:
            return 0

        header_map: dict[str, int] = {}
        data_start_index = 0
        for index, row in enumerate(rows):
            normalized_row = [SeguimientoClienteComercialService._normalize_tsv_header(cell) for cell in row]
            if any(cell == "FECHA CONTACTO" for cell in normalized_row) and any(cell == "ESTADO CLIENTE" for cell in normalized_row):
                for column_index, cell in enumerate(normalized_row):
                    if not cell:
                        continue
                    canonical_header = {
                        "N": "NO",
                        "NO": "NO",
                        "FECHA CONTACTO": "FECHA CONTACTO",
                        "PERSONA CONTACTO": "PERSONA CONTACTO",
                        "CELULAR": "CELULAR",
                        "EMAIL": "EMAIL",
                        "RAZON SOCIAL": "RAZON SOCIAL",
                        "RUC": "RUC",
                        "ASESOR": "ASESOR",
                        "CONTACTO": "CONTACTO",
                        "RUBRO": "RUBRO",
                        "ESTADO CLIENTE": "ESTADO CLIENTE",
                        "SERVICIO SOLICITADO": "SERVICIO SOLICITADO",
                        "F ULTIMO CONTACTO": "F ULTIMO CONTACTO",
                        "OBSERVACIONES": "OBSERVACIONES",
                        "N COTIZACION": "N COTIZACION",
                        "ESTADO SEGUIMIENTO": "ESTADO SEGUIMIENTO",
                    }.get(cell, cell)
                    header_map[canonical_header] = column_index
                data_start_index = index + 1
                break

        if not header_map:
            raise ValueError("El TXT no contiene una fila de encabezados válida para Seguimiento Comercial.")

        def get_value(row: list[str], header: str) -> Optional[str]:
            column_index = header_map.get(header)
            if column_index is None or column_index >= len(row):
                return None
            value = row[column_index].strip()
            return value or None

        db.query(SeguimientoClienteComercial).delete()
        db.commit()

        inserted_count = 0
        for row in rows[data_start_index:]:
            if not any(cell.strip() for cell in row if cell):
                continue

            values = {
                "no": get_value(row, "NO"),
                "fecha_contacto": get_value(row, "FECHA CONTACTO"),
                "persona_contacto": get_value(row, "PERSONA CONTACTO"),
                "numero_celular": get_value(row, "CELULAR"),
                "email": get_value(row, "EMAIL"),
                "razon_social": get_value(row, "RAZON SOCIAL"),
                "ruc": get_value(row, "RUC"),
                "asesor": get_value(row, "ASESOR"),
                "contacto": get_value(row, "CONTACTO"),
                "rubro": get_value(row, "RUBRO"),
                "estado_cliente": get_value(row, "ESTADO CLIENTE"),
                "servicio_solicitado": get_value(row, "SERVICIO SOLICITADO"),
                "fecha_ultimo_contacto": get_value(row, "F ULTIMO CONTACTO"),
                "observaciones": get_value(row, "OBSERVACIONES"),
                "numero_cotizacion": get_value(row, "N COTIZACION"),
                "estado_seguimiento": get_value(row, "ESTADO SEGUIMIENTO"),
            }

            db_item = SeguimientoClienteComercialService._build_seguimiento_item_from_values(values, creado_por=creado_por)
            if not db_item:
                continue

            db.add(db_item)
            inserted_count += 1
            if inserted_count % 100 == 0:
                db.flush()

        db.commit()
        return inserted_count

    @staticmethod
    def crear_plantilla_excel_defecto() -> io.BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SEG.CLIENTE"

        sheet.cell(row=1, column=1).value = "SEGUIMIENTO CLIENTE COMERCIAL"
        sheet.cell(row=2, column=1).value = "Plantilla generada automáticamente"

        for index, header in enumerate(SEG_CLIENTE_HEADERS, start=1):
            sheet.cell(row=4, column=index).value = header

        lista_sheet = workbook.create_sheet("LISTA")
        lista_sheet.cell(row=1, column=1).value = "ASESORES"
        lista_sheet.cell(row=1, column=2).value = "CONTACTOS"
        lista_sheet.cell(row=1, column=3).value = "RUBROS"
        lista_sheet.cell(row=1, column=4).value = "ESTADOS"

        max_len = max(
            len(PREDEFINED_ASESORES),
            len(PREDEFINED_CONTACTOS),
            len(PREDEFINED_RUBROS),
            len(PREDEFINED_ESTADOS),
        )

        for row_index in range(max_len):
            if row_index < len(PREDEFINED_ASESORES):
                lista_sheet.cell(row=row_index + 2, column=1).value = PREDEFINED_ASESORES[row_index]
            if row_index < len(PREDEFINED_CONTACTOS):
                lista_sheet.cell(row=row_index + 2, column=2).value = PREDEFINED_CONTACTOS[row_index]
            if row_index < len(PREDEFINED_RUBROS):
                lista_sheet.cell(row=row_index + 2, column=3).value = PREDEFINED_RUBROS[row_index]
            if row_index < len(PREDEFINED_ESTADOS):
                lista_sheet.cell(row=row_index + 2, column=4).value = PREDEFINED_ESTADOS[row_index]

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def listar_seguimientos(
        db: Session,
        *,
        search: Optional[str] = None,
        asesor: Optional[str] = None,
        estado_cliente: Optional[str] = None,
        limit: int = 100,
        offset: int = 0
    ) -> Tuple[int, List[SeguimientoClienteComercial]]:
        """
        Retrieves paginated and filtered tracking records from the database.
        """
        query = db.query(SeguimientoClienteComercial)
        
        # Apply search filter (search across razon_social, persona_contacto, ruc, email, celular, n_cotizacion)
        if search:
            search_pattern = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    SeguimientoClienteComercial.razon_social.ilike(search_pattern),
                    SeguimientoClienteComercial.persona_contacto.ilike(search_pattern),
                    SeguimientoClienteComercial.ruc.ilike(search_pattern),
                    SeguimientoClienteComercial.email.ilike(search_pattern),
                    SeguimientoClienteComercial.numero_celular.ilike(search_pattern),
                    SeguimientoClienteComercial.numero_cotizacion.ilike(search_pattern)
                )
            )
            
        # Apply advisor filter
        if asesor:
            query = query.filter(SeguimientoClienteComercial.asesor.ilike(asesor.strip()))
            
        # Apply state filter
        if estado_cliente:
            query = query.filter(SeguimientoClienteComercial.estado_cliente.ilike(estado_cliente.strip()))
            
        # Order by: first by 'no' descending or 'id' descending to show newest first
        query = query.order_by(
            SeguimientoClienteComercial.no.desc().nullslast(),
            SeguimientoClienteComercial.id.desc()
        )
        
        total = query.count()
        items = query.offset(max(0, offset)).limit(max(1, min(limit, 200))).all()
        return total, items

    @staticmethod
    def obtener_seguimiento(db: Session, id: int) -> Optional[SeguimientoClienteComercial]:
        """
        Finds a single record by ID.
        """
        return db.query(SeguimientoClienteComercial).filter(SeguimientoClienteComercial.id == id).first()

    @staticmethod
    def crear_seguimiento(
        db: Session,
        *,
        data: SeguimientoClienteComercialCreate,
        creado_por: Optional[str] = None
    ) -> SeguimientoClienteComercial:
        """
        Creates a new tracking record.
        Auto-increments the 'no' field based on the maximum 'no' currently in the table.
        """
        max_no = db.query(func.max(SeguimientoClienteComercial.no)).scalar() or 0
        
        db_item = SeguimientoClienteComercial(
            no=max_no + 1,
            fecha_contacto=data.fecha_contacto or date.today(),
            persona_contacto=data.persona_contacto,
            numero_celular=data.numero_celular,
            email=data.email,
            razon_social=data.razon_social,
            ruc=data.ruc,
            asesor=data.asesor,
            contacto=data.contacto,
            rubro=data.rubro,
            estado_cliente=data.estado_cliente,
            servicio_solicitado=data.servicio_solicitado,
            fecha_ultimo_contacto=data.fecha_ultimo_contacto,
            observaciones=data.observaciones,
            numero_cotizacion=data.numero_cotizacion,
            estado_seguimiento=data.estado_seguimiento,
            creado_por=creado_por
        )
        db.add(db_item)
        db.commit()
        db.refresh(db_item)
        return db_item

    @staticmethod
    def actualizar_seguimiento(
        db: Session,
        id: int,
        data: SeguimientoClienteComercialUpdate
    ) -> Optional[SeguimientoClienteComercial]:
        """
        Updates an entire record.
        """
        db_item = SeguimientoClienteComercialService.obtener_seguimiento(db, id)
        if not db_item:
            return None
            
        update_data = data.dict(exclude_unset=True)
        for key, val in update_data.items():
            setattr(db_item, key, val)
            
        db.commit()
        db.refresh(db_item)
        return db_item

    @staticmethod
    def patch_seguimiento(
        db: Session,
        id: int,
        data: dict
    ) -> Optional[SeguimientoClienteComercial]:
        """
        Partially updates a record (used for inline grid edits).
        """
        db_item = SeguimientoClienteComercialService.obtener_seguimiento(db, id)
        if not db_item:
            return None
            
        # Define allowed fields for patching (excluding metadata)
        allowed_fields = {
            "no", "fecha_contacto", "persona_contacto", "numero_celular",
            "email", "razon_social", "ruc", "asesor", "contacto", "rubro",
            "estado_cliente", "servicio_solicitado", "fecha_ultimo_contacto",
            "observaciones", "numero_cotizacion", "estado_seguimiento"
        }
        
        for key, val in data.items():
            if key in allowed_fields:
                # Type conversions if necessary
                if key in ("fecha_contacto", "fecha_ultimo_contacto") and val:
                    if isinstance(val, str):
                        try:
                            val = datetime.strptime(val.split("T")[0], "%Y-%m-%d").date()
                        except ValueError:
                            val = None
                setattr(db_item, key, val)
                
        db.commit()
        db.refresh(db_item)
        return db_item

    @staticmethod
    def eliminar_seguimiento(db: Session, id: int) -> bool:
        """
        Deletes a record.
        """
        db_item = SeguimientoClienteComercialService.obtener_seguimiento(db, id)
        if not db_item:
            return False
        db.delete(db_item)
        db.commit()
        return True

    @staticmethod
    def obtener_catalogos(db: Session) -> dict:
        """
        Returns merged distinct catalog values for autocomplete dropdowns.
        """
        # Query distinct values from db
        db_asesores = db.query(SeguimientoClienteComercial.asesor).distinct().all()
        db_contactos = db.query(SeguimientoClienteComercial.contacto).distinct().all()
        db_rubros = db.query(SeguimientoClienteComercial.rubro).distinct().all()
        db_estados = db.query(SeguimientoClienteComercial.estado_cliente).distinct().all()

        # Merge utility helper
        def merge_catalogs(predefined: list, db_results: list, aliases: Optional[dict[str, str]] = None) -> list[str]:
            merged: list[str] = []
            seen: set[str] = set()

            def add_value(raw_value: object) -> None:
                value = SeguimientoClienteComercialService._normalize_catalog_value(raw_value, predefined, aliases)
                if value and value not in seen:
                    seen.add(value)
                    merged.append(value)

            for item in predefined:
                add_value(item)

            # Add database results
            for row in db_results:
                add_value(row[0])

            return merged

        return {
            "asesores": merge_catalogs(PREDEFINED_ASESORES, db_asesores, ADVISOR_ALIASES),
            "contactos": merge_catalogs(PREDEFINED_CONTACTOS, db_contactos),
            "rubros": merge_catalogs(PREDEFINED_RUBROS, db_rubros),
            "estados": merge_catalogs(PREDEFINED_ESTADOS, db_estados, STATE_ALIASES),
        }

    @staticmethod
    def importar_excel(db: Session, file_content: bytes, creado_por: Optional[str] = None) -> int:
        """
        Imports database records from the Excel file contents.
        This deletes all existing records in 'seguimiento_cliente_comercial' and inserts from row 5 onwards.
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception:
            return SeguimientoClienteComercialService._import_from_text_tsv(db, file_content, creado_por=creado_por)

        if 'SEG.CLIENTE' not in wb.sheetnames:
            raise ValueError("La hoja 'SEG.CLIENTE' no fue encontrada en el archivo de Excel.")
            
        sheet = wb['SEG.CLIENTE']
        
        # Clear existing table data to prevent duplicate keys
        db.query(SeguimientoClienteComercial).delete()
        db.commit()
        
        inserted_count = 0
        
        def normalize_catalog(val: object, allowed_values: list[str], aliases: Optional[dict[str, str]] = None) -> Optional[str]:
            return SeguimientoClienteComercialService._normalize_catalog_value(val, allowed_values, aliases)
            
        def to_int(val) -> Optional[int]:
            if val is None:
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        def to_str(val) -> Optional[str]:
            if val is None:
                return None
            val_str = str(val).strip()
            if not val_str:
                return None
            # Handle float representation (e.g. RUC, cellular converted to float with .0)
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return val_str

        def to_date(val) -> Optional[date]:
            return SeguimientoClienteComercialService._parse_date_value(val) or SeguimientoClienteComercialService._parse_text_date(val)

        # Parse rows starting from row 5
        for r in range(5, sheet.max_row + 1):
            no_val = sheet.cell(row=r, column=1).value
            fecha_contacto_val = sheet.cell(row=r, column=2).value
            persona_contacto_val = sheet.cell(row=r, column=3).value
            celular_val = sheet.cell(row=r, column=4).value
            email_val = sheet.cell(row=r, column=5).value
            razon_social_val = sheet.cell(row=r, column=6).value
            ruc_val = sheet.cell(row=r, column=7).value
            asesor_val = sheet.cell(row=r, column=8).value
            contacto_val = sheet.cell(row=r, column=9).value
            rubro_val = sheet.cell(row=r, column=10).value
            estado_cliente_val = sheet.cell(row=r, column=11).value
            servicio_val = sheet.cell(row=r, column=12).value
            fecha_ultimo_val = sheet.cell(row=r, column=13).value
            observaciones_val = sheet.cell(row=r, column=14).value
            cotizacion_val = sheet.cell(row=r, column=15).value
            estado_seg_val = sheet.cell(row=r, column=16).value
            
            # If the row is entirely empty, skip it. Specifically check if crucial fields are missing
            if not any([no_val, fecha_contacto_val, persona_contacto_val, razon_social_val, ruc_val]):
                continue
                
            db_item = SeguimientoClienteComercialService._build_seguimiento_item_from_values(
                {
                    "no": no_val,
                    "fecha_contacto": fecha_contacto_val,
                    "persona_contacto": persona_contacto_val,
                    "numero_celular": celular_val,
                    "email": email_val,
                    "razon_social": razon_social_val,
                    "ruc": ruc_val,
                    "asesor": asesor_val,
                    "contacto": contacto_val,
                    "rubro": rubro_val,
                    "estado_cliente": estado_cliente_val,
                    "servicio_solicitado": servicio_val,
                    "fecha_ultimo_contacto": fecha_ultimo_val,
                    "observaciones": observaciones_val,
                    "numero_cotizacion": cotizacion_val,
                    "estado_seguimiento": estado_seg_val,
                },
                creado_por=creado_por,
            )
            if not db_item:
                continue
            db.add(db_item)
            inserted_count += 1
            
            # Flush periodically to avoid huge memory use
            if inserted_count % 100 == 0:
                db.flush()
                
        db.commit()
        return inserted_count

    @staticmethod
    def exportar_excel(db: Session, template_path: str) -> io.BytesIO:
        """
        Loads the template file, populates columns A-P starting from row 5 with database records,
        and returns the result as a BytesIO file.
        """
        if not os.path.exists(template_path):
            workbook_io = SeguimientoClienteComercialService.crear_plantilla_excel_defecto()
            workbook = openpyxl.load_workbook(workbook_io, data_only=False)
        else:
            workbook = openpyxl.load_workbook(template_path, data_only=False)
            
        if 'SEG.CLIENTE' not in workbook.sheetnames:
            raise ValueError("La hoja 'SEG.CLIENTE' no fue encontrada en el template de Excel.")
            
        sheet = workbook['SEG.CLIENTE']
        
        # Get all records sorted by 'no' ascending
        records = db.query(SeguimientoClienteComercial).order_by(
            SeguimientoClienteComercial.no.asc().nullslast(),
            SeguimientoClienteComercial.id.asc()
        ).all()
        
        # Clean existing template data rows (from row 5 onwards, up to sheet.max_row)
        for r in range(5, max(sheet.max_row + 1, len(records) + 10)):
            for col in range(1, 17):
                sheet.cell(row=r, column=col).value = None

        # Write data keeping styles
        for idx, rec in enumerate(records):
            r = 5 + idx
            sheet.cell(row=r, column=1).value = rec.no
            sheet.cell(row=r, column=2).value = rec.fecha_contacto
            sheet.cell(row=r, column=3).value = rec.persona_contacto
            sheet.cell(row=r, column=4).value = rec.numero_celular
            sheet.cell(row=r, column=5).value = rec.email
            sheet.cell(row=r, column=6).value = rec.razon_social
            sheet.cell(row=r, column=7).value = rec.ruc
            sheet.cell(row=r, column=8).value = rec.asesor
            sheet.cell(row=r, column=9).value = rec.contacto
            sheet.cell(row=r, column=10).value = rec.rubro
            sheet.cell(row=r, column=11).value = rec.estado_cliente
            sheet.cell(row=r, column=12).value = rec.servicio_solicitado
            sheet.cell(row=r, column=13).value = rec.fecha_ultimo_contacto
            sheet.cell(row=r, column=14).value = rec.observaciones
            sheet.cell(row=r, column=15).value = rec.numero_cotizacion
            sheet.cell(row=r, column=16).value = rec.estado_seguimiento

        # Save workbook to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
