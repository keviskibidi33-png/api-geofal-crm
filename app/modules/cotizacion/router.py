from fastapi import APIRouter, HTTPException, Response, UploadFile, File, Query
from psycopg2.extras import RealDictCursor
from datetime import date, datetime
from typing import List, Optional
import io
import json
import asyncio
from pathlib import Path

from .schemas import QuoteExportRequest, NextNumberResponse
from .service import (
    _has_database_url, _get_connection, _ensure_sequence_table, 
    _next_quote_sequential, _save_quote_to_folder, register_quote_in_db, 
    _upload_to_supabase_storage, _delete_from_supabase_storage,
    QUOTES_FOLDER, get_condiciones_textos,
    update_quote_db, _get_safe_filename
)
from .excel import _get_template_path, generate_quote_excel

router = APIRouter(prefix="", tags=["Cotizaciones"]) # Prefix empty to match legacy /export paths if needed, but we should probably use /quotes

# Legacy paths support
@router.post("/quote/next-number")
async def get_next_quote_number():
    """Obtiene el siguiente número de cotización"""
    try:
        if _has_database_url():
            _ensure_sequence_table()
        
        year = date.today().year
        if _has_database_url():
            sequential = _next_quote_sequential(year)
            number = f"{sequential:03d}"
        else:
            number = "001"
        
        year_suffix = str(year)[-2:]
        token = f"{number}-{year_suffix}"
        return {"number": number, "year": year, "token": token}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/export")
async def export_quote(payload: QuoteExportRequest) -> Response:
    try:
        # Prepare Data for Excel Generation
        template_path = _get_template_path(payload.template_id)
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        # Generar número de cotización if missing
        fecha_emision = payload.fecha_emision or date.today()
        cotizacion_numero = payload.cotizacion_numero
        if not cotizacion_numero:
            if _has_database_url():
                try:
                    sequential = _next_quote_sequential(fecha_emision.year)
                    cotizacion_numero = f"{sequential:03d}"
                except Exception:
                    cotizacion_numero = "000"
            else:
                cotizacion_numero = "000"
        
        # Load conditions texts
        condiciones_textos = get_condiciones_textos(payload.condiciones_ids) if payload.condiciones_ids else []

        # Prepare dict for export_xlsx_direct
        export_data = {
            'cotizacion_numero': cotizacion_numero,
            'fecha_emision': fecha_emision,
            'cliente': payload.cliente or '',
            'ruc': payload.ruc or '',
            'contacto': payload.contacto or '',
            'telefono': payload.telefono_contacto or '',
            'telefono_contacto': payload.telefono_contacto or '', # Add for safety
            'email': payload.correo or '',
            'correo': payload.correo_vendedor or payload.correo or '', 
            'fecha_solicitud': payload.fecha_solicitud,
            'proyecto': payload.proyecto or '',
            'ubicacion': payload.ubicacion or '',
            'personal_comercial': payload.personal_comercial or '',
            'telefono_comercial': payload.telefono_comercial or '',
            'plazo_dias': payload.plazo_dias or 0,
            'condicion_pago': payload.condicion_pago or '',
            'condiciones_ids': payload.condiciones_ids or [],
            'condiciones_textos': condiciones_textos,
            'items': [
                {
                    'codigo': item.codigo,
                    'descripcion': item.descripcion,
                    'norma': item.norma,
                    'acreditado': item.acreditado,
                    'costo_unitario': item.costo_unitario,
                    'cantidad': item.cantidad,
                }
                for item in payload.items
            ],
            'include_igv': payload.include_igv,
            'igv_rate': payload.igv_rate,
        }

        # Generate Excel
        xlsx_bytes = generate_quote_excel(payload)
        
        # Save mechanics
        year = fecha_emision.year
        
        # Save to local folder
        xlsx_bytes.seek(0)
        filepath = _save_quote_to_folder(
            io.BytesIO(xlsx_bytes.read()), 
            cotizacion_numero, 
            year, 
            payload.cliente or "SinCliente"
        )
        
        # Register in DB and Upload (Sanitized path for cloud)
        safe_cliente_cloud = _get_safe_filename(payload.cliente or "S-N", None)
        cloud_path = f"{year}/COT-{year}-{cotizacion_numero}-{safe_cliente_cloud}.xlsx"
        db_registered = False
        
        # --- Step 1: DB Registration (critical) ---
        try:
            await asyncio.to_thread(
                register_quote_in_db,
                cotizacion_numero, year, payload.cliente, str(filepath), payload,
                cloud_path,  # object_key kwarg
            )
            db_registered = True
        except Exception as e:
            import traceback
            print(f"[COT-{year}-{cotizacion_numero}] DB Registration FAILED: {e}")
            traceback.print_exc()
            # Retry once
            try:
                print(f"[COT-{year}-{cotizacion_numero}] RETRY DB registration...")
                await asyncio.to_thread(
                    register_quote_in_db,
                    cotizacion_numero, year, payload.cliente, str(filepath), payload,
                    cloud_path,
                )
                db_registered = True
                print(f"[COT-{year}-{cotizacion_numero}] RETRY DB registration OK")
            except Exception as retry_err:
                print(f"[COT-{year}-{cotizacion_numero}] RETRY FAILED: {retry_err}")
        
        # --- Step 2: Storage Upload (non-critical, best-effort) ---
        try:
            xlsx_bytes.seek(0)
            await asyncio.to_thread(
                _upload_to_supabase_storage, xlsx_bytes, "cotizaciones", cloud_path
            )
        except Exception as e:
            print(f"[COT-{year}-{cotizacion_numero}] Storage upload error (non-critical): {e}")

        # Return Response with registration status header
        xlsx_bytes.seek(0)
        headers = {
            "Content-Disposition": f'attachment; filename="COT-{year}-{cotizacion_numero}.xlsx"',
            "X-DB-Registered": str(db_registered).lower(),
        }
        return Response(
            content=xlsx_bytes.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/export/xlsx")
async def export_quote_xlsx(payload: QuoteExportRequest) -> Response:
    """Alias para /export"""
    return await export_quote(payload)

@router.get("/by-token/{token}")
async def get_quote_by_token(token: str):
    """
    Get quote by token (e.g., '123-26' -> Number 123, Year 2026)
    """
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")

    # Parse token
    try:
        parts = token.split('-')
        if len(parts) != 2:
            raise ValueError("Invalid token format")
        
        number = parts[0].zfill(3) # Ensure 3 digits for querying
        year_suffix = parts[1]
        
        # Validate parts
        if not number.isdigit() or not year_suffix.isdigit():
             raise ValueError("Token parts must be numeric")

        # Assume 20xx
        year = int(f"20{year_suffix}")
        
    except ValueError:
         raise HTTPException(status_code=400, detail="Invalid token format (Expected NNN-YY)")

    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, numero, year, cliente_nombre as cliente, cliente_ruc as ruc, 
                       cliente_contacto as contacto, cliente_telefono as telefono, 
                       cliente_email as email,
                       proyecto, ubicacion, items_json
                FROM cotizaciones
                WHERE numero = %s AND year = %s
                LIMIT 1
            """, (number, year))
            
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            return {"data": dict(row), "success": True}
    except Exception as e:
        print(f"Error fetching quote by token: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/quotes")
async def list_quotes(year: int = None, limit: int = 50):
    """Lista las cotizaciones guardadas"""
    quotes = []
    if _has_database_url():
        conn = _get_connection()
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                if year:
                    cur.execute("""
                        SELECT 
                            id, numero, year, cliente_nombre, cliente_ruc, proyecto, 
                            total, estado, moneda, fecha_emision, archivo_path as filepath, 
                            created_at, cliente_id, proyecto_id, vendedor_id
                        FROM cotizaciones
                        WHERE year = %s
                        ORDER BY created_at DESC
                        LIMIT %s
                    """, (year, limit))
                else:
                    cur.execute("""
                        SELECT 
                            id, numero, year, cliente_nombre, cliente_ruc, proyecto, 
                            total, estado, moneda, fecha_emision, archivo_path as filepath, 
                            created_at, cliente_id, proyecto_id, vendedor_id
                        FROM cotizaciones
                        ORDER BY created_at DESC
                        LIMIT %s
                    """, (limit,))
                quotes = [dict(row) for row in cur.fetchall()]
        finally:
            conn.close()
    else:
        # File based fallback
        target_year = year or date.today().year
        year_folder = QUOTES_FOLDER / str(target_year)
        if year_folder.exists():
            for f in sorted(year_folder.glob("*.xlsx"), reverse=True)[:limit]:
                quotes.append({
                    "filename": f.name,
                    "filepath": str(f),
                    "year": target_year,
                    "created_at": f.stat().st_mtime
                })
    return {"quotes": quotes, "total": len(quotes)}

@router.get("/quotes/{quote_id}/download")
async def download_quote(quote_id: str):
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT archivo_path FROM cotizaciones WHERE id = %s", (quote_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            filepath = Path(row['archivo_path'])
            if not filepath.exists():
                raise HTTPException(status_code=404, detail="File not found")
            
            with open(filepath, 'rb') as f:
                content = f.read()
            
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filepath.name}"'},
            )
    finally:
        conn.close()

@router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str):
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT archivo_path as filepath FROM cotizaciones WHERE id = %s", (quote_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            filepath = Path(row['filepath'])
            if filepath.exists():
                try:
                    filepath.unlink()
                except:
                    pass
            
            cur.execute("DELETE FROM cotizaciones WHERE id = %s", (quote_id,))
            conn.commit()
            return {"success": True, "message": "Quote deleted successfully"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/quotes/{quote_id}")
async def get_quote_by_id(quote_id: str):
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, numero, year, cliente_nombre as cliente, cliente_ruc as ruc, cliente_contacto as contacto, 
                       cliente_telefono as telefono_contacto, cliente_email as correo,
                       proyecto, ubicacion, personal_comercial, telefono_comercial, 
                       items_json, total, estado, moneda, fecha_emision, fecha_solicitud, 
                       proyecto_id, cliente_id, plazo_dias, condicion_pago, condiciones_ids, correo_vendedor,
                       include_igv, created_at, updated_at
                FROM cotizaciones
                WHERE id = %s
            """, (quote_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            data = dict(row)
            val = data.get('condiciones_ids')
            if isinstance(val, str):
                data['condiciones_ids'] = [v.strip() for v in val.strip("{}").split(",") if v.strip()]
            elif val is None:
                data['condiciones_ids'] = []
            elif isinstance(val, list):
                data['condiciones_ids'] = [str(v) for v in val if v]
            
            return {"data": data, "success": True}
    finally:
        conn.close()

@router.put("/quotes/{quote_id}")
async def update_quote(quote_id: str, payload: QuoteExportRequest):
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        # 1. Existing check
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT numero, year, archivo_path as filepath, object_key FROM cotizaciones WHERE id = %s", (quote_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Quote not found")
        
        # 2. Texts
        condiciones_textos = get_condiciones_textos(payload.condiciones_ids) if payload.condiciones_ids else []
        
        # 3. Generate Excel
        export_data = {
            "cotizacion_numero": existing['numero'],
            "fecha_emision": payload.fecha_emision or datetime.now().strftime("%Y-%m-%d"),
            "fecha_solicitud": payload.fecha_solicitud or datetime.now().strftime("%Y-%m-%d"),
            "cliente": payload.cliente or "",
            "ruc": payload.ruc or "",
            "contacto": payload.contacto or "",
            "telefono": payload.telefono_contacto or "",
            "telefono_contacto": payload.telefono_contacto or "",
            "email": payload.correo or "",
            "correo": payload.correo_vendedor or payload.correo or "",
            "proyecto": payload.proyecto or "",
            "ubicacion": payload.ubicacion or "",
            "personal_comercial": payload.personal_comercial or "",
            "telefono_comercial": payload.telefono_comercial or "",
            "plazo_dias": payload.plazo_dias or 0,
            "condicion_pago": payload.condicion_pago or "",
            "condiciones_textos": condiciones_textos,
            "items": [
                {
                    "codigo": it.codigo or "",
                    "descripcion": it.descripcion or "",
                    "norma": it.norma or "",
                    "acreditado": it.acreditado or "NO",
                    "cantidad": it.cantidad,
                    "costo_unitario": it.costo_unitario,
                }
                for it in payload.items
            ],
            'include_igv': payload.include_igv,
            'igv_rate': payload.igv_rate,
        }
        
        # 3. Generate Excel
        xlsx_bytes = generate_quote_excel(payload)
        
        # 4. Save file physics
        if existing['filepath']:
            fp = Path(existing['filepath'])
            if fp.exists():
                try: fp.unlink() 
                except: pass
        
        year = existing['year']
        year_folder = QUOTES_FOLDER / str(year)
        year_folder.mkdir(parents=True, exist_ok=True)
        filepath = year_folder / f"COT-{year}-{existing['numero']}.xlsx"
        
        xlsx_bytes.seek(0)
        with open(filepath, "wb") as f:
            f.write(xlsx_bytes.read())
            
        # 5. DB Update (Sanitize key)
        safe_cliente_cloud = _get_safe_filename(payload.cliente or "S-N", None)
        cloud_path = f"{year}/COT-{year}-{existing['numero']}-{safe_cliente_cloud}.xlsx"
        
        await asyncio.to_thread(update_quote_db, quote_id, payload, str(filepath), cloud_path)
        
        # 6. Storage Update (best-effort)
        try:
            xlsx_bytes.seek(0)
            await asyncio.to_thread(
                _upload_to_supabase_storage, xlsx_bytes, "cotizaciones", cloud_path
            )
        except Exception as e:
            print(f"Error updating storage: {e}")
                
        return {"success": True, "message": "Quote updated successfully", "quote_id": quote_id}
            
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# Plantillas Endpoints
@router.get("/plantillas")
async def get_plantillas(vendedor_id: str):
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, nombre, descripcion, items_json, condiciones_ids,
                       plazo_dias, condicion_pago, veces_usada, created_at, updated_at
                FROM plantillas_cotizacion
                WHERE vendedor_id = %s AND activo = true
                ORDER BY veces_usada DESC, nombre ASC
            """, (vendedor_id,))
            return [dict(p) for p in cur.fetchall()]
    finally:
        conn.close()

@router.post("/plantillas")
async def create_plantilla(payload: dict):
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO plantillas_cotizacion 
                (nombre, descripcion, vendedor_id, items_json, condiciones_ids, 
                 plazo_dias, condicion_pago)
                VALUES (%s, %s, %s, %s, %s::uuid[], %s, %s)
                RETURNING id
            """, (
                payload.get('nombre'),
                payload.get('descripcion'),
                payload.get('vendedor_id'),
                json.dumps(payload.get('items'), ensure_ascii=False),
                payload.get('condiciones_ids', []),
                payload.get('plazo_dias'),
                payload.get('condicion_pago')
            ))
            pid = cur.fetchone()[0]
            conn.commit()
            return {"success": True, "plantilla_id": str(pid)}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/plantillas/{plantilla_id}")
async def get_plantilla(plantilla_id: str):
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, nombre, descripcion, items_json, condiciones_ids,
                       plazo_dias, condicion_pago, veces_usada
                FROM plantillas_cotizacion
                WHERE id = %s AND activo = true
            """, (plantilla_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Plantilla no encontrada")
            
            # Increment usage
            cur.execute("UPDATE plantillas_cotizacion SET veces_usada = veces_usada + 1 WHERE id = %s", (plantilla_id,))
            conn.commit()
            return dict(row)
    finally:
        conn.close()

@router.put("/plantillas/{plantilla_id}")
async def update_plantilla(plantilla_id: str, payload: dict):
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE plantillas_cotizacion
                SET nombre = %s, descripcion = %s, items_json = %s, 
                    condiciones_ids = %s::uuid[], plazo_dias = %s, condicion_pago = %s,
                    updated_at = NOW()
                WHERE id = %s
                RETURNING id
            """, (
                payload.get('nombre'),
                payload.get('descripcion'),
                json.dumps(payload.get('items'), ensure_ascii=False),
                payload.get('condiciones_ids', []),
                payload.get('plazo_dias'),
                payload.get('condicion_pago'),
                plantilla_id
            ))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Plantilla no encontrada")
            conn.commit()
            return {"success": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.delete("/plantillas/{plantilla_id}")
async def delete_plantilla(plantilla_id: str):
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE plantillas_cotizacion SET activo = false, updated_at = NOW() WHERE id = %s RETURNING id", (plantilla_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Plantilla no encontrada")
            conn.commit()
            return {"success": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

def _parse_excel_for_import(content: bytes) -> dict:
    """
    Parsea un archivo Excel (.xlsx) para extraer todos los datos de cotización.
    Retorna un dict con cliente, items, condiciones, plazo, condicion_pago, etc.
    """
    import openpyxl
    import re as _re
    
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    
    # Buscar la hoja de cotización
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() in ("MORT2", "COTIZACION", "COTIZACIÓN"):
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active or wb[wb.sheetnames[0]]
    
    def safe_cell(cell_ref: str) -> str:
        try:
            val = ws[cell_ref].value
            return str(val).strip() if val is not None else ""
        except Exception:
            return ""
    
    def safe_number(cell_ref: str) -> float:
        try:
            val = ws[cell_ref].value
            if val is None:
                return 0.0
            return float(val)
        except (ValueError, TypeError):
            return 0.0
    
    # --- Datos del cliente ---
    cliente = safe_cell("D5")
    ruc = safe_cell("D6")
    contacto = safe_cell("D7")
    telefono = safe_cell("E8")
    email = safe_cell("D9")
    proyecto = safe_cell("L5")
    ubicacion = safe_cell("L7")
    personal_comercial = safe_cell("L8")
    telefono_comercial = safe_cell("L9")
    titulo = safe_cell("G3")
    
    # --- Items (empiezan en fila 17) ---
    items = []
    row = 17
    max_empty_rows = 5
    empty_count = 0
    
    while empty_count < max_empty_rows and row < 200:
        desc = safe_cell(f"C{row}")
        codigo = safe_cell(f"B{row}")
        norma = safe_cell(f"J{row}")
        acreditado = safe_cell(f"K{row}")
        costo = safe_number(f"L{row}")
        cantidad = safe_number(f"M{row}")
        
        if not desc and costo == 0:
            empty_count += 1
            row += 1
            continue
        
        empty_count = 0
        items.append({
            "codigo": codigo,
            "descripcion": desc,
            "norma": norma,
            "acreditado": acreditado,
            "costo_unitario": costo,
            "cantidad": cantidad if cantidad > 0 else 1,
        })
        row += 1
    
    # --- Calcular extra_rows para localizar filas dinámicas ---
    extra_rows = max(0, len(items) - 1)
    
    # --- Condiciones Específicas (fila 23 + extra_rows, celda B) ---
    condiciones_especificas_texto = ""
    condiciones_especificas_lista = []
    row_condiciones = 23 + extra_rows
    condiciones_raw = safe_cell(f"B{row_condiciones}")
    if condiciones_raw and "CONDICIONES" in condiciones_raw.upper():
        condiciones_especificas_texto = condiciones_raw
        # Extraer items individuales (líneas que empiezan con -)
        for line in condiciones_raw.split("\n"):
            line = line.strip()
            if line.startswith("-"):
                condiciones_especificas_lista.append(line.lstrip("- ").strip())
    
    # --- Plazo Estimado (fila 24 + extra_rows, celda B) ---
    plazo_dias = 0
    plazo_texto = ""
    row_plazo = 24 + extra_rows
    plazo_raw = safe_cell(f"B{row_plazo}")
    if plazo_raw and "PLAZO" in plazo_raw.upper():
        plazo_texto = plazo_raw
        # Intentar extraer número de días
        match = _re.search(r'(\d+)\s*d[ií]as?\s*h[aá]biles', plazo_raw, _re.IGNORECASE)
        if match:
            plazo_dias = int(match.group(1))
    
    # --- Condición de Pago (fila 34 + extra_rows, celda B) ---
    condicion_pago_key = ""
    condicion_pago_texto = ""
    row_condicion = 34 + extra_rows
    condicion_raw = safe_cell(f"B{row_condicion}")
    if condicion_raw and "CONDICI" in condicion_raw.upper():
        condicion_pago_texto = condicion_raw
        # Detectar el tipo de condición de pago por texto
        texto_lower = condicion_raw.lower()
        if "valorización" in texto_lower or "valorizacion" in texto_lower:
            condicion_pago_key = "valorizacion"
        elif "50%" in condicion_raw:
            condicion_pago_key = "50_adelanto"
        elif "adelantado" in texto_lower:
            condicion_pago_key = "adelantado"
        elif "30 días" in texto_lower or "30 dias" in texto_lower:
            condicion_pago_key = "credito_30"
        elif "15 días" in texto_lower or "15 dias" in texto_lower:
            condicion_pago_key = "credito_15"
        elif "7 días" in texto_lower or "7 dias" in texto_lower:
            condicion_pago_key = "credito_7"
    
    sheet_names = list(wb.sheetnames) if hasattr(wb, 'sheetnames') else []
    wb.close()
    
    # --- Calcular totales ---
    subtotal = sum(it["costo_unitario"] * it["cantidad"] for it in items)
    igv = subtotal * 0.18
    total = subtotal + igv
    
    return {
        "cliente": cliente,
        "ruc": ruc,
        "contacto": contacto,
        "telefono": telefono,
        "email": email,
        "proyecto": proyecto,
        "ubicacion": ubicacion,
        "personal_comercial": personal_comercial,
        "telefono_comercial": telefono_comercial,
        "titulo_original": titulo,
        "items": items,
        "items_count": len(items),
        "subtotal": round(subtotal, 2),
        "igv": round(igv, 2),
        "total": round(total, 2),
        "plazo_dias": plazo_dias,
        "plazo_texto": plazo_texto,
        "condicion_pago_key": condicion_pago_key,
        "condicion_pago_texto": condicion_pago_texto,
        "condiciones_especificas_texto": condiciones_especificas_texto,
        "condiciones_especificas_lista": condiciones_especificas_lista,
        "hojas_disponibles": sheet_names,
    }


@router.post("/import-excel/check-number")
async def check_quote_number(numero: str = Query(...), year: Optional[int] = Query(None)):
    """Verifica si un número de cotización ya existe para el año dado."""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    check_year = year or date.today().year
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT id, numero, year, cliente_nombre, total, estado FROM cotizaciones WHERE numero = %s AND year = %s AND visibilidad = 'visible'",
                (numero, check_year)
            )
            existing = cur.fetchone()
            if existing:
                return {
                    "exists": True,
                    "quote": {
                        "id": existing["id"],
                        "numero": existing["numero"],
                        "year": existing["year"],
                        "cliente": existing["cliente_nombre"],
                        "total": float(existing["total"]) if existing["total"] else 0,
                        "estado": existing["estado"],
                    }
                }
            return {"exists": False}
    finally:
        conn.close()


@router.post("/import-excel")
async def import_excel_quote(
    file: UploadFile = File(...),
    user_id: Optional[str] = Query(None),
    user_name: Optional[str] = Query(None),
    custom_numero: Optional[str] = Query(None),
    condiciones_ids: Optional[str] = Query(None)
):
    """
    Importa un archivo Excel existente como nueva cotización.
    Parsea el contenido para detectar cliente, items y totales automáticamente,
    asigna un nuevo número de cotización y lo registra en el sistema.
    Si custom_numero se proporciona, usa ese número en lugar de generar uno nuevo.
    """
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    filename = file.filename or ""
    ext = Path(filename).suffix.lower()
    if ext not in ['.xlsx']:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx) para importación")
    
    content = await file.read()
    
    try:
        parsed = _parse_excel_for_import(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {str(e)}")
    
    items = parsed["items"]
    cliente = parsed["cliente"]
    
    if not items and not cliente:
        raise HTTPException(
            status_code=400, 
            detail="No se pudieron detectar datos en el Excel. Verifique que el formato sea compatible con el template de Geofal."
        )
    
    # --- Determinar número de cotización ---
    _ensure_sequence_table()
    year = date.today().year
    year_suffix = str(year)[-2:]
    
    if custom_numero and custom_numero.strip():
        cotizacion_numero = custom_numero.strip()
        # Verificar si ya existe para advertir (no bloquear, se hace UPSERT)
    else:
        sequential = _next_quote_sequential(year)
        cotizacion_numero = f"{sequential:03d}"
    
    # --- Preparar nombre de archivo para storage ---
    safe_cliente = _get_safe_filename(cliente or "IMPORTADO", "").rstrip('.')
    cloud_path = f"{year}/COT-{year}-{cotizacion_numero}_{safe_cliente}.xlsx"
    
    # --- Subir el Excel original a Supabase Storage ---
    file_bytes = io.BytesIO(content)
    uploaded_path = await asyncio.to_thread(
        _upload_to_supabase_storage, file_bytes, "cotizaciones", cloud_path
    )
    
    # --- Guardar copia local ---
    local_path = ""
    try:
        year_folder = QUOTES_FOLDER / str(year)
        year_folder.mkdir(exist_ok=True)
        local_filename = f"COT-{year}-{cotizacion_numero}_{safe_cliente}.xlsx"
        local_filepath = year_folder / local_filename
        with open(local_filepath, "wb") as f:
            f.write(content)
        local_path = str(local_filepath)
    except Exception as e:
        print(f"Warning: No se pudo guardar copia local: {e}")
    
    # --- Registrar en DB ---
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            items_json = json.dumps(items, ensure_ascii=False)
            
            # Parsear condiciones_ids (viene como string separado por comas)
            condiciones_uuid_list = []
            if condiciones_ids:
                condiciones_uuid_list = [c.strip() for c in condiciones_ids.split(",") if c.strip()]
            
            cur.execute("""
                INSERT INTO cotizaciones (
                    numero, year, cliente_nombre, cliente_ruc, cliente_contacto,
                    cliente_telefono, cliente_email, proyecto, ubicacion,
                    personal_comercial, telefono_comercial, fecha_emision,
                    subtotal, igv, total, include_igv, estado, moneda,
                    archivo_path, items_json, items_count, object_key,
                    vendedor_nombre, user_created, visibilidad,
                    plazo_dias, condicion_pago, condiciones_ids
                ) VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s::uuid[]
                )
                ON CONFLICT (year, numero) DO UPDATE SET
                    cliente_nombre = EXCLUDED.cliente_nombre,
                    cliente_ruc = EXCLUDED.cliente_ruc,
                    cliente_contacto = EXCLUDED.cliente_contacto,
                    cliente_telefono = EXCLUDED.cliente_telefono,
                    cliente_email = EXCLUDED.cliente_email,
                    proyecto = EXCLUDED.proyecto,
                    ubicacion = EXCLUDED.ubicacion,
                    archivo_path = EXCLUDED.archivo_path,
                    object_key = EXCLUDED.object_key,
                    items_json = EXCLUDED.items_json,
                    items_count = EXCLUDED.items_count,
                    total = EXCLUDED.total,
                    subtotal = EXCLUDED.subtotal,
                    igv = EXCLUDED.igv,
                    plazo_dias = EXCLUDED.plazo_dias,
                    condicion_pago = EXCLUDED.condicion_pago,
                    condiciones_ids = EXCLUDED.condiciones_ids,
                    visibilidad = 'visible',
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id
            """, (
                cotizacion_numero, year, cliente, parsed["ruc"], parsed["contacto"],
                parsed["telefono"], parsed["email"], parsed["proyecto"], parsed["ubicacion"],
                parsed["personal_comercial"], parsed["telefono_comercial"], date.today(),
                parsed["subtotal"], parsed["igv"], parsed["total"], True, 'borrador', 'PEN',
                local_path, items_json, len(items), cloud_path,
                user_name or parsed["personal_comercial"] or '', user_id, 'visible',
                parsed["plazo_dias"], parsed["condicion_pago_key"],
                condiciones_uuid_list if condiciones_uuid_list else None
            ))
            
            result = cur.fetchone()
            conn.commit()
            quote_id = result[0] if result else None
    except Exception as e:
        conn.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error registrando cotización: {str(e)}")
    finally:
        conn.close()
    
    return {
        "success": True,
        "quote_id": quote_id,
        "numero": cotizacion_numero,
        "year": year,
        "token": f"{cotizacion_numero}-{year_suffix}",
        "cloud_path": cloud_path,
        "parsed_data": parsed
    }


@router.post("/import-excel/preview")
async def preview_import_excel(file: UploadFile = File(...)):
    """
    Pre-visualiza los datos que se extraerían de un Excel sin crear la cotización.
    Permite al usuario verificar antes de confirmar la importación.
    """
    filename = file.filename or ""
    ext = Path(filename).suffix.lower()
    if ext not in ['.xlsx']:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx)")
    
    content = await file.read()
    
    try:
        parsed = _parse_excel_for_import(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {str(e)}")
    
    # Obtener el próximo número sugerido
    suggested_numero = ""
    try:
        if _has_database_url():
            _ensure_sequence_table()
            year = date.today().year
            # Peek at next number without consuming it
            conn = _get_connection()
            try:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT last_value FROM quote_sequences WHERE year = %s", (year,))
                    row = cur.fetchone()
                    next_val = (int(row["last_value"]) + 1) if row else 1
                    suggested_numero = f"{next_val:03d}"
            finally:
                conn.close()
    except Exception:
        suggested_numero = "001"
    
    parsed["suggested_numero"] = suggested_numero
    parsed["suggested_year"] = date.today().year
    
    # --- Cargar condiciones_especificas de la DB y hacer matching ---
    all_condiciones = []
    matched_condiciones_ids = []
    try:
        if _has_database_url():
            conn2 = _get_connection()
            try:
                with conn2.cursor(cursor_factory=RealDictCursor) as cur2:
                    cur2.execute(
                        "SELECT id, texto, categoria, orden FROM condiciones_especificas WHERE activo = true ORDER BY orden ASC"
                    )
                    all_condiciones = [
                        {"id": str(r["id"]), "texto": r["texto"], "categoria": r.get("categoria", ""), "orden": r.get("orden", 0)}
                        for r in cur2.fetchall()
                    ]
            finally:
                conn2.close()
            
            # Matching: comparar texto extraído del Excel con texto de DB
            if parsed.get("condiciones_especificas_lista") and all_condiciones:
                import re as _re2
                def _normalize(t: str) -> str:
                    """Normaliza texto para comparación flexible"""
                    t = t.lower().strip()
                    t = t.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
                    t = _re2.sub(r'[^\w\s]', '', t)
                    t = _re2.sub(r'\s+', ' ', t)
                    return t
                
                db_map = {_normalize(c["texto"]): c["id"] for c in all_condiciones}
                for extracted in parsed["condiciones_especificas_lista"]:
                    norm_extracted = _normalize(extracted)
                    # Coincidencia exacta normalizada
                    if norm_extracted in db_map:
                        matched_condiciones_ids.append(db_map[norm_extracted])
                        continue
                    # Coincidencia parcial: si el texto de DB contiene al extraído o viceversa
                    for db_norm, db_id in db_map.items():
                        if norm_extracted in db_norm or db_norm in norm_extracted:
                            if db_id not in matched_condiciones_ids:
                                matched_condiciones_ids.append(db_id)
                            break
    except Exception as e:
        print(f"Warning: No se pudieron cargar condiciones_especificas: {e}")
    
    parsed["all_condiciones"] = all_condiciones
    parsed["matched_condiciones_ids"] = matched_condiciones_ids
    
    return {
        "success": True,
        "preview": parsed
    }


@router.post("/{quote_id}/manual-upload")
async def manual_upload_quote(quote_id: str, file: UploadFile = File(...)):
    """Permite subir un archivo manual (PDF o Excel) para reemplazar el existente"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    # 1. Verificar que la cotización existe
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT numero, year, object_key, archivo_path FROM cotizaciones WHERE id = %s", (quote_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Quote not found")
        
        # 2. Leer contenido del archivo y determinar extensión
        content = await file.read()
        xlsx_bytes = io.BytesIO(content)
        
        filename = file.filename or ""
        ext = Path(filename).suffix.lower()
        if ext not in ['.xlsx', '.xls', '.pdf']:
             raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx) o PDF (.pdf)")

        # 2b. Validar que el nombre del archivo contiene el código de la cotización
        year = existing['year']
        numero = existing['numero']
        expected_code = f"COT-{year}-{numero}"
        if expected_code not in filename.upper():
             raise HTTPException(status_code=400, detail=f"El nombre del archivo debe contener el código {expected_code} para mantener el orden.")

        # 3. Determinar path de storage (usar el existente pero con la extensión correcta)
        year = existing['year']
        numero = existing['numero']
        
        safe_cliente_cloud = "MANUAL-UPLOAD" # Fallback if client name is complex
        # Construct new cloud path to ensure extension is correct
        cloud_path = f"{year}/COT-{year}-{numero}{ext}"
        
        # 3b. Eliminar archivo anterior de Storage si la extensión cambió (evitar archivos huérfanos)
        old_object_key = existing.get('object_key')
        if old_object_key and old_object_key != cloud_path:
            print(f"Manual upload: Eliminando archivo antiguo de Storage: {old_object_key}")
            await asyncio.to_thread(_delete_from_supabase_storage, "cotizaciones", old_object_key)
        
        # 4. Subir a Storage (el service ya usa x-upsert: true)
        xlsx_bytes.seek(0)
        from .service import _upload_to_supabase_storage # Ensure imported
        uploaded_path = await asyncio.to_thread(
            _upload_to_supabase_storage, xlsx_bytes, "cotizaciones", cloud_path
        )
        
        if not uploaded_path:
            raise HTTPException(status_code=500, detail="Error uploading to storage")
            
        # 5. Actualizar archivo local
        current_local_path = existing.get('archivo_path')
        new_local_path = None
        
        if current_local_path:
            # Si existe, derivamos el nuevo path local con la extensión correcta
            old_path = Path(current_local_path)
            new_path = old_path.with_suffix(ext)
            try:
                # Remove old file if extension changed
                if old_path.exists() and old_path != new_path:
                    try: old_path.unlink()
                    except: pass
                
                # Write new file
                new_path.parent.mkdir(parents=True, exist_ok=True)
                with open(new_path, "wb") as f:
                    f.write(content)
                new_local_path = str(new_path)
            except Exception as e:
                print(f"Error updating local file copy: {e}")
                new_local_path = current_local_path # Fallback

        # 6. Actualizar object_key y archivo_path en DB
        with conn.cursor() as cur:
            if new_local_path:
                cur.execute("UPDATE cotizaciones SET object_key = %s, archivo_path = %s, updated_at = NOW() WHERE id = %s", (cloud_path, new_local_path, quote_id))
            else:
                cur.execute("UPDATE cotizaciones SET object_key = %s, updated_at = NOW() WHERE id = %s", (cloud_path, quote_id))
            conn.commit()

        return {"success": True, "message": "Quote file replaced successfully", "object_key": cloud_path}
            
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals() and conn:
            conn.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()
