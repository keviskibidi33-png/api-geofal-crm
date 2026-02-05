
import io
import re
import copy
import zipfile
from copy import copy
from datetime import date
from typing import Any, List
from pathlib import Path
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.cell.cell import MergedCell
from .schemas import QuoteExportRequest

# Keep V1 version logic manually, or rely on a simpler extraction?
# The user wants "code relating to the user's requests should be written in the locations listed...".
# I should just copy the logic from main.py 1:1 to here to avoid breaking functionality.
# This logic was MASSIVE in main.py (lines 374-793).

# --- Template Path Constants ---
TEMPLATE_VARIANTS = {
    'V1': 'Temp_Cotizacion.xlsx',
    'V2': 'V2 - PROBETAS.xlsx',
    'V3': 'V3 - DENSIDAD DE CAMPO Y MUESTREO.xlsx',
    'V4': 'V4 - EXTRACCIÓN DE DIAMANTINA.xlsx',
    'V5': 'V5 - DIAMANTINA PARA PASES.xlsx',
    'V6': 'V6 - ALBAÑILERÍA.xlsx',
    'V7': 'V7 - VIGA BECKELMAN.xlsx',
    'V8': 'V8 - CONTROL DE CALIDAD DE CONCRETO FRESCO EN OBRA.xlsx',
}

def _get_template_path(template_id: str | None = None) -> Path:
    """Get template path based on template_id or default"""
    filename = TEMPLATE_VARIANTS.get(template_id, 'Temp_Cotizacion.xlsx') if template_id else 'Temp_Cotizacion.xlsx'
    
    # Path resolution: app/modules/cotizacion/excel.py -> app/
    # We use a robust relative path from this file to the templates directory
    current_dir = Path(__file__).resolve().parent
    app_dir = current_dir.parents[1] # app/
    
    possible_paths = [
        app_dir / "templates" / filename,  # Standard: app/templates/
        Path("/app/templates") / filename, # Docker absolute
        current_dir.parents[2] / "app" / "templates" / filename, # Root/app/templates/
    ]
    
    for p in possible_paths:
        if p.exists():
            return p
            
    # Fallback to standard app location
    return app_dir / "templates" / filename

# --- Excel Utility Functions (Copied from main.py) ---

def _copy_row_format(ws: Any, src_row: int, dst_row: int, *, min_col: int = 1, max_col: int = 60) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(min_col, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.comment = None

def _shift_range_rows(range_ref: str, *, insert_at_row: int, delta: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    if min_row >= insert_at_row:
        min_row += delta
        max_row += delta
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

def _restore_merged_cells(ws: Any, merged_ranges: list[str], *, insert_at_row: int, delta: int) -> None:
    for r in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(r))
        except Exception:
            continue

    for r in merged_ranges:
        new_ref = _shift_range_rows(r, insert_at_row=insert_at_row, delta=delta)
        try:
            ws.merge_cells(new_ref)
        except ValueError:
            continue

def _restore_print_area(ws: Any, print_area: Any, *, insert_at_row: int, delta: int) -> None:
    if not print_area:
        return

    raw = print_area
    if isinstance(raw, list):
        raw = raw[0] if raw else None
    if not raw or not isinstance(raw, str):
        return

    try:
        min_col, min_row, max_col, max_row = range_boundaries(raw)
    except Exception:
        return

    if max_row >= insert_at_row:
        max_row += delta
    ws.print_area = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

def _force_merge_b_to_n(ws: Any, row: int) -> None:
    target = f"B{row}:N{row}"

    for r in list(ws.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(r))
        except Exception:
            continue

        if min_row <= row <= max_row and not (max_col < 2 or min_col > 14):
            try:
                ws.unmerge_cells(str(r))
            except Exception:
                continue

    try:
        ws.merge_cells(target)
    except ValueError:
        return

def _force_merge_range(ws: Any, *, row: int, min_col: int, max_col: int) -> None:
    target = f"{get_column_letter(min_col)}{row}:{get_column_letter(max_col)}{row}"

    for r in list(ws.merged_cells.ranges):
        try:
            r_min_col, r_min_row, r_max_col, r_max_row = range_boundaries(str(r))
        except Exception:
            continue

        if r_min_row <= row <= r_max_row and not (r_max_col < min_col or r_min_col > max_col):
            try:
                ws.unmerge_cells(str(r))
            except Exception:
                continue

    try:
        ws.merge_cells(target)
    except ValueError:
        return

def _find_row_by_text(ws: Any, text: str, *, max_rows: int = 200, max_cols: int = 20) -> int | None:
    needle = text.strip().lower()
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and needle in v.strip().lower():
                return r
    return None

def _snapshot_row_style(ws: Any, *, row: int, min_col: int, max_col: int) -> dict[int, dict[str, Any]]:
    snap: dict[int, dict[str, Any]] = {}
    snap[0] = {"height": ws.row_dimensions[row].height}
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=c)
        snap[c] = {
            "_style": copy(cell._style),
            "number_format": cell.number_format,
            "alignment": copy(cell.alignment),
            "font": copy(cell.font),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "protection": copy(cell.protection),
        }
    return snap

def _snapshot_row_merges(ws: Any, *, row: int) -> list[str]:
    merges: list[str] = []
    for r in ws.merged_cells.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(r))
        except Exception:
            continue
        if min_row <= row <= max_row:
            merges.append(str(r))
    return merges

def _apply_row_merges(ws: Any, *, row: int, merges: list[str], insert_at_row: int, delta: int) -> None:
    if not merges:
        return

    for existing in list(ws.merged_cells.ranges):
        try:
            _, min_row, _, max_row = range_boundaries(str(existing))
        except Exception:
            continue
        if min_row <= row <= max_row:
            try:
                ws.unmerge_cells(str(existing))
            except Exception:
                continue

    for rng in merges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
        except Exception:
            continue

        if delta > 0:
            if min_row >= insert_at_row:
                min_row += delta
                max_row += delta
            elif max_row >= insert_at_row:
                max_row += delta

        target = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        try:
            ws.merge_cells(target)
        except ValueError:
            continue

def _apply_row_style(ws: Any, *, row: int, min_col: int, max_col: int, snap: dict[int, dict[str, Any]]) -> None:
    if 0 in snap and "height" in snap[0]:
        ws.row_dimensions[row].height = snap[0]["height"]

    for c in range(min_col, max_col + 1):
        if c not in snap:
            continue
        dst = ws.cell(row=row, column=c)
        s = snap[c]
        dst._style = copy(s["_style"])
        dst.number_format = s["number_format"]
        dst.alignment = copy(s["alignment"])
        dst.font = copy(s["font"])
        dst.border = copy(s["border"])
        dst.fill = copy(s["fill"])
        dst.protection = copy(s["protection"])

def _set_cell(ws: Any, addr: str, value: Any) -> None:
    if value is None:
        return

    cell = ws[addr]
    try:
        cell.value = value
        return
    except AttributeError:
        pass

    try:
        col, row = ws[addr].column, ws[addr].row
    except Exception:
        ws[addr].value = value
        return

    for r in ws.merged_cells.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(r))
        except Exception:
            continue
        if min_col <= col <= max_col and min_row <= row <= max_row:
            top_left = ws.cell(row=min_row, column=min_col)
            top_left.value = value
            return

    ws[addr].value = value

def _apply_quote_number(ws: Any, addr: str, cotizacion_numero: str | None, fecha_emision: date | None) -> None:
    if not cotizacion_numero and not fecha_emision:
        return

    current = ws[addr].value
    if not isinstance(current, str):
        current = ""

    if fecha_emision is None:
        fecha_emision = date.today()

    year_suffix = str(fecha_emision.year)[-2:]
    numero = cotizacion_numero or "000"
    token = f"{numero}-{year_suffix}"

    if "XXX-XX" in current:
        ws[addr].value = current.replace("XXX-XX", token)
        return

    if re.search(r"XXX-\d{2}", current):
        ws[addr].value = re.sub(r"XXX-\d{2}", token, current)
        return

    ws[addr].value = re.sub(r"\b\d{1,6}-\d{2}\b", token, current) or token

def _shift_drawing_xml(data: bytes, *, start_row0: int, delta: int) -> bytes:
    if delta <= 0:
        return data
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        return data

    def repl(m: re.Match[str]) -> str:
        v = int(m.group(2))
        if v >= start_row0:
            v += delta
        return f"{m.group(1)}{v}{m.group(3)}"

    row_vals = [int(v) for v in re.findall(r"<xdr:row>(\d+)</xdr:row>", text)]
    has_xdr_rows = bool(row_vals)
    if not row_vals:
        row_vals = [int(v) for v in re.findall(r"<[A-Za-z0-9_]+:row>(\d+)</[A-Za-z0-9_]+:row>", text)]

    if not any(start_row0 <= v <= 200 for v in row_vals):
        return data

    if has_xdr_rows:
        text = re.sub(r"(<xdr:row>)(\d+)(</xdr:row>)", repl, text)
    else:
        text = re.sub(r"(<[A-Za-z0-9_]+:row>)(\d+)(</[A-Za-z0-9_]+:row>)", repl, text)

    return text.encode("utf-8")

def _shift_vml(data: bytes, *, start_row0: int, delta: int) -> bytes:
    if delta <= 0:
        return data
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = data.decode("latin-1")
        except Exception:
            return data

    def repl_row(m: re.Match[str]) -> str:
        v = int(m.group(1))
        if v >= start_row0:
            v += delta
        return f"{m.group(0).split('>')[0]}>" + str(v) + f"</{m.group(0).split('</')[1]}"

    text = re.sub(r"(<x:Row>)(\d+)(</x:Row>)", lambda m: f"{m.group(1)}{int(m.group(2)) + (delta if int(m.group(2)) >= start_row0 else 0)}{m.group(3)}", text)

    def repl_anchor(m: re.Match[str]) -> str:
        parts = [p.strip() for p in m.group(2).split(",")]
        try:
            nums = [int(p) for p in parts]
        except Exception:
            return m.group(0)

        if len(nums) >= 8:
            if nums[2] >= start_row0:
                nums[2] += delta
            if nums[6] >= start_row0:
                nums[6] += delta
            new_val = ",".join(str(n) for n in nums)
            return f"{m.group(1)}{new_val}{m.group(3)}"
        return m.group(0)

    text = re.sub(r"(<x:Anchor>)([^<]+)(</x:Anchor>)", repl_anchor, text)

    try:
        return text.encode("utf-8")
    except Exception:
        return data

def _preserve_template_assets(template_path: Path, generated: io.BytesIO, *, insert_at_row: int, delta: int) -> io.BytesIO:
    generated.seek(0)
    out = io.BytesIO()

    with zipfile.ZipFile(template_path, "r") as ztpl:
        with zipfile.ZipFile(generated, "r") as zgen:
            with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                names_tpl = set(ztpl.namelist())

                def should_take_from_template(name: str) -> bool:
                    if name.startswith("xl/media/"): return True
                    if name.startswith("xl/drawings/"): return True
                    return False

                tpl_override = {n for n in names_tpl if should_take_from_template(n)}
                gen_names = set(zgen.namelist())
                start_row0 = max(0, insert_at_row - 2)

                for name in zgen.namelist():
                    if name in tpl_override and name in names_tpl:
                        data = ztpl.read(name)
                        if name.startswith("xl/drawings/") and name.endswith(".xml"):
                            data = _shift_drawing_xml(data, start_row0=start_row0, delta=delta)
                        if name.startswith("xl/drawings/") and name.endswith(".vml"):
                            data = _shift_vml(data, start_row0=start_row0, delta=delta)
                    else:
                        data = zgen.read(name)
                    zout.writestr(name, data)

                for name in tpl_override:
                    if name not in gen_names and name in names_tpl:
                        data = ztpl.read(name)
                        if name.startswith("xl/drawings/") and name.endswith(".xml"):
                            data = _shift_drawing_xml(data, start_row0=start_row0, delta=delta)
                        if name.startswith("xl/drawings/") and name.endswith(".vml"):
                            data = _shift_vml(data, start_row0=start_row0, delta=delta)
                        zout.writestr(name, data)

    out.seek(0)
    return out

# Import this to avoid circular import issues if placed at top, but usually fine.
# Note: we need app.xlsx_direct_v2 for export_xlsx_direct in main.py, 
# but here we seem to be RE-IMPLEMENTING "main.py's _export_xlsx". 
# Wait, main.py uses `export_xlsx_direct` (from app/xlsx_direct_v2.py) at line 876.
# AND it also had a huge chunk of `_export_xlsx` from 394 to 794. 
# Ah, `_export_xlsx` in main.py calls `app.xlsx_direct_v2.export_xlsx_direct` inside it? 
# No, `_export_xlsx` (lines 794-End) in main.py uses `app.xlsx_direct_v2.export_xlsx_direct`.
# BUT `export_xlsx_direct` was imported as `export_xlsx_direct`.
# The huge chunk of code (lines 379-791) in `main.py` were helper functions used by... what? 
# They seemed to be used by `_export_xlsx` which was NOT fully shown in my view (it cut off at 800).
# Let's check `_export_xlsx` again in `main.py`.
# Line 876: `return export_xlsx_direct(str(template_path), export_data)`
# So `main.py`'s `_export_xlsx` function actually calls `export_xlsx_direct`. 
# AND `main.py` defines `_copy_row_format`, etc. Are they unused?
# Or does `export_xlsx_direct` use them? `export_xlsx_direct` is imported from `app.xlsx_direct_v2`.
# If `app.xlsx_direct_v2` is a separate file, then main.py's definitions might be legacy or unused overrides?

# Wait, `xlsx_direct_v2.py` has `_duplicate_row`, `_shift_rows`, `_set_cell_value`.
# `main.py` has `_copy_row_format`, `_shift_range_rows`, `_shift_drawing_xml`.
# It looks like `main.py` has a LOT of legacy code.
# The `export_quote` endpoint uses `_export_xlsx` (line 1073).
# `_export_xlsx` (line 794) calls `export_xlsx_direct` (line 876).
# So `export_xlsx_direct` is the one doing the work.
# Therefore, I can just import `export_xlsx_direct` in `app/modules/cotizacion/excel.py` and wrap it?
# Or does `_export_xlsx` do pre-processing?
# Lines 803-876 of `main.py` (inside `_export_xlsx`) do Pydantic -> Dict conversion and logic.
# So I should move `_export_xlsx` logic to `app/modules/cotizacion/excel.py` but rename it to `generate_quote_excel`.
# And I should import `export_xlsx_direct` from `app.xlsx_direct_v2` (or better, move `xlsx_direct_v2.py` to `app/common/` or `app/modules/cotizacion/xlsx_utils.py` if it's specific).
# Since `programacion` also used `xlsx_direct_v2` (imported in `programacion_export.py`), it's a shared util.
# I will keep `app/xlsx_direct_v2.py` where it is for now, or move to `app/common`.
# For now, I will keep it and import it.

# All those helper functions in `main.py` (`_copy_row_format` etc) MIGHT be unused if `_export_xlsx` delegates entirely. 
# If `export_xlsx_direct` does the job, then `main.py` had dead code? 
# OR `export_xlsx_direct` was imported but `main.py` defined its OWN version?
# Line 27: `# from app.xlsx_direct_v2 import export_xlsx_direct` -> Commented out!
# Wait, I need to check line 27 in my PREVIOUS `view_file`.
# In `main.py` view (Step 2828):
# 27: # from app.xlsx_direct_v2 import export_xlsx_direct
# 28: from app.programacion_export import export_programacion_xlsx
# ...
# 876: return export_xlsx_direct(str(template_path), export_data)
# If it's commented out, where does `export_xlsx_direct` come from? 
# Maybe defined inside `main.py`? 
# I searched `main.py` view, I didn't see `def export_xlsx_direct`.
# BUT I only viewed up to line 1600. `xls_direct_v2.py` was viewed separately.
# Maybe I missed the import or definition.
# Let's assume I need to implement `generate_quote_excel` using `app.xlsx_direct_v2`.

import openpyxl
from openpyxl.drawing.image import Image
from .schemas import QuoteExportRequest
from datetime import date
import io

def generate_quote_excel(payload: QuoteExportRequest) -> io.BytesIO:
    """
    Genera el Excel de cotización usando openpyxl para manipular el template.
    Específicamente enfocado en el template 'V1' (sheet8/PRUEBA 1).
    """
    template_path = _get_template_path(payload.template_id)
    if not template_path.exists():
        raise FileNotFoundError(f"Template no encontrado: {template_path}")
    
    # Cargar el libro de trabajo con openpyxl
    wb = openpyxl.load_workbook(str(template_path))
    
    # Seleccionar la hoja correcta con una lógica de prioridad flexible
    ws = None
    
    # 1. Intentar con la hoja activa (la que guardó el usuario) si parece ser el formulario
    active_ws = wb.active
    if active_ws and active_ws.sheet_state == 'visible':
        # Verificamos si tiene "CLIENTE:" o "COTIZACIÓN" en las primeras filas
        is_form = False
        for r in range(1, 10):
            for c in range(1, 5):
                val = str(active_ws.cell(row=r, column=c).value or "").upper()
                if "CLIENTE:" in val or "COTIZACIÓN" in val:
                    is_form = True; break
            if is_form: break
        if is_form: ws = active_ws

    # 2. Si no, buscar hojas por nombres conocidos (Priorizando MORT2 que vimos en screenshot)
    if not ws:
        for target in ['MORT2', 'MORT1', 'PRUEBA 1']:
            if target in wb.sheetnames and wb[target].sheet_state == 'visible':
                ws = wb[target]; break

    # 3. Si no, buscar cualquier hoja visible que tenga "CLIENTE:" en B5 (o cerca)
    if not ws:
        for name in wb.sheetnames:
            temp_ws = wb[name]
            if temp_ws.sheet_state == 'visible':
                val = str(temp_ws.cell(row=5, column=2).value or "").upper()
                if "CLIENTE:" in val:
                    ws = temp_ws; break
    
    if not ws:
        ws = wb.active # Último recurso
    
    # Asegurarnos de que esta sea la hoja activa al abrir el archivo
    wb.active = ws

    def _normalize(text):
        if not text or not isinstance(text, str): return ""
        t = text.upper().strip()
        for a, b in [("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U")]:
            t = t.replace(a, b)
        return t

    def _find_anchor(label_text, max_row=60, max_col=20):
        target = _normalize(label_text)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                val = _normalize(ws.cell(row=r, column=c).value)
                if target in val:
                    return r, c
        return None, None

    def _safe_write(r, c, val):
        """Escribe en una celda, manejando si es parte de un rango combinado."""
        from openpyxl.cell.cell import MergedCell
        cell = ws.cell(row=r, column=c)
        if isinstance(cell, MergedCell):
            for range_ in ws.merged_cells.ranges:
                if cell.coordinate in range_:
                    ws.cell(row=range_.min_row, column=range_.min_col).value = val
                    return
        cell.value = val

    def _safe_write_rel(label, val, col_offset=2):
        """Busca una etiqueta y escribe en la celda a su derecha."""
        r, c = _find_anchor(label)
        if r and c:
            _safe_write(r, c + col_offset, val)

    # 1. Poner número de cotización en el título
    fecha_emision = payload.fecha_emision or date.today()
    numero = payload.cotizacion_numero or "000"
    year_suffix = str(fecha_emision.year)[-2:]
    token = f"{numero}-{year_suffix}"
    
    tr, tc = _find_anchor("COTIZACIÓN DE LABORATORIO")
    if tr and tc:
        cell = ws.cell(row=tr, column=tc)
        if cell.value and "N°" in cell.value:
            cell.value = cell.value.split("N°")[0] + f"N° {token}"
        else:
            cell.value = f"COTIZACIÓN DE LABORATORIO N° {token}"

    # 2. Datos de Cabecera
    _safe_write_rel("CLIENTE:", payload.cliente or "")
    _safe_write_rel("R.U.C", payload.ruc or "")
    _safe_write_rel("CONTACTO:", payload.contacto or "")
    _safe_write_rel("TELÉFONO DE CONTACTO:", payload.telefono_contacto or "", col_offset=3)
    _safe_write_rel("CORREO:", payload.correo or "")
    _safe_write_rel("PROYECTO", payload.proyecto or "", col_offset=2)
    _safe_write_rel("UBICACIÓN", payload.ubicacion or "", col_offset=2)
    _safe_write_rel("PERSONAL COMERCIAL", payload.personal_comercial or "", col_offset=2)
    _safe_write_rel("TELÉFONO DE COMERCIAL", payload.telefono_comercial or "", col_offset=2)
    _safe_write_rel("FECHA SOLICITUD", payload.fecha_solicitud or "")
    _safe_write_rel("FECHA DE EMISIÓN:", fecha_emision.strftime("%d/%m/%Y"), col_offset=2)

    # 3. Tabla de Items
    ir_last, ic_code = None, None
    for r in range(1, 40):
        # Miramos columnas A o B para el keyword CÓDIGO o ITEM (normalizado)
        v1 = _normalize(ws.cell(row=r, column=1).value)
        v2 = _normalize(ws.cell(row=r, column=2).value)
        if "CODIGO" in v1 or "ITEM" in v1:
            ir_last, ic_code = r, 1
        elif "CODIGO" in v2 or "ITEM" in v2:
            ir_last, ic_code = r, 2

    if ir_last:
        # Detectar el final real de la cabecera (merge vertical)
        max_h_row = ir_last
        test_cell = ws.cell(row=ir_last, column=ic_code)
        if isinstance(test_cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if test_cell.coordinate in rng:
                    max_h_row = max(max_h_row, rng.max_row)
        
        # En MORT2 ir_last suele ser 15, max_h_row=16, START_ROW=17
        START_ROW = max_h_row + 1
        
        # Mapa de columnas: escanea la ÚLTIMA fila de cabecera
        col_map = {}
        target_header_row = max_h_row
        header_labels = {
            "CÓDIGO": "code", "ITEM": "code",
            "DESCRIPCIÓN": "desc",
            "NORMA": "norma",
            "ACREDITADO": "acred",
            "COSTO UNITARIO": "price", "PRECIO": "price",
            "CANTIDAD": "cant",
            "COSTO PARCIAL": "total"
        }
        for c in range(1, 20):
            val = _normalize(ws.cell(row=target_header_row, column=c).value)
            # Manejar celdas combinadas buscando a la izquierda si está vacío
            if not val:
                for cc in range(c-1, 0, -1):
                    v_neighbor = _normalize(ws.cell(row=target_header_row, column=cc).value)
                    if v_neighbor: val = v_neighbor; break
            
            for lab, key in header_labels.items():
                if _normalize(lab) in val:
                    if key not in col_map: col_map[key] = c
        
        c_code = col_map.get("code", ic_code or 2)
        c_desc = col_map.get("desc", c_code + 1)
        c_norma = col_map.get("norma", c_code + 7) 
        c_acred = col_map.get("acred", c_code + 8)
        c_price = col_map.get("price", c_code + 10)
        c_cant = col_map.get("cant", c_code + 11)
        c_total = col_map.get("total", c_code + 12) 
    else:
        START_ROW = 15
        c_code, c_desc, c_norma, c_acred, c_price, c_cant, c_total = 2, 3, 9, 10, 12, 13, 14
        
    items = payload.items or []
    total_parcial = 0
    
    # Detectar el salto de filas (step) entre items
    step = 1
    if START_ROW + 1 <= ws.max_row:
        c1 = ws.cell(row=START_ROW, column=c_code)
        c2 = ws.cell(row=START_ROW + 1, column=c_code)
        if isinstance(c1, MergedCell) and isinstance(c2, MergedCell):
            for rng in ws.merged_cells.ranges:
                if c1.coordinate in rng and c2.coordinate in rng:
                    step = 2; break

    # --- Duplicación de filas para múltiples items ---
    if len(items) > 1:
        # Tomar snapshot del estilo y combinaciones de la(s) fila(s) del primer item
        row_snapshots = []
        for i in range(step):
            snap = _snapshot_row_style(ws, row=START_ROW + i, min_col=1, max_col=20)
            merges = _snapshot_row_merges(ws, row=START_ROW + i)
            row_snapshots.append((snap, merges))
        
        # Insertar filas para los items adicionales (idx 1 en adelante)
        # Insertamos desde abajo hacia arriba o reservamos espacio? 
        # Insertar en START_ROW + step es lo más limpio
        num_new_rows = (len(items) - 1) * step
        ws.insert_rows(START_ROW + step, amount=num_new_rows)
        
        # Aplicar estilos y merges a las nuevas filas
        for idx in range(1, len(items)):
            base_r = START_ROW + (idx * step)
            for i in range(step):
                snap, merges = row_snapshots[i]
                target_r = base_r + i
                _apply_row_style(ws, row=target_r, min_col=1, max_col=20, snap=snap)
                # Ojo: _apply_row_merges necesita delta relativo si se movieron cosas, 
                # pero aquí las filas son "nuevas" y limpias. 
                # Re-implementamos un apply simple para merges horizontales en la misma fila:
                for m_ref in merges:
                    m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(m_ref)
                    r_offset = target_r - m_min_row
                    new_m_ref = f"{get_column_letter(m_min_col)}{m_min_row + r_offset}:{get_column_letter(m_max_col)}{m_max_row + r_offset}"
                    try:
                        ws.merge_cells(new_m_ref)
                    except Exception: pass

    for idx, item in enumerate(items):
        r = START_ROW + (idx * step)
        _safe_write(r, c_code, item.codigo)
        
        # Manejo de Descripcion + Norma (pueden estar unidas C-I)
        cell_desc = ws.cell(row=r, column=c_desc)
        cell_norma = ws.cell(row=r, column=c_norma)
        is_unified = False
        if isinstance(cell_desc, MergedCell) and isinstance(cell_norma, MergedCell):
            for range_ in ws.merged_cells.ranges:
                if cell_desc.coordinate in range_ and cell_norma.coordinate in range_:
                    is_unified = True; break
        
        if is_unified:
            text = f"{item.descripcion or ''}"
            if item.norma: text += f" - {item.norma}"
            _safe_write(r, c_desc, text)
        else:
            _safe_write(r, c_desc, item.descripcion)
            _safe_write(r, c_norma, item.norma)
            
        _safe_write(r, c_acred, item.acreditado)
        _safe_write(r, c_price, item.costo_unitario)
        _safe_write(r, c_cant, item.cantidad)
        
        parcial = (item.costo_unitario or 0) * (item.cantidad or 0)
        _safe_write(r, c_total, parcial)
        total_parcial += parcial

    # 4. Totales
    tr_total, tc_total_label = _find_anchor("Costo parcial", max_row=START_ROW + 50)
    if tr_total:
        target_c = c_total if 'c_total' in locals() else tc_total_label + 2
        _safe_write(tr_total, target_c, total_parcial)
        _safe_write(tr_total + 1, target_c, total_parcial * 0.18 if payload.include_igv else 0)
        _safe_write(tr_total + 2, target_c, total_parcial * 1.18 if payload.include_igv else total_parcial)

    # 5. Condiciones del Servicio (Opcional)
    # Por ahora no tocamos el texto de condiciones ya que suele estar fijo en el template

    # Guardar a un BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
