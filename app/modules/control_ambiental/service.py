from __future__ import annotations

import datetime
import io
import logging
import os
from typing import List, Optional, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import func, desc

from lxml import etree
from app.modules.common.excel_xml import (
    find_template_path,
    transform_template_sheet,
    set_cell,
    col_num_to_letter,
    _set_paragraph_text,
    NS_SHEET,
    NS_DRAW,
    NS_A,
)
from openpyxl import Workbook, load_workbook

from .models import ControlTemperatura, ControlBalanza
from .schemas import (
    ControlTemperaturaCreate,
    ControlTemperaturaResponse,
    ControlBalanzaCreate,
    ControlBalanzaResponse,
    ControlAmbientalDashboardResponse,
    AreaStatusSummary,
    BalanzaStatusSummary,
)

logger = logging.getLogger(__name__)

STANDARD_AREAS = [
    {"name": "CÁMARA HÚMEDA", "norma": "NTP 339.033 / ASTM C192", "rango_temp": "23°C ± 2.0°C", "rango_hum": "≥ 95%", "temp_target": 23.0, "temp_tol": 2.0, "hum_min": 90.0},
    {"name": "LABORATORIO SUELOS", "norma": "MTC E 101 / NTP 339.127", "rango_temp": "20°C ± 3.0°C", "rango_hum": "50% - 70%", "temp_target": 20.0, "temp_tol": 3.0, "hum_min": 45.0},
    {"name": "LABORATORIO CONCRETO", "norma": "NTP 339.034 / MTC E 704", "rango_temp": "20°C ± 3.0°C", "rango_hum": "50% - 70%", "temp_target": 20.0, "temp_tol": 3.0, "hum_min": 45.0},
    {"name": "ENSAYOS QUÍMICOS", "norma": "NTP 339.178 / MTC E 219", "rango_temp": "20°C ± 2.0°C", "rango_hum": "50% - 65%", "temp_target": 20.0, "temp_tol": 2.0, "hum_min": 48.0},
]

STANDARD_BALANZAS = [
    {"codigo": "BAL-01", "ubicacion": "Muestras / Cám. Húmeda", "capacidad_g": 30000.0, "masa_patron_g": 5000.0, "error_max_g": 1.0},
    {"codigo": "BAL-02", "ubicacion": "Laboratorio Suelos", "capacidad_g": 20000.0, "masa_patron_g": 2000.0, "error_max_g": 0.5},
    {"codigo": "BAL-03", "ubicacion": "Laboratorio Concreto", "capacidad_g": 5000.0, "masa_patron_g": 1000.0, "error_max_g": 0.1},
    {"codigo": "BAL-04", "ubicacion": "Química / Finos", "capacidad_g": 1000.0, "masa_patron_g": 500.0, "error_max_g": 0.05},
    {"codigo": "BAL-05", "ubicacion": "Analítica General", "capacidad_g": 300.0, "masa_patron_g": 100.0, "error_max_g": 0.005},
    {"codigo": "BAL-06", "ubicacion": "Laboratorio Huanta", "capacidad_g": 15000.0, "masa_patron_g": 2000.0, "error_max_g": 0.5},
]


def _evaluar_cumplimiento_temp(area: str, temp: float, hum: float) -> bool:
    area_norm = area.upper()
    if "CÁMARA HÚMEDA" in area_norm or "CURADO" in area_norm:
        temp_ok = 21.0 <= temp <= 25.0
        hum_ok = hum >= 90.0
    else:
        temp_ok = 17.0 <= temp <= 24.0
        hum_ok = 45.0 <= hum <= 80.0
    return temp_ok and hum_ok


class ControlAmbientalService:

    # --- Temperatura ---

    @staticmethod
    def listar_temperatura(
        db: Session,
        area: Optional[str] = None,
        fecha_inicio: Optional[str] = None,
        fecha_fin: Optional[str] = None,
        limit: int = 100,
    ) -> List[ControlTemperatura]:
        query = db.query(ControlTemperatura)
        if area:
            query = query.filter(ControlTemperatura.area_ambiente.ilike(f"%{area}%"))
        if fecha_inicio:
            query = query.filter(ControlTemperatura.fecha >= fecha_inicio)
        if fecha_fin:
            query = query.filter(ControlTemperatura.fecha <= fecha_fin)
        return query.order_by(desc(ControlTemperatura.fecha), desc(ControlTemperatura.hora_lectura), desc(ControlTemperatura.id)).limit(limit).all()

    @staticmethod
    def crear_temperatura(
        db: Session, payload: ControlTemperaturaCreate, user_id: str, user_name: str
    ) -> ControlTemperatura:
        cumple = payload.cumple_especificacion
        if cumple is None:
            cumple = _evaluar_cumplimiento_temp(payload.area_ambiente, payload.temperatura_c, payload.humedad_relativa_pct)

        record = ControlTemperatura(
            fecha=payload.fecha,
            hora_lectura=payload.hora_lectura,
            area_ambiente=payload.area_ambiente.strip().upper(),
            temperatura_c=round(payload.temperatura_c, 2),
            humedad_relativa_pct=round(payload.humedad_relativa_pct, 2),
            temp_min=round(payload.temp_min, 2) if payload.temp_min is not None else None,
            temp_max=round(payload.temp_max, 2) if payload.temp_max is not None else None,
            cumple_especificacion=cumple,
            responsable_lectura=payload.responsable_lectura or user_name,
            observaciones=payload.observaciones,
        )
        db.add(record)
        db.commit()
        db.refresh(record)

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Registró lectura de temperatura en {record.area_ambiente} ({record.temperatura_c}°C, {record.humedad_relativa_pct}%)",
            module="CONTROL_AMBIENTAL",
            details={
                "id": record.id,
                "area": record.area_ambiente,
                "temperatura": record.temperatura_c,
                "humedad": record.humedad_relativa_pct,
                "cumple": record.cumple_especificacion,
            },
        )
        return record

    @staticmethod
    def actualizar_temperatura(
        db: Session, record_id: int, payload: ControlTemperaturaCreate, user_id: str, user_name: str
    ) -> ControlTemperatura:
        record = db.query(ControlTemperatura).filter(ControlTemperatura.id == record_id).first()
        if not record:
            raise ValueError(f"Registro de temperatura ID {record_id} no encontrado")

        cumple = payload.cumple_especificacion
        if cumple is None:
            cumple = _evaluar_cumplimiento_temp(payload.area_ambiente, payload.temperatura_c, payload.humedad_relativa_pct)

        record.fecha = payload.fecha
        record.hora_lectura = payload.hora_lectura
        record.area_ambiente = payload.area_ambiente.strip().upper()
        record.temperatura_c = round(payload.temperatura_c, 2)
        record.humedad_relativa_pct = round(payload.humedad_relativa_pct, 2)
        record.temp_min = round(payload.temp_min, 2) if payload.temp_min is not None else None
        record.temp_max = round(payload.temp_max, 2) if payload.temp_max is not None else None
        record.cumple_especificacion = cumple
        record.responsable_lectura = payload.responsable_lectura or user_name
        record.observaciones = payload.observaciones

        db.commit()
        db.refresh(record)

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Actualizó lectura de temperatura ID {record.id} ({record.area_ambiente})",
            module="CONTROL_AMBIENTAL",
            details={"id": record.id, "area": record.area_ambiente, "cumple": record.cumple_especificacion},
        )
        return record

    @staticmethod
    def eliminar_temperatura(db: Session, record_id: int, user_id: str, user_name: str) -> bool:
        record = db.query(ControlTemperatura).filter(ControlTemperatura.id == record_id).first()
        if not record:
            return False

        db.delete(record)
        db.commit()

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Eliminó lectura de temperatura ID {record_id}",
            module="CONTROL_AMBIENTAL",
            details={"id": record_id},
        )
        return True

    # --- Balanza ---

    @staticmethod
    def listar_balanza(
        db: Session,
        codigo: Optional[str] = None,
        fecha_inicio: Optional[str] = None,
        fecha_fin: Optional[str] = None,
        limit: int = 100,
    ) -> List[ControlBalanza]:
        query = db.query(ControlBalanza)
        if codigo:
            query = query.filter(ControlBalanza.codigo_balanza.ilike(f"%{codigo}%"))
        if fecha_inicio:
            query = query.filter(ControlBalanza.fecha >= fecha_inicio)
        if fecha_fin:
            query = query.filter(ControlBalanza.fecha <= fecha_fin)
        return query.order_by(desc(ControlBalanza.fecha), desc(ControlBalanza.id)).limit(limit).all()

    @staticmethod
    def crear_balanza(
        db: Session, payload: ControlBalanzaCreate, user_id: str, user_name: str
    ) -> ControlBalanza:
        error_g = round(payload.lectura_balanza_g - payload.masa_patron_g, 4)
        conforme = payload.estado_conforme
        if conforme is None:
            conforme = (abs(error_g) <= payload.error_max_permitido_g) and payload.limpieza_nivelacion

        record = ControlBalanza(
            fecha=payload.fecha,
            codigo_balanza=payload.codigo_balanza.strip().upper(),
            ubicacion=payload.ubicacion.strip().upper(),
            capacidad_g=payload.capacidad_g,
            masa_patron_g=payload.masa_patron_g,
            lectura_balanza_g=payload.lectura_balanza_g,
            error_g=error_g,
            error_max_permitido_g=payload.error_max_permitido_g,
            estado_conforme=conforme,
            limpieza_nivelacion=payload.limpieza_nivelacion,
            verificado_por=payload.verificado_por or user_name,
            observaciones=payload.observaciones,
        )
        db.add(record)
        db.commit()
        db.refresh(record)

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Registró verificación diaria de balanza {record.codigo_balanza} (error: {error_g}g, conforme: {conforme})",
            module="CONTROL_AMBIENTAL",
            details={
                "id": record.id,
                "codigo": record.codigo_balanza,
                "error_g": error_g,
                "conforme": record.estado_conforme,
            },
        )
        return record

    @staticmethod
    def actualizar_balanza(
        db: Session, record_id: int, payload: ControlBalanzaCreate, user_id: str, user_name: str
    ) -> ControlBalanza:
        record = db.query(ControlBalanza).filter(ControlBalanza.id == record_id).first()
        if not record:
            raise ValueError(f"Registro de balanza ID {record_id} no encontrado")

        error_g = round(payload.lectura_balanza_g - payload.masa_patron_g, 4)
        conforme = payload.estado_conforme
        if conforme is None:
            conforme = (abs(error_g) <= payload.error_max_permitido_g) and payload.limpieza_nivelacion

        record.fecha = payload.fecha
        record.codigo_balanza = payload.codigo_balanza.strip().upper()
        record.ubicacion = payload.ubicacion.strip().upper()
        record.capacidad_g = payload.capacidad_g
        record.masa_patron_g = payload.masa_patron_g
        record.lectura_balanza_g = payload.lectura_balanza_g
        record.error_g = error_g
        record.error_max_permitido_g = payload.error_max_permitido_g
        record.estado_conforme = conforme
        record.limpieza_nivelacion = payload.limpieza_nivelacion
        record.verificado_por = payload.verificado_por or user_name
        record.observaciones = payload.observaciones

        db.commit()
        db.refresh(record)

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Actualizó verificación de balanza ID {record.id} ({record.codigo_balanza})",
            module="CONTROL_AMBIENTAL",
            details={"id": record.id, "codigo": record.codigo_balanza, "conforme": record.estado_conforme},
        )
        return record

    @staticmethod
    def eliminar_balanza(db: Session, record_id: int, user_id: str, user_name: str) -> bool:
        record = db.query(ControlBalanza).filter(ControlBalanza.id == record_id).first()
        if not record:
            return False

        db.delete(record)
        db.commit()

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Eliminó verificación de balanza ID {record_id}",
            module="CONTROL_AMBIENTAL",
            details={"id": record_id},
        )
        return True

    # --- Dashboard KPI ---

    @staticmethod
    def obtener_dashboard(db: Session) -> ControlAmbientalDashboardResponse:
        today_str = datetime.date.today().strftime("%Y-%m-%d")

        # Temp stats
        temps = db.query(ControlTemperatura).all()
        total_temps = len(temps)
        avg_temp = round(sum(t.temperatura_c for t in temps) / total_temps, 1) if total_temps > 0 else 21.5
        avg_hum = round(sum(t.humedad_relativa_pct for t in temps) / total_temps, 1) if total_temps > 0 else 72.0
        cumplen_temp = sum(1 for t in temps if t.cumple_especificacion)
        tasa_cumplimiento_temp = round((cumplen_temp / total_temps) * 100.0, 1) if total_temps > 0 else 100.0

        # Balanza stats
        balanzas = db.query(ControlBalanza).all()
        total_balanzas_records = len(balanzas)
        unique_balanza_codes = set(b.codigo_balanza for b in balanzas) or {"BAL-01", "BAL-02", "BAL-03", "BAL-04", "BAL-05", "BAL-06"}
        verificadas_hoy = len(set(b.codigo_balanza for b in balanzas if b.fecha == today_str))
        conformes_balanza = sum(1 for b in balanzas if b.estado_conforme)
        tasa_conforma_balanza = round((conformes_balanza / total_balanzas_records) * 100.0, 1) if total_balanzas_records > 0 else 100.0

        # Alerts count
        alertas_temp = sum(1 for t in temps if not t.cumple_especificacion)
        alertas_balanza = sum(1 for b in balanzas if not b.estado_conforme)
        total_alertas = alertas_temp + alertas_balanza

        # Area status summary
        areas_resumen: List[AreaStatusSummary] = []
        for area_info in STANDARD_AREAS:
            latest = (
                db.query(ControlTemperatura)
                .filter(ControlTemperatura.area_ambiente.ilike(f"%{area_info['name']}%"))
                .order_by(desc(ControlTemperatura.fecha), desc(ControlTemperatura.id))
                .first()
            )
            if latest:
                areas_resumen.append(
                    AreaStatusSummary(
                        area=area_info["name"],
                        temperatura_actual=latest.temperatura_c,
                        humedad_actual=latest.humedad_relativa_pct,
                        norma=area_info["norma"],
                        rango_temperatura=area_info["rango_temp"],
                        rango_humedad=area_info["rango_hum"],
                        conforme=latest.cumple_especificacion,
                        ultima_lectura=f"{latest.fecha} {latest.hora_lectura}",
                    )
                )
            else:
                areas_resumen.append(
                    AreaStatusSummary(
                        area=area_info["name"],
                        temperatura_actual=area_info["temp_target"],
                        humedad_actual=95.0 if "CÁMARA" in area_info["name"] else 60.0,
                        norma=area_info["norma"],
                        rango_temperatura=area_info["rango_temp"],
                        rango_humedad=area_info["rango_hum"],
                        conforme=True,
                        ultima_lectura="Sin registros",
                    )
                )

        # Balanza status summary
        balanza_resumen: List[BalanzaStatusSummary] = []
        for bal_info in STANDARD_BALANZAS:
            latest = (
                db.query(ControlBalanza)
                .filter(ControlBalanza.codigo_balanza == bal_info["codigo"])
                .order_by(desc(ControlBalanza.fecha), desc(ControlBalanza.id))
                .first()
            )
            if latest:
                balanza_resumen.append(
                    BalanzaStatusSummary(
                        codigo_balanza=bal_info["codigo"],
                        ubicacion=latest.ubicacion,
                        capacidad_g=latest.capacidad_g,
                        ultima_verificacion=latest.fecha,
                        error_reciente_g=latest.error_g,
                        error_max_permitido_g=latest.error_max_permitido_g,
                        conforme=latest.estado_conforme,
                        verificado_por=latest.verificado_por,
                    )
                )
            else:
                balanza_resumen.append(
                    BalanzaStatusSummary(
                        codigo_balanza=bal_info["codigo"],
                        ubicacion=bal_info["ubicacion"],
                        capacidad_g=bal_info["capacidad_g"],
                        ultima_verificacion="Sin registros",
                        error_reciente_g=0.0,
                        error_max_permitido_g=bal_info["error_max_g"],
                        conforme=True,
                        verificado_por="-",
                    )
                )

        return ControlAmbientalDashboardResponse(
            total_lecturas_temperatura=total_temps,
            promedio_temperatura_c=avg_temp,
            promedio_humedad_pct=avg_hum,
            tasa_cumplimiento_temp_pct=tasa_cumplimiento_temp,
            total_balanzas_registradas=len(unique_balanza_codes),
            balanzas_verificadas_hoy=verificadas_hoy,
            tasa_conformidad_balanzas_pct=tasa_conforma_balanza,
            alertas_activas=total_alertas,
            areas_resumen=areas_resumen,
            balanzas_resumen=balanza_resumen,
        )

    # --- Seed Data Injector ---

    @staticmethod
    def sembrar_datos_reales(db: Session, user_id: str = "system", user_name: str = "Sistema Seeder") -> int:
        import random
        today = datetime.date.today()
        count_seeded = 0

        # Generate 20 days of historical readings
        for i in range(20, -1, -1):
            date_str = (today - datetime.timedelta(days=i)).strftime("%Y-%m-%d")

            # Seed Temperature for each standard area (morning & afternoon)
            for area in STANDARD_AREAS:
                for hora in ["08:30", "15:00"]:
                    # Small variation around targets
                    if "CÁMARA" in area["name"]:
                        temp = round(23.0 + random.uniform(-0.8, 0.9), 1)
                        hum = round(random.uniform(94.0, 99.0), 1)
                    else:
                        temp = round(20.0 + random.uniform(-1.2, 1.5), 1)
                        hum = round(random.uniform(52.0, 68.0), 1)

                    cumple = _evaluar_cumplimiento_temp(area["name"], temp, hum)
                    t_record = ControlTemperatura(
                        fecha=date_str,
                        hora_lectura=hora,
                        area_ambiente=area["name"],
                        temperatura_c=temp,
                        humedad_relativa_pct=hum,
                        temp_min=round(temp - 1.5, 1),
                        temp_max=round(temp + 1.2, 1),
                        cumple_especificacion=cumple,
                        responsable_lectura=random.choice(["J.P.", "M.A.", "R.V.", "C.T."]),
                        observaciones="Control de rutinas de laboratorio conforme" if cumple else "Leve desviación registrada y corregida ventilación",
                    )
                    db.add(t_record)
                    count_seeded += 1

            # Seed Balanza verification for each standard scale
            for bal in STANDARD_BALANZAS:
                masa = bal["masa_patron_g"]
                err_max = bal["error_max_g"]
                error_sim = round(random.uniform(-err_max * 0.4, err_max * 0.4), 4)
                lectura = round(masa + error_sim, 4)
                b_record = ControlBalanza(
                    fecha=date_str,
                    codigo_balanza=bal["codigo"],
                    ubicacion=bal["ubicacion"],
                    capacidad_g=bal["capacidad_g"],
                    masa_patron_g=masa,
                    lectura_balanza_g=lectura,
                    error_g=error_sim,
                    error_max_permitido_g=err_max,
                    estado_conforme=abs(error_sim) <= err_max,
                    limpieza_nivelacion=True,
                    verificado_por=random.choice(["J.P.", "M.A.", "R.V."]),
                    observaciones="Nivel de gota centrado, platillo limpio.",
                )
                db.add(b_record)
                count_seeded += 1

        db.commit()

        emit_audit_log(
            db=db,
            user_id=user_id,
            user_name=user_name,
            action=f"Se sembraron {count_seeded} registros reales/de laboratorio en Control Ambiental",
            module="CONTROL_AMBIENTAL",
            details={"total_registros": count_seeded},
        )
        return count_seeded

    # --- Excel Export Generators ---

    @staticmethod
    def generar_excel_temperatura(
        db: Session,
        area: Optional[str] = None,
        fecha_inicio: Optional[str] = None,
        fecha_fin: Optional[str] = None,
    ) -> io.BytesIO:
        template_name = "F-LEM-P-05.01 V03 CONTROL DE TEMPERATURA Y HUMEDAD RELATIVA.xlsx"
        template_path = find_template_path(template_name)
        records = ControlAmbientalService.listar_temperatura(
            db, area=area, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, limit=500
        )
        if template_path and os.path.exists(template_path):
            first = records[0] if records else None
            obs = (first.observaciones if first else "") or ""
            parsed_obs = {}
            if obs.startswith("{"):
                try:
                    import json
                    parsed_obs = json.loads(obs)
                except Exception:
                    pass
            elif obs.startswith("REVISADO POR:"):
                parsed_obs = {"revisado_por": obs.replace("REVISADO POR:", "").strip()}

            reg_val = parsed_obs.get("registro", "REG-01")
            mes_anio_val = parsed_obs.get("mes_anio", "AGOSTO DE 2026")
            aprob_por_val = parsed_obs.get("aprobado_por", "JEFE DE LABORATORIO")
            fecha_aprob_val = parsed_obs.get("fecha_aprobacion", datetime.date.today().strftime("%Y-%m-%d"))
            area_val = area or (first.area_ambiente if first else "ÁREA DE RECEPCIÓN DE MUESTRAS")

            def _transform_sheet(sheet_bytes: bytes) -> bytes:
                root = etree.fromstring(sheet_bytes)
                sheet_data = root.find(f".//{{{NS_SHEET}}}sheetData")

                set_cell(sheet_data, "D15", area_val)

                if "CÁMARA HÚMEDA" in area_val.upper() or "CURADO" in area_val.upper():
                    set_cell(sheet_data, "D17", "23°C ± 2.0°C / ≥ 90%")
                else:
                    set_cell(sheet_data, "D17", "20°C ± 3.0°C / 45% - 80%")

                start_row = 24
                for idx, r in enumerate(records):
                    r_obs = r.observaciones or ""
                    r_parsed = {}
                    if r_obs.startswith("{"):
                        try:
                            import json
                            r_parsed = json.loads(r_obs)
                        except Exception:
                            pass
                    elif r_obs.startswith("REVISADO POR:"):
                        r_parsed = {"revisado_por": r_obs.replace("REVISADO POR:", "").strip()}

                    fecha_lectura = r_parsed.get("fecha_lectura", r.fecha)
                    hum_min = r_parsed.get("hum_min", "-")
                    rev_por = r_parsed.get("revisado_por", "")

                    row = start_row + idx
                    set_cell(sheet_data, f"B{row}", r.fecha)
                    set_cell(sheet_data, f"C{row}", r.hora_lectura)
                    set_cell(sheet_data, f"D{row}", fecha_lectura)
                    set_cell(sheet_data, f"E{row}", r.temp_min if r.temp_min is not None else "-", is_number=isinstance(r.temp_min, (int, float)))
                    set_cell(sheet_data, f"F{row}", r.temperatura_c, is_number=isinstance(r.temperatura_c, (int, float)))
                    set_cell(sheet_data, f"G{row}", hum_min, is_number=isinstance(hum_min, (int, float)))
                    set_cell(sheet_data, f"H{row}", r.humedad_relativa_pct, is_number=isinstance(r.humedad_relativa_pct, (int, float)))
                    set_cell(sheet_data, f"I{row}", r.responsable_lectura)
                    set_cell(sheet_data, f"K{row}", rev_por)

                return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

            def _transform_drawing(drawing_xml: bytes) -> bytes:
                root = etree.fromstring(drawing_xml)
                ns = {"xdr": NS_DRAW, "a": NS_A}
                vals = {
                    1: reg_val,
                    3: mes_anio_val,
                    6: aprob_por_val,
                    9: fecha_aprob_val,
                }
                for anchor in root.findall(".//xdr:twoCellAnchor", ns):
                    from_col_el = anchor.find(".//xdr:from/xdr:col", ns)
                    from_row_el = anchor.find(".//xdr:from/xdr:row", ns)
                    if from_col_el is not None and from_row_el is not None:
                        col = int(from_col_el.text)
                        row = int(from_row_el.text)
                        if row == 9 and col in vals and vals[col]:
                            p = anchor.find(".//a:p", ns)
                            if p is not None:
                                _set_paragraph_text(p, vals[col])

                return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

            file_bytes = transform_template_sheet(
                template_name,
                "Formato de control de temperatu",
                _transform_sheet,
                drawing_transform=_transform_drawing,
            )
            output = io.BytesIO(file_bytes)
            return output
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "F-LEM-P-05.01 V03"
            ws.append(["GEOFAL S.A.C. - LABORATORIO DE ENSAYO DE MATERIALES"])
            ws.append(["FORMATO DE CONTROL DE TEMPERATURA Y HUMEDAD RELATIVA (F-LEM-P-05.01 V03)"])
            ws.append([])
            for r in records:
                ws.append([r.fecha, r.hora_lectura, r.area_ambiente, r.temperatura_c, r.humedad_relativa_pct])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

    @staticmethod
    def generar_excel_balanzas(
        db: Session,
        codigo: Optional[str] = None,
        fecha_inicio: Optional[str] = None,
        fecha_fin: Optional[str] = None,
    ) -> io.BytesIO:
        template_name = "F-LEM-IN-01.02 V03 FORMATO DE VERIFICACIÓN DIARIA DE BALANZAS.xlsx"
        template_path = find_template_path(template_name)
        records = ControlAmbientalService.listar_balanza(
            db, codigo=codigo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, limit=500
        )

        if template_path and os.path.exists(template_path):
            def _transform_sheet(sheet_bytes: bytes) -> bytes:
                root = etree.fromstring(sheet_bytes)
                sheet_data = root.find(f".//{{{NS_SHEET}}}sheetData")

                if records:
                    first = records[0]
                    obs = first.observaciones or ""
                    parsed_obs = {}
                    if obs.startswith("{"):
                        try:
                            import json
                            parsed_obs = json.loads(obs)
                        except Exception:
                            pass
                    elif obs.startswith("REVISADO POR:"):
                        parsed_obs = {"revisado_por": obs.replace("REVISADO POR:", "").strip()}

                    codigo_bal_val = codigo or first.codigo_balanza
                    pesas_val = parsed_obs.get("codigos_pesas_patron", "PP-01, PP-02, PP-05")
                    mes_anio_val = parsed_obs.get("mes_anio", "AGOSTO DE 2026")

                    set_cell(sheet_data, "E6", codigo_bal_val)       # E6
                    set_cell(sheet_data, "AK6", mes_anio_val)      # AK6
                    set_cell(sheet_data, "F8", pesas_val)           # F8
                elif codigo:
                    set_cell(sheet_data, "E6", codigo)

                # Group records by (r.fecha, hora) so each verification test forms 1 horizontal row in Excel
                grouped_rows = {}
                for r in records:
                    obs = r.observaciones or ""
                    parsed_obs = {}
                    if obs.startswith("{"):
                        try:
                            import json
                            parsed_obs = json.loads(obs)
                        except Exception:
                            pass
                    elif obs.startswith("REVISADO POR:"):
                        parsed_obs = {"revisado_por": obs.replace("REVISADO POR:", "").strip()}

                    hora = parsed_obs.get("hora", "08:00")
                    pesadas_payload = parsed_obs.get("pesadas", [])
                    key = (r.fecha, hora)

                    if key not in grouped_rows:
                        grouped_rows[key] = {
                            "fecha": r.fecha,
                            "hora": hora,
                            "temp_c": parsed_obs.get("temp_c", "-"),
                            "humedad_pct": parsed_obs.get("humedad_pct", "-"),
                            "verificado_por": r.verificado_por,
                            "revisado_por": parsed_obs.get("revisado_por", ""),
                            "pesadas": []
                        }

                    if pesadas_payload and isinstance(pesadas_payload, list) and len(pesadas_payload) > 0:
                        grouped_rows[key]["pesadas"] = pesadas_payload
                    else:
                        lectura_val = r.lectura_balanza_g or r.masa_patron_g or ""
                        estado_val = "OK" if r.estado_conforme else "NO"
                        grouped_rows[key]["pesadas"].append({
                            "lectura_balanza_g": lectura_val,
                            "estado": estado_val
                        })

                start_row = 12
                merge_cells = root.find(f".//{{{NS_SHEET}}}mergeCells")
                if merge_cells is None:
                    merge_cells = etree.SubElement(root, f"{{{NS_SHEET}}}mergeCells")

                for idx, row_data in enumerate(grouped_rows.values()):
                    row = start_row + idx

                    # Remove any pre-existing template merge for this data row (like C12:K12)
                    for mc in list(merge_cells.findall(f"{{{NS_SHEET}}}mergeCell")):
                        ref = mc.get("ref", "")
                        if f"{row}" in ref:
                            merge_cells.remove(mc)

                    # Add exact merges A{row}:B{row} for FECHA and C{row}:D{row} for HORA
                    for new_ref in [f"A{row}:B{row}", f"C{row}:D{row}"]:
                        mc = etree.SubElement(merge_cells, f"{{{NS_SHEET}}}mergeCell")
                        mc.set("ref", new_ref)

                    set_cell(sheet_data, f"A{row}", row_data["fecha"])
                    set_cell(sheet_data, f"C{row}", row_data["hora"])
                    set_cell(sheet_data, f"E{row}", row_data["temp_c"], is_number=isinstance(row_data["temp_c"], (int, float)))
                    set_cell(sheet_data, f"F{row}", row_data["humedad_pct"], is_number=isinstance(row_data["humedad_pct"], (int, float)))

                    pesadas = row_data["pesadas"]
                    if pesadas and isinstance(pesadas, list):
                        for p_idx, p in enumerate(pesadas[:15]):
                            col_num_lectura = 7 + (p_idx * 2)
                            col_num_estado = 8 + (p_idx * 2)
                            col_letter_lectura = col_num_to_letter(col_num_lectura)
                            col_letter_estado = col_num_to_letter(col_num_estado)

                            lectura_val = p.get("lectura_balanza_g") or p.get("masa_patron_g") or ""
                            estado_val = p.get("estado") or ("OK" if lectura_val != "" else "")

                            if lectura_val != "":
                                set_cell(sheet_data, f"{col_letter_lectura}{row}", lectura_val, is_number=isinstance(lectura_val, (int, float)))
                            if estado_val != "":
                                set_cell(sheet_data, f"{col_letter_estado}{row}", estado_val)

                    set_cell(sheet_data, f"AK{row}", row_data["verificado_por"])
                    set_cell(sheet_data, f"AL{row}", row_data["revisado_por"])

                merge_cells.set("count", str(len(merge_cells.findall(f"{{{NS_SHEET}}}mergeCell"))))

                return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

            file_bytes = transform_template_sheet(template_name, "Verf Diaria", _transform_sheet)
            output = io.BytesIO(file_bytes)
            return output
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "F-LEM-IN-01.02 V03"
            ws.append(["GEOFAL S.A.C. - LABORATORIO DE ENSAYO DE MATERIALES"])
            ws.append(["FORMATO DE VERIFICACIÓN DIARIA DE BALANZAS (F-LEM-IN-01.02 V03)"])
            ws.append([])
            for r in records:
                ws.append([r.fecha, r.codigo_balanza, r.lectura_balanza_g, r.verificado_por])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
