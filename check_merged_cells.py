import openpyxl

template_path = r"C:\Users\Lenovo\Documents\crmnew\api-geofal-crm\app\templates\informes\Densidad Huantar\1-INF.-N-001-26-SU06-DEN-V05.xlsx"
wb = openpyxl.load_workbook(template_path)

for name in ["F LEM1", "F LEM2", "Datos"]:
    ws = wb[name]
    print(f"=== MERGED CELLS IN {name} ===")
    ranges = list(ws.merged_cells.ranges)
    # Print ranges that intersect with rows 27 to 45
    for r in ranges:
        min_col, min_row, max_col, max_row = r.bounds
        if min_row >= 25 and max_row <= 45:
            print(f"  Range: {r.coord} (Rows: {min_row}-{max_row}, Cols: {openpyxl.utils.get_column_letter(min_col)}-{openpyxl.utils.get_column_letter(max_col)})")
