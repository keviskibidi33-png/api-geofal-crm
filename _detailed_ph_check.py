from openpyxl import load_workbook

wb = load_workbook('app/templates/Template_PH.xlsx')
ws = wb.worksheets[1]  # PH sheet

print("=== DETAILED CELL INSPECTION ===\n")

print("ROW 17 (SECADO AL AIRE):")
for col in range(1, 10):  # A-I
    cell = ws.cell(17, col)
    if cell.value or ws.cell(17, col).coordinate in [c.coord for mc in ws.merged_cells for c in mc.cells]:
        print(f"  Col {cell.column_letter}: '{cell.value}'")

print("\nROW 18 (SECADO HORNO):")
for col in range(1, 10):
    cell = ws.cell(18, col)
    if cell.value:
        print(f"  Col {cell.column_letter}: '{cell.value}'")

print("\nROW 24 (TEMPERATURA):")
for col in range(1, 10):
    cell = ws.cell(24, col)
    if cell.value:
        print(f"  Col {cell.column_letter}: '{cell.value}'")

print("\nROW 25 (PH):")
for col in range(1, 10):
    cell = ws.cell(25, col)
    if cell.value:
        print(f"  Col {cell.column_letter}: '{cell.value}'")

print("\n=== MERGED CELLS DETAILED ===")
for mc in ws.merged_cells:
    min_row, min_col, max_row, max_col = mc.bounds
    if min_row >= 17 and max_row <= 25:
        start_cell = ws.cell(min_row, min_col)
        end_cell = ws.cell(max_row, max_col)
        print(f"{start_cell.coordinate}:{end_cell.coordinate} - Contains: '{start_cell.value}'")
